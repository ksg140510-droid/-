"""멤브레인 홀막힘 카운팅 프로그램 v5.3"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import cv2
import numpy as np
from PIL import Image, ImageTk, ImageDraw, ImageFont
import threading
import time
import datetime
import openpyxl
import json
import os
import winsound
import base64
import io
import math as _math
import ctypes
import ctypes.wintypes
import re
import shutil
import sqlite3
import subprocess
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference


class _DNX64:
    """DNX64 SDK(디노라이트 제조사 공식 제어 DLL) 지연 로딩 싱글턴.
    SDK가 없는 PC에서도 앱이 정상 동작해야 하므로 실패하면 조용히 None."""
    _dll = None
    _tried = False
    _idx = 0
    # LED 유지용 백그라운드 스레드와 메인 스레드(노출고정/화질고정 등)가
    # 동시에 DLL을 호출하면 장비가 명령을 잘못 알아듣거나 먹통이 될 수 있어
    # 모든 DNX64 호출은 이 락을 잡고 나서 하도록 한다.
    lock = threading.Lock()

    @classmethod
    def get(cls):
        if not cls._tried:
            cls._tried = True
            try:
                path = r'C:\Program Files\DNX64\DNX64.dll'
                if not os.path.exists(path):
                    return None
                dll = ctypes.CDLL(path)
                dll.Init.restype = ctypes.c_bool
                dll.GetVideoDeviceCount.restype = ctypes.c_int
                dll.SetVideoDeviceIndex.argtypes = [ctypes.c_int]
                dll.GetVideoDeviceIndex.restype = ctypes.c_int
                dll.GetAutoExposure.argtypes = [ctypes.c_int]
                dll.GetAutoExposure.restype = ctypes.c_long
                dll.SetAutoExposure.argtypes = [ctypes.c_int, ctypes.c_long]
                dll.SetAutoExposure.restype = None
                dll.GetExposureValue.argtypes = [ctypes.c_int]
                dll.GetExposureValue.restype = ctypes.c_long
                dll.SetExposureValue.argtypes = [ctypes.c_int, ctypes.c_long]
                dll.SetExposureValue.restype = None
                dll.SetLEDState.argtypes = [ctypes.c_int, ctypes.c_int]
                dll.SetLEDState.restype = None
                dll.GetAETarget.argtypes = [ctypes.c_int]
                dll.GetAETarget.restype = ctypes.c_long
                dll.SetAETarget.argtypes = [ctypes.c_int, ctypes.c_long]
                dll.SetAETarget.restype = None
                dll.GetVideoProcAmp.argtypes = [ctypes.c_int]
                dll.GetVideoProcAmp.restype = ctypes.c_long
                dll.SetVideoProcAmp.argtypes = [ctypes.c_int, ctypes.c_long]
                dll.SetVideoProcAmp.restype = None
                dll.StopMonitoring.restype = None
                dll.Init()
                time.sleep(0.3)
                if dll.GetVideoDeviceCount() <= 0:
                    return None
                dll.SetVideoDeviceIndex(0)
                time.sleep(0.1)
                cls._idx = dll.GetVideoDeviceIndex()
                cls._dll = dll
            except Exception:
                cls._dll = None
        return cls._dll


LOGO_PATH = os.path.join(os.path.expanduser('~'), 'Desktop', '1.jfif')

MODELS = {
    '120H': 8200, '120L': 7700,
    '140H': 9100, '140L': 9600,
    '160H': 10600, '160L': 10500,
    '180H': 11800, '180L': 12000,
    '200H': 13200, '200L': 13100,
}
DEFECT_LIMIT   = 4.0
DEFECT_REASONS = ['오염', '파손', '이물질', '가공불량', '기타']
SCALE_MARGIN   = 28    # 스케일 바 좌하단 여백
SCALE_SNAP_R   = 20    # 캘리브레이션 클릭 스냅 반경(px)
SCALE_TARGET_PX = 180  # 스케일 바 목표 픽셀 폭 (nice값 계산 기준)
# 바탕화면 기준 경로 사용 — exe(onefile) 실행 시 __file__ 이 매번 삭제되는 임시
# 압축해제 폴더를 가리켜 캘리브레이션이 재시작마다 초기화되는 문제 방지
_CAL_FILE      = os.path.join(os.path.expanduser('~'), 'Desktop', '홀막힘_캘리브레이션.json')

def _nice_scale(px_per_mm, target_px=SCALE_TARGET_PX):
    """px_per_mm 와 목표 픽셀에 맞는 (bar_px, bar_mm) 반환."""
    if px_per_mm <= 0:
        return 180, None
    target_mm = target_px / px_per_mm
    for v in (0.01, 0.02, 0.05, 0.1, 0.2, 0.5, 1.0, 2.0, 5.0, 10.0):
        if v >= target_mm * 0.5:
            return max(int(v * px_per_mm), 20), v
    return max(int(10.0 * px_per_mm), 20), 10.0

def _xl_safe(v):
    """Excel 수식 인젝션 방지: =,+,-,@ 로 시작하는 사용자 입력값 앞에 인용부호 추가"""
    s = str(v)
    if s and s[0] in ('=', '+', '-', '@'):
        return "'" + s
    return s

_INVALID_FS_CHARS = re.compile(r'[\\/:*?"<>|\x00-\x1f]')

def _safe_filename(name, fallback='UNKNOWN'):
    """LOT/일련번호 등 자유입력값을 폴더·파일명으로 쓸 때 경로 조작(..)과
    Windows 금지문자를 제거해 안전한 이름으로 변환."""
    s = str(name).replace('..', '_')
    s = _INVALID_FS_CHARS.sub('_', s).strip().strip('.').strip()
    return s if s else fallback

MODEL_CLR = {
    'H': {'normal': '#7b1a1a', 'select': '#e53935', 'text': '#ffffff'},
    'L': {'normal': '#0d2d5e', 'select': '#1976d2', 'text': '#ffffff'},
}
CAM_W, CAM_H   = 1280, 720
PREVIEW_W, PREVIEW_H = 800, 560
FPS_LIMIT = 30
# ── HMI 색상 팔레트 ─────────────────────────────────────────────────────────
# 다크 모드 (기본)
BG       = '#0a0e14'
CARD_BG  = '#141b24'
PANEL_BG = '#1a2332'
HDR_BG   = '#0d1e35'
BORDER   = '#253448'
ACC_RED  = '#e53935'
ACC_GRN  = '#00c853'
ACC_BLU  = '#1e90ff'
ACC_YEL  = '#ffb300'
TXT_W    = '#d8e4f0'
TXT_G    = '#6b84a0'
TXT_DIM  = '#3d5470'

# 라이트 모드 대응색
LT_BG      = '#eef2f7'
LT_CARD    = '#dde5f0'
LT_PANEL   = '#ccd8ea'
LT_HDR     = '#1a4a8a'
LT_BORDER  = '#9ab0cc'
LT_TXT     = '#0d1e2e'
LT_TXT_G   = '#3a5570'

# 컷팅 캡처(③) 패널 전용 다크/라이트 색상표 — 이 패널은 전역 자동 리테마
# (_D2L/_L2D 색상표 치환)에서 제외하고 _apply_cut_capture_theme()이 직접 칠함.
# (여러 다크색이 같은 라이트색으로 겹쳐 매핑되는 전역 표의 모호성을 피하고,
# 라이트 모드에서도 100% 대비가 보장되는 값을 명시적으로 지정하기 위함)
CUT_CAPTURE_COLORS = {
    'dark': dict(
        panel_bg='#1a1a2e', border='#5b6bc0',
        title_fg='#8c9eff', subtitle_fg='#9aa8d4', label_fg='#ffffff',
        combo_bg='#2d3542', combo_fg='#ffffff', combo_border='#8c9eff',
        combo_insert='#ffd54a',
        count_unsel_bg='#2d3542', count_unsel_fg='#ffffff',
        count_sel_bg='#3949ab', count_sel_fg='#ffffff',
        entry_bg='#2d3542', entry_fg='#ffffff', entry_insert='#ffd54a',
        apply_bg='#2a3540', apply_fg='#ffffff',
        status_fg='#ffd54a', capture_bg='#3949ab', capture_fg='#ffffff',
    ),
    'light': dict(
        panel_bg='#e3e6fa', border='#5b6bc0',
        title_fg='#33409e', subtitle_fg='#4a5590', label_fg='#10192e',
        combo_bg='#ffffff', combo_fg='#10192e', combo_border='#5b6bc0',
        combo_insert='#10192e',
        count_unsel_bg='#c6ccef', count_unsel_fg='#10192e',
        count_sel_bg='#3949ab', count_sel_fg='#ffffff',
        entry_bg='#ffffff', entry_fg='#10192e', entry_insert='#10192e',
        apply_bg='#c6ccef', apply_fg='#10192e',
        status_fg='#8a5a00', capture_bg='#3949ab', capture_fg='#ffffff',
    ),
}

HISTORY_FILE = os.path.join(os.path.expanduser('~'), 'Desktop', '홀막힘_이력.json')
HISTORY_DB_FILE = os.path.join(os.path.expanduser('~'), 'Desktop', '홀막힘_이력.db')
CAPTURE_DIR  = os.path.join(os.path.expanduser('~'), 'Desktop', '홀막힘_캡처')
CUTTING_CAPTURE_DIR = os.path.join(os.path.expanduser('~'), 'Desktop', '칼날컷팅')
CUTTING_RECORD_FILE = os.path.join(os.path.expanduser('~'), 'Desktop', '칼날컷팅_기록.json')
CONFIG_FILE  = os.path.join(os.path.expanduser('~'), 'Desktop', '홀막힘_설정.json')
HISTORY_BACKUP_DIR = os.path.join(os.path.expanduser('~'), 'Desktop', '홀막힘_이력_백업')
HISTORY_BACKUP_KEEP = 50   # 보관할 최대 백업 개수 (초과 시 오래된 것부터 자동 삭제)

_HISTORY_FIELDS = ('date', 'time', 'lot', 'operator', 'serial', 'model',
                    'fibers', 'count', 'rate', 'verdict', 'defect_reason', 'deviation')


def _backup_history_file(reason='manual'):
    """이력 JSON을 타임스탬프 파일명으로 백업 폴더에 복사 (기존 백업을 덮어쓰지 않음).
    반환값: 성공 시 백업 파일 경로, 대상 파일이 없거나 실패 시 None."""
    if not os.path.exists(HISTORY_FILE):
        return None
    try:
        os.makedirs(HISTORY_BACKUP_DIR, exist_ok=True)
        ts  = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        dst = os.path.join(HISTORY_BACKUP_DIR, f'홀막힘_이력_{ts}_{reason}.json')
        shutil.copy2(HISTORY_FILE, dst)
        backups = sorted(
            f for f in os.listdir(HISTORY_BACKUP_DIR)
            if f.startswith('홀막힘_이력_') and f.endswith('.json'))
        while len(backups) > HISTORY_BACKUP_KEEP:
            oldest = backups.pop(0)
            try:
                os.remove(os.path.join(HISTORY_BACKUP_DIR, oldest))
            except Exception:
                pass
        return dst
    except Exception:
        return None


def _backup_history_db(reason='manual'):
    """이력 SQLite DB를 타임스탬프 파일명으로 백업 폴더에 복사.
    반환값: 성공 시 백업 파일 경로, 대상 파일이 없거나 실패 시 None."""
    if not os.path.exists(HISTORY_DB_FILE):
        return None
    try:
        os.makedirs(HISTORY_BACKUP_DIR, exist_ok=True)
        ts  = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        dst = os.path.join(HISTORY_BACKUP_DIR, f'홀막힘_이력_{ts}_{reason}.db')
        shutil.copy2(HISTORY_DB_FILE, dst)
        backups = sorted(
            f for f in os.listdir(HISTORY_BACKUP_DIR)
            if f.startswith('홀막힘_이력_') and f.endswith('.db'))
        while len(backups) > HISTORY_BACKUP_KEEP:
            oldest = backups.pop(0)
            try:
                os.remove(os.path.join(HISTORY_BACKUP_DIR, oldest))
            except Exception:
                pass
        return dst
    except Exception:
        return None


# ── 이력 저장소 (SQLite) ─────────────────────────────────────────────────────
# v5.3까지는 홀막힘_이력.json 전체를 매번 통째로 읽고 써서, 이력이 쌓일수록
# 느려지고 파일이 깨질 위험도 커졌다. v5.4부터는 SQLite(홀막힘_이력.db)를
# 실제 저장소로 쓰고, 기존 JSON은 최초 1회 자동으로 옮겨 담은 뒤 원본 그대로
# 보존한다 (삭제하지 않음 — 문제 생기면 언제든 대조/복구 가능).

def _db_connect():
    conn = sqlite3.connect(HISTORY_DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def _db_init_schema():
    conn = _db_connect()
    try:
        conn.execute('''CREATE TABLE IF NOT EXISTS inspections (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            date          TEXT NOT NULL DEFAULT '',
            time          TEXT NOT NULL DEFAULT '',
            lot           TEXT NOT NULL DEFAULT '',
            operator      TEXT NOT NULL DEFAULT '',
            serial        TEXT NOT NULL DEFAULT '',
            model         TEXT NOT NULL DEFAULT '',
            fibers        INTEGER,
            count         INTEGER,
            rate          REAL,
            verdict       TEXT NOT NULL DEFAULT '',
            defect_reason TEXT NOT NULL DEFAULT '',
            deviation     REAL
        )''')
        for col in ('date', 'lot', 'operator', 'model', 'serial', 'verdict'):
            conn.execute(f'CREATE INDEX IF NOT EXISTS idx_insp_{col} ON inspections({col})')
        conn.execute('''CREATE TABLE IF NOT EXISTS _migration_meta (
            key TEXT PRIMARY KEY, value TEXT)''')
        conn.commit()
    finally:
        conn.close()


def _db_row_to_dict(row):
    d = {k: row[k] for k in _HISTORY_FIELDS}
    d['id'] = row['id']
    return d


def _db_insert(entry):
    """이력 1건 삽입. 반환값: 새로 생성된 행 id."""
    conn = _db_connect()
    try:
        cur = conn.execute(
            'INSERT INTO inspections '
            '(date,time,lot,operator,serial,model,fibers,count,rate,verdict,defect_reason,deviation) '
            'VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
            (entry.get('date', ''), entry.get('time', ''), entry.get('lot', ''),
             entry.get('operator', ''), entry.get('serial', ''), entry.get('model', ''),
             entry.get('fibers'), entry.get('count'), entry.get('rate'),
             entry.get('verdict', ''), entry.get('defect_reason', ''), entry.get('deviation')))
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def _db_delete_by_id(row_id):
    conn = _db_connect()
    try:
        conn.execute('DELETE FROM inspections WHERE id=?', (row_id,))
        conn.commit()
    finally:
        conn.close()


def _db_delete_by_serial_time(serial, time_str):
    """반환값: 실제로 삭제된 행이 있으면 True."""
    conn = _db_connect()
    try:
        cur = conn.execute('DELETE FROM inspections WHERE serial=? AND time=?',
                            (serial, time_str))
        conn.commit()
        return cur.rowcount > 0
    finally:
        conn.close()


def _db_delete_all():
    conn = _db_connect()
    try:
        conn.execute('DELETE FROM inspections')
        conn.commit()
    finally:
        conn.close()


def _db_query_all(order='id DESC'):
    """필터 없이 전체 이력을 dict 리스트로 반환 (order: 'id DESC'=최신순)."""
    conn = _db_connect()
    try:
        cur = conn.execute(f'SELECT * FROM inspections ORDER BY {order}')
        return [_db_row_to_dict(r) for r in cur.fetchall()]
    finally:
        conn.close()


def _db_count_all():
    conn = _db_connect()
    try:
        return conn.execute('SELECT COUNT(*) FROM inspections').fetchone()[0]
    finally:
        conn.close()


def _migrate_json_to_db_if_needed():
    """앱 시작 시 1회 호출. 기존 홀막힘_이력.json을 SQLite로 옮겨 담는다.
    - 이미 이전 완료된 경우 즉시 반환 (재실행해도 중복 삽입되지 않음)
    - 원본 JSON 파일은 절대 수정/삭제하지 않고 그대로 둔다
    - 옮기기 전 타임스탬프 백업을 별도로 하나 더 남긴다
    - 옮긴 건수와 원본 건수가 정확히 일치할 때만 '이전 완료'로 표시한다
      (불일치·오류 시에는 표시하지 않아 다음 실행 때 다시 시도한다)"""
    _db_init_schema()

    conn = _db_connect()
    try:
        row = conn.execute(
            "SELECT value FROM _migration_meta WHERE key='migrated'").fetchone()
        already_migrated = bool(row and row[0] == '1')
    finally:
        conn.close()
    if already_migrated:
        return

    if not os.path.exists(HISTORY_FILE):
        conn = _db_connect()
        try:
            conn.execute(
                "INSERT OR REPLACE INTO _migration_meta (key,value) VALUES ('migrated','1')")
            conn.execute(
                "INSERT OR REPLACE INTO _migration_meta (key,value) VALUES ('source_row_count','0')")
            conn.commit()
        finally:
            conn.close()
        return

    records = None
    try:
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            loaded = json.load(f)
        if isinstance(loaded, list):
            records = loaded
    except Exception:
        records = None

    if records is None:
        try:
            messagebox.showwarning(
                '이력 데이터베이스 준비 실패',
                f'{HISTORY_FILE}\n'
                '파일을 읽을 수 없어 이력 데이터베이스로 옮기지 못했습니다.\n'
                f'백업 폴더({HISTORY_BACKUP_DIR})에서 최근 백업을 확인해주세요.\n\n'
                '프로그램은 계속 실행되며, 다음 실행 시 다시 시도합니다.')
        except Exception:
            pass
        return

    try:
        _backup_history_file(reason='migration')
    except Exception:
        pass

    conn = _db_connect()
    try:
        conn.execute('BEGIN')
        for h in records:
            conn.execute(
                'INSERT INTO inspections '
                '(date,time,lot,operator,serial,model,fibers,count,rate,verdict,defect_reason,deviation) '
                'VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                (h.get('date', ''), h.get('time', ''), h.get('lot', ''), h.get('operator', ''),
                 h.get('serial', ''), h.get('model', ''), h.get('fibers'), h.get('count'),
                 h.get('rate'), h.get('verdict', ''), h.get('defect_reason', ''),
                 h.get('deviation')))
        db_count = conn.execute('SELECT COUNT(*) FROM inspections').fetchone()[0]
        if db_count == len(records):
            conn.execute(
                "INSERT OR REPLACE INTO _migration_meta (key,value) VALUES ('migrated','1')")
            conn.execute(
                "INSERT OR REPLACE INTO _migration_meta (key,value) VALUES ('source_row_count',?)",
                (str(len(records)),))
            conn.commit()
        else:
            conn.rollback()
    except Exception:
        conn.rollback()
    finally:
        conn.close()


# 다크 → 라이트 색상 매핑 (전체 커버리지)
_D2L = {
    # ── 배경 ────────────────────────────────────────────────────────────────
    '#0a0e14': LT_BG,       '#141b24': LT_CARD,
    '#1a2332': LT_PANEL,    '#0d1e35': LT_HDR,
    '#253448': LT_BORDER,
    # ── 텍스트 ──────────────────────────────────────────────────────────────
    '#d8e4f0': LT_TXT,      '#6b84a0': LT_TXT_G,
    '#3d5470': '#6080a0',   '#aaccee': '#dceeff',
    '#e6edf3': '#1a2a3a',   '#8b949e': '#4a5a6a',
    '#ffffff':  LT_TXT,     '#111111': '#ffffff',
    # ── 기존 배경 계열 ──────────────────────────────────────────────────────
    '#0d1117': LT_BG,       '#1c2128': LT_CARD,
    '#111820': '#ccd6e4',   '#21262d': '#cdd8e4',
    '#30363d': '#aabbcc',   '#2a2a2a': '#dde5ef',
    '#c8ddf0': '#1a3a6b',
    # ── 헤더/버튼 파랑 계열 ─────────────────────────────────────────────────
    '#1a3a6b': '#1565c0',   '#2c5f9e': '#0d47a1',
    '#0d2d5e': '#0d47a1',   '#1a3a5c': '#0d3060',
    '#1976d2': '#1255a0',   '#5b9bd5': '#1565c0',
    '#4a9fd4': '#1565c0',
    # ── 기능 버튼 색상 ──────────────────────────────────────────────────────
    '#1a5c2e': '#1b5e20',   '#0a2010': '#c8e6c9',
    '#7a5c00': '#e65100',   '#7b1a1a': '#ffcdd2',
    '#5c1a1a': '#ef9a9a',
    # ── 강조색 ──────────────────────────────────────────────────────────────
    '#e53935': '#c62828',   '#00c853': '#1b5e20',
    '#1e90ff': '#1255a0',   '#ffb300': '#e65100',
    '#f5c518': '#b36a00',   '#2e8b57': '#1b5e20',
    # ── 태그/상태 ────────────────────────────────────────────────────────────
    '#ff8080': '#b00020',   '#ffc107': '#b36a00',
    '#aaffaa': '#1b5e20',   '#ff6b6b': '#c62828',
    '#cc0000': '#c62828',   '#dde3ea': '#b0bec8',
    # ── 다크 모드 전용 추가 ──────────────────────────────────────────────────
    '#3d1010': '#ffe5e5',
    # ── 이탈거리 측정 패널 배경 (누락 시 라이트모드에서 저대비 발생) ──────────
    '#151008': '#fff3e0',   '#0a1a0a': '#e8f5e9',
    '#0d2214': '#e8f5e9',
}
_L2D = {v: k for k, v in _D2L.items()}

_CURRENT_THEME = 'dark'     # 모듈 전역 테마 상태

# ── HMI Flat 아이콘 시스템 (PIL 기반, 24~28px) ────────────────────────────────

# 기능별 고유 아이콘 색상: (다크 RGB, 라이트 RGB)
_ICO_CLR = {
    'camera':     ((30,  150, 255),   (0,   90, 210)),   # 파랑  — 화면캡처
    'disk':       ((0,   210, 100),   (0,  140,  55)),   # 초록  — 저장
    'undo':       ((255, 140,  0),    (200,  90,   0)),   # 주황  — 취소
    'reset':      ((150, 175, 200),   (70,  100, 140)),   # 회청  — 초기화
    'xmark':      ((230,  55,  55),   (185,  25,  25)),   # 빨강  — 삭제/NG
    'check':      ((0,   220,  85),   (0,  160,  55)),   # 밝초록 — OK/양호
    'folder':     ((0,   195, 215),   (0,  130, 165)),   # 청록  — 이력조회
    'excel':      ((0,   185,  80),   (0,  130,  50)),   # 엑셀초록
    'printer':    ((100, 165, 230),   (40,  100, 190)),   # 스틸블루
    'next':       ((40,  165, 255),   (0,  105, 210)),   # 파랑  — 다음
    'graph':      ((165, 100, 245),   (110,  50, 200)),   # 보라  — 통계
    'gear':       ((140, 175, 210),   (70,  110, 160)),   # 스틸  — 설정
    'bell':       ((255, 205,  40),   (185, 135,   0)),   # 노랑  — 알람
    'bell_off':   ((120, 130, 140),   ( 90, 100, 115)),   # 회색  — 알람끔
    'sun':        ((255, 215,  55),   (200, 150,   0)),   # 황금  — 라이트모드
    'moon':       ((155, 195, 255),   ( 70, 115, 220)),   # 연청  — 다크모드
    'home':       ((200, 225, 255),   ( 55, 100, 180)),   # 청백  — 홈
    'fullscreen': ((185, 215, 245),   ( 55,  95, 165)),   # 청백  — 전체화면
    'windowed':   ((185, 215, 245),   ( 55,  95, 165)),   # 청백  — 창모드
    'grid':       ((0,   205, 195),   (0,  145, 140)),   # 청록  — 가이드
    'zoom_plus':  ((80,  225, 135),   (0,  165,  70)),   # 밝초록 — 확대
    'zoom_minus': ((255, 165,  40),   (195, 110,   0)),   # 주황  — 축소
    'select':     ((185, 205, 230),   ( 75, 110, 155)),   # 연청  — 영역선택
}

_ICON_CACHE: dict = {}


def _make_icon(name: str, size: int = 26, fg=(216, 228, 240)) -> Image.Image:
    """산업용 HMI Flat-design 아이콘을 PIL로 생성"""
    img = Image.new('RGBA', (size, size), (0, 0, 0, 0))
    d   = ImageDraw.Draw(img)
    c   = (*fg, 255)
    dim = (*tuple(max(0, v - 80) for v in fg), 200)
    S, cx, cy = size, size // 2, size // 2
    m  = max(2, S // 13)
    lw = max(2, S // 13)

    try:
        if name == 'home':
            d.polygon([(cx, m), (m, cy + 2), (S - m, cy + 2)], fill=c)
            d.rectangle([m + 4, cy + 2, S - m - 4, S - m], fill=c)
            dw = max(2, S // 8)
            d.rectangle([cx - dw, S * 3 // 5, cx + dw, S - m], fill=(0, 0, 0, 0))

        elif name == 'folder':
            d.rounded_rectangle([m, S // 3, S - m, S - m], radius=2, fill=c)
            d.rounded_rectangle([m, S // 3 - 2, S // 3 + 4, S // 2 + 1], radius=2, fill=c)

        elif name == 'graph':
            d.line([m, S - m, S - m, S - m], fill=c, width=lw)
            bw = max(3, (S - 2 * m - 6) // 3)
            for i, h in enumerate([S // 4, S // 2, S * 3 // 4 - m]):
                x = m + i * (bw + 2)
                d.rectangle([x, S - m - h - lw, x + bw, S - m - lw], fill=c)

        elif name == 'gear':
            r_out  = cx - m
            r_in   = max(3, r_out - max(3, S // 8))
            r_hole = max(2, r_in  - max(2, S // 9))
            for i in range(8):
                a1 = _math.radians(i * 45 - 16)
                a2 = _math.radians(i * 45 + 16)
                pts = [
                    (cx + r_in * _math.cos(a1),  cy + r_in * _math.sin(a1)),
                    (cx + r_out * _math.cos(a1), cy + r_out * _math.sin(a1)),
                    (cx + r_out * _math.cos(a2), cy + r_out * _math.sin(a2)),
                    (cx + r_in * _math.cos(a2),  cy + r_in * _math.sin(a2)),
                ]
                d.polygon([(int(x), int(y)) for x, y in pts], fill=c)
            d.ellipse([cx - r_in, cy - r_in, cx + r_in, cy + r_in], fill=c)
            d.ellipse([cx - r_hole, cy - r_hole, cx + r_hole, cy + r_hole], fill=(0, 0, 0, 0))

        elif name == 'camera':
            d.rounded_rectangle([m, S // 3, S - m, S - m], radius=2, fill=c)
            d.rectangle([S // 3, m + 2, S * 2 // 3, S // 3 + 2], fill=c)
            r = max(3, S // 5)
            d.ellipse([cx - r, cy - r + 1, cx + r, cy + r + 1], outline=c, width=lw)
            r2 = max(1, r - 3)
            d.ellipse([cx - r2, cy - r2 + 1, cx + r2, cy + r2 + 1], fill=c)
            d.ellipse([cx - 2, cy - 1, cx + 2, cy + 3], fill=(0, 0, 0, 0))

        elif name == 'disk':
            d.rounded_rectangle([m, m, S - m, S - m], radius=2, fill=c)
            d.rectangle([m + 3, m + 3, S - m - 3, S // 2], fill=dim)
            sw = max(4, S // 5)
            d.rectangle([cx - sw // 2, m + 3, cx + sw // 2, S // 3], fill=c)
            d.rounded_rectangle([m + 4, S * 2 // 3, S - m - 4, S - m - 3], radius=1, fill=dim)

        elif name == 'undo':
            r = cx - m
            d.arc([m, m + 2, S - m, S - m + 2], start=140, end=390, fill=c, width=lw + 1)
            ax = int(cx + r * _math.cos(_math.radians(140)))
            ay = int(cy + r * _math.sin(_math.radians(140))) + 2
            ah = max(4, S // 7)
            d.polygon([(ax + ah, ay - ah // 2), (ax - ah // 3, ay + ah // 2), (ax + ah, ay + ah)], fill=c)

        elif name == 'reset':
            r = cx - m
            d.arc([m, m, S - m, S - m], start=55, end=335, fill=c, width=lw + 1)
            ax = int(cx + r * _math.cos(_math.radians(55)))
            ay = int(cy + r * _math.sin(_math.radians(55)))
            ah = max(4, S // 7)
            d.polygon([(ax - ah, ay + ah // 2), (ax + ah // 2, ay - ah), (ax + ah // 2, ay + ah)], fill=c)

        elif name == 'check':
            d.line([m + 2, S * 3 // 5, cx - 1, S - m - 2, S - m - 1, m + 4],
                   fill=(0, 210, 90, 255), width=lw + 2)

        elif name == 'xmark':
            d.line([m + 3, m + 3, S - m - 3, S - m - 3], fill=(220, 60, 60, 255), width=lw + 1)
            d.line([S - m - 3, m + 3, m + 3, S - m - 3], fill=(220, 60, 60, 255), width=lw + 1)

        elif name == 'printer':
            d.rounded_rectangle([m + 3, S // 3, S - m - 3, S * 3 // 4], radius=2, fill=c)
            d.rectangle([m + 5, m + 2, S - m - 5, S // 3 + 2], fill=c)
            d.rectangle([m + 5, S * 3 // 4 - 1, S - m - 5, S - m - 2], fill=c)
            d.rectangle([m + 5, S // 2, S - m - 5, S // 2 + lw], fill=(0, 0, 0, 160))

        elif name == 'excel':
            d.rounded_rectangle([m + 2, m, S - m - 2, S - m], radius=2, fill=c)
            corner = max(4, S // 6)
            d.polygon([(S - m - 2 - corner, m), (S - m - 2, m + corner), (S - m - 2, m)], fill=dim)
            grn = (34, 180, 34, 255)
            for dy in [-4, 0, 4]:
                d.line([m + 5, cy + dy, S - m - 5, cy + dy], fill=grn, width=lw)

        elif name == 'zoom_plus':
            r = max(4, cx - m - 2)
            d.arc([m, m, m + r * 2, m + r * 2], 0, 360, fill=c, width=lw)
            hx, hy = m + r, m + r
            d.line([hx - 3, hy, hx + 3, hy], fill=c, width=lw)
            d.line([hx, hy - 3, hx, hy + 3], fill=c, width=lw)
            d.line([m + r * 2 - 1, m + r * 2 - 1, S - m - 1, S - m - 1], fill=c, width=lw + 1)

        elif name == 'zoom_minus':
            r = max(4, cx - m - 2)
            d.arc([m, m, m + r * 2, m + r * 2], 0, 360, fill=c, width=lw)
            hx, hy = m + r, m + r
            d.line([hx - 3, hy, hx + 3, hy], fill=c, width=lw)
            d.line([m + r * 2 - 1, m + r * 2 - 1, S - m - 1, S - m - 1], fill=c, width=lw + 1)

        elif name == 'select':
            dl = max(2, S // 9)
            for x in range(m, S - m, dl * 2):
                d.line([x, m, min(x + dl, S - m), m], fill=c, width=lw)
                d.line([x, S - m, min(x + dl, S - m), S - m], fill=c, width=lw)
            for y in range(m, S - m, dl * 2):
                d.line([m, y, m, min(y + dl, S - m)], fill=c, width=lw)
                d.line([S - m, y, S - m, min(y + dl, S - m)], fill=c, width=lw)

        elif name == 'grid':
            cell = max(4, (S - 2 * m - 4) // 3)
            for row in range(3):
                for col in range(3):
                    x, y = m + col * (cell + 2), m + row * (cell + 2)
                    d.rectangle([x, y, x + cell, y + cell], fill=c)

        elif name == 'next':
            d.polygon([(m + 4, m + 3), (S - m - 3, cy), (m + 4, S - m - 3)], fill=c)

        elif name == 'bell':
            d.chord([m + 1, m + 2, S - m - 1, S - m - 4], 200, 340, fill=c)
            d.arc([cx - 3, m - 1, cx + 3, m + 5], 180, 360, fill=c, width=lw)
            d.ellipse([cx - 2, S - m - 5, cx + 2, S - m - 1], fill=c)

        elif name == 'bell_off':
            d.chord([m + 1, m + 2, S - m - 1, S - m - 4], 200, 340,
                    fill=(100, 100, 100, 200))
            d.line([m, m, S - m, S - m], fill=(220, 60, 60, 255), width=lw + 1)

        elif name == 'sun':
            r = max(3, S // 5)
            d.ellipse([cx - r, cy - r, cx + r, cy + r], fill=c)
            ro = cx - m
            for i in range(8):
                a = _math.radians(i * 45)
                x1, y1 = cx + (r + 2) * _math.cos(a), cy + (r + 2) * _math.sin(a)
                x2, y2 = cx + ro * _math.cos(a), cy + ro * _math.sin(a)
                d.line([int(x1), int(y1), int(x2), int(y2)], fill=c, width=lw)

        elif name == 'moon':
            ro = cx - m
            d.ellipse([cx - ro, cy - ro, cx + ro, cy + ro], fill=c)
            off = max(3, ro // 2)
            d.ellipse([cx - ro + off, cy - ro, cx + ro + off, cy + ro], fill=(0, 0, 0, 0))

        elif name == 'fullscreen':
            cl = max(4, S // 6)
            for sx, sy in [(-1, -1), (1, -1), (1, 1), (-1, 1)]:
                bx = cx + sx * (cx - m - 1)
                by = cy + sy * (cy - m - 1)
                d.line([bx, by, bx + sx * cl, by], fill=c, width=lw)
                d.line([bx, by, bx, by + sy * cl], fill=c, width=lw)

        elif name == 'windowed':
            cl = max(4, S // 6)
            for sx, sy in [(-1, -1), (1, -1), (1, 1), (-1, 1)]:
                bx = cx + sx * cl
                by = cy + sy * cl
                d.line([bx, by, bx - sx * cl, by], fill=c, width=lw)
                d.line([bx, by, bx, by - sy * cl], fill=c, width=lw)

    except Exception:
        pass
    return img


def _ph(name: str, size: int = 22, theme: str = None) -> ImageTk.PhotoImage:
    """테마·기능별 고유 색상을 가진 HMI 아이콘 PhotoImage 반환"""
    t = theme or _CURRENT_THEME
    if name in _ICO_CLR:
        fg = _ICO_CLR[name][0] if t == 'dark' else _ICO_CLR[name][1]
    else:
        fg = (216, 228, 240) if t == 'dark' else (30, 50, 80)
    key = (name, size, fg)
    if key not in _ICON_CACHE:
        _ICON_CACHE[key] = ImageTk.PhotoImage(_make_icon(name, size, fg))
    return _ICON_CACHE[key]


def _setup_hover(btn: tk.Button, bg_norm: str, bg_hover: str,
                 fg_norm: str = '#ffffff', fg_hover: str = '#ffffff'):
    """버튼 Hover(마우스오버) 색상 강조 바인딩"""
    btn.configure(bg=bg_norm, fg=fg_norm, activebackground=bg_hover,
                  activeforeground=fg_hover)
    btn.bind('<Enter>', lambda e: btn.configure(bg=bg_hover))
    btn.bind('<Leave>', lambda e: btn.configure(bg=bg_norm))


class HistoryWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self._parent = parent
        # 이 창은 열릴 때마다 새로 만들어지는 Toplevel이라, 메인 창 테마 토글에
        # 실시간으로 반응할 필요는 없음 — 여는 시점의 테마를 한 번만 읽어서
        # 그 값으로 팔레트를 고정한다(라이브 리테마 불필요).
        self._theme = getattr(parent, 'current_theme', 'dark')
        if self._theme == 'light':
            self._c = dict(
                bg=LT_BG, card=LT_CARD, panel=LT_PANEL, border=LT_BORDER,
                txt=LT_TXT, txt_g=LT_TXT_G, txt_dim='#8a9db5',
                header_bg=LT_HDR, header_fg='#ffffff', header_sub='#dbe8fb',
                entry_bg='#ffffff', entry_fg=LT_TXT,
                tree_bg='#ffffff', tree_heading_bg=LT_PANEL, tree_heading_fg=LT_TXT_G,
                tree_sel='#bcd4f0',
                tag_fail_bg='#fbdede', tag_fail_fg='#a5040d',
                tag_warn_bg='#fff2cc', tag_warn_fg='#8a5a00',
                tag_ok_bg='#ffffff',   tag_ok_fg='#1b7a1b',
                btn_neutral_bg='#dfe7f2', btn_neutral_fg=LT_TXT,
                canvas_bg='#fbfcfe', canvas_grid='#dde3ea', canvas_axis='#9fb0c4',
                canvas_label='#5a6b80', canvas_title=LT_TXT,
                canvas_fail_zone='#fdecea', canvas_ok_zone='#eaf7ea',
                canvas_limit='#c62828', canvas_avg='#3a6bbf', canvas_fill='#dbe9fb',
                canvas_good='#1b8a3a', canvas_bad='#c62828',
                tip_bg='#ffffff', tip_fg=LT_TXT, tip_border=LT_BORDER,
            )
        else:
            self._c = dict(
                bg=BG, card=CARD_BG, panel=PANEL_BG, border=BORDER,
                txt=TXT_W, txt_g=TXT_G, txt_dim=TXT_DIM,
                header_bg='#1a3a6b', header_fg=TXT_W, header_sub='#aaccee',
                entry_bg='#21262d', entry_fg=TXT_W,
                tree_bg=CARD_BG, tree_heading_bg='#21262d', tree_heading_fg=TXT_G,
                tree_sel='#2c5f9e',
                tag_fail_bg='#3d1010', tag_fail_fg='#ff8080',
                tag_warn_bg='#3d3000', tag_warn_fg='#ffc107',
                tag_ok_bg=CARD_BG,     tag_ok_fg='#aaffaa',
                btn_neutral_bg='#21262d', btn_neutral_fg=TXT_W,
                canvas_bg='#0d1117', canvas_grid='#1e2d3d', canvas_axis='#4a5568',
                canvas_label='#7a8fa6', canvas_title=TXT_W,
                canvas_fail_zone='#1a0808', canvas_ok_zone='#081a08',
                canvas_limit='#cc3333', canvas_avg='#6080c0', canvas_fill='#1a2a3a',
                canvas_good='#50c050', canvas_bad='#e05050',
                tip_bg='#21262d', tip_fg=TXT_W, tip_border='#30363d',
            )
        self.title('검사 이력 조회')
        self.configure(bg=self._c['bg'])
        self.geometry('1280x760')
        self.resizable(True, True)
        self.minsize(900, 550)
        self._history = self._load()
        self._build()

    # ── 데이터 로드 ──────────────────────────────────────────────────────────

    def _load(self):
        try:
            return _db_query_all(order='id DESC')   # 최신순
        except Exception:
            return []

    def _unique(self, key):
        """이력에서 key 값 목록(정렬, 빈값 제외)"""
        return ['전체'] + sorted({h.get(key, '') for h in self._history if h.get(key, '')})

    # ── 상단 헤더 + 탭 구성 ──────────────────────────────────────────────────

    def _build(self):
        hb = self._c['header_bg']; hf = self._c['header_fg']
        hdr = tk.Frame(self, bg=hb)
        hdr.pack(fill='x')
        tk.Label(hdr, text='  검사 이력 조회', font=('맑은 고딕', 14, 'bold'),
                 bg=hb, fg=hf).pack(side='left', pady=8, padx=8)

        total   = len(self._history)
        defects = sum(1 for h in self._history if h.get('verdict') == '불량')
        self.lbl_summary = tk.Label(
            hdr,
            text=f'총 {total}건  |  불량 {defects}건  |  양호 {total - defects}건',
            font=('맑은 고딕', 11), bg=hb, fg=self._c['header_sub'])
        self.lbl_summary.pack(side='right', padx=16)

        s = ttk.Style()
        s.theme_use('clam')
        s.configure('Hist.TNotebook', background=self._c['bg'], borderwidth=0)
        s.configure('Hist.TNotebook.Tab', background=self._c['panel'],
                    foreground=self._c['txt'], padding=(10, 6))
        s.map('Hist.TNotebook.Tab',
              background=[('selected', hb)], foreground=[('selected', hf)])

        nb = ttk.Notebook(self, style='Hist.TNotebook')
        nb.pack(fill='both', expand=True, padx=8, pady=8)

        bg = self._c['bg']
        t1 = tk.Frame(nb, bg=bg);  nb.add(t1, text='  이력 목록  ')
        t2 = tk.Frame(nb, bg=bg);  nb.add(t2, text='  모델별 조회  ')
        t3 = tk.Frame(nb, bg=bg);  nb.add(t3, text='  LOT별 조회  ')
        t4 = tk.Frame(nb, bg=bg);  nb.add(t4, text='  작업자별 조회  ')
        t5 = tk.Frame(nb, bg=bg);  nb.add(t5, text='  불량률 추이  ')

        self._build_list(t1)
        self._build_model_stats(t2)
        self._build_lot_stats(t3)
        self._build_operator_stats(t4)
        self._build_trend(t5)

    # ── 공용 필터 바 헬퍼 ────────────────────────────────────────────────────

    def _filter_bar(self, parent, specs):
        """specs = [('라벨', var, values, width), ...]  → 필터 프레임 반환"""
        bar = tk.Frame(parent, bg=self._c['bg'])
        bar.pack(fill='x', padx=6, pady=(6, 2))
        for label, var, values, width in specs:
            tk.Label(bar, text=label, font=('맑은 고딕', 10),
                     bg=self._c['bg'], fg=self._c['txt_g']).pack(side='left', padx=(8, 2))
            cb = ttk.Combobox(bar, textvariable=var, values=values,
                               width=width, state='readonly',
                               font=('맑은 고딕', 10))
            cb.pack(side='left', padx=(0, 4))
        return bar

    def _make_treeview(self, parent, cols, widths, style_name='Hist.Treeview'):
        c = self._c
        s = ttk.Style()
        s.theme_use('clam')
        s.configure(style_name,
                     background=c['tree_bg'], foreground=c['txt'],
                     rowheight=24, fieldbackground=c['tree_bg'],
                     font=('맑은 고딕', 10))
        s.configure(f'{style_name}.Heading',
                     background=c['tree_heading_bg'], foreground=c['tree_heading_fg'],
                     font=('맑은 고딕', 10, 'bold'))
        s.map(style_name, background=[('selected', c['tree_sel'])])
        tv = ttk.Treeview(parent, columns=cols, show='headings', style=style_name)
        for col in cols:
            tv.heading(col, text=col)
            tv.column(col, width=widths.get(col, 80), anchor='center')
        tv.tag_configure('fail', background=c['tag_fail_bg'], foreground=c['tag_fail_fg'])
        tv.tag_configure('warn', background=c['tag_warn_bg'], foreground=c['tag_warn_fg'])
        tv.tag_configure('ok',   background=c['tag_ok_bg'],   foreground=c['tag_ok_fg'])
        sb = ttk.Scrollbar(parent, orient='vertical', command=tv.yview)
        tv.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        tv.pack(side='left', fill='both', expand=True)
        return tv

    # ── Tab 1: 이력 목록 (모델 + LOT + 작업자 + 판정 필터) ──────────────────

    def _build_list(self, parent):
        c = self._c
        # ── 필터 변수 ──────────────────────────────────────────────────────────
        self.fv_model   = tk.StringVar(value='전체')
        self.fv_lot     = tk.StringVar(value='전체')
        self.fv_op      = tk.StringVar(value='전체')
        self.fv_verdict   = tk.StringVar(value='전체')
        self.fv_date_from = tk.StringVar(value='')
        self.fv_date_to   = tk.StringVar(value='')
        self.fv_serial  = tk.StringVar(value='')
        self._checked_keys = set()   # 체크 선택 (serial+date+time 키) — 페이지 이동해도 유지됨
        self._iid_to_entry = {}
        self._page = 1
        self._page_size = 200

        # ── 필터 행 1: Combobox ───────────────────────────────────────────────
        bar = tk.Frame(parent, bg=c['bg'])
        bar.pack(fill='x', padx=6, pady=(6, 1))

        for label, var, values, w in [
            ('모델:',   self.fv_model,   ['전체'] + list(MODELS.keys()),  9),
            ('LOT:',    self.fv_lot,     self._unique('lot'),             13),
            ('작업자:', self.fv_op,      self._unique('operator'),        9),
            ('판정:',   self.fv_verdict, ['전체', '양호', '불량'],         7),
        ]:
            tk.Label(bar, text=label, font=('맑은 고딕', 9),
                     bg=c['bg'], fg=c['txt_g']).pack(side='left', padx=(5, 1))
            cb = ttk.Combobox(bar, textvariable=var, values=values,
                               width=w, state='readonly', font=('맑은 고딕', 9))
            cb.pack(side='left', padx=(0, 2))
            cb.bind('<<ComboboxSelected>>', lambda e: self._apply_filters())

        tk.Label(bar, text='SN:', font=('맑은 고딕', 9),
                 bg=c['bg'], fg=c['txt_g']).pack(side='left', padx=(6, 1))
        sn_e = tk.Entry(bar, textvariable=self.fv_serial, width=13,
                        font=('맑은 고딕', 9), bg=c['entry_bg'], fg=c['entry_fg'],
                        insertbackground=c['entry_fg'], relief='flat')
        sn_e.pack(side='left', padx=(0, 2))
        sn_e.bind('<KeyRelease>', lambda e: self._apply_filters())

        # ── 필터 행 2: 기간(날짜 범위) ────────────────────────────────────────
        date_bar = tk.Frame(parent, bg=c['bg'])
        date_bar.pack(fill='x', padx=6, pady=(0, 2))
        tk.Label(date_bar, text='기간:', font=('맑은 고딕', 9),
                 bg=c['bg'], fg=c['txt_g']).pack(side='left', padx=(5, 1))
        for var in (self.fv_date_from, self.fv_date_to):
            e = tk.Entry(date_bar, textvariable=var, width=11,
                          font=('맑은 고딕', 9), bg=c['entry_bg'], fg=c['entry_fg'],
                          insertbackground=c['entry_fg'], relief='flat')
            e.pack(side='left', padx=2)
            e.bind('<KeyRelease>', lambda ev: self._apply_filters())
            if var is self.fv_date_from:
                tk.Label(date_bar, text='~', font=('맑은 고딕', 9),
                         bg=c['bg'], fg=c['txt_g']).pack(side='left')
        tk.Label(date_bar, text='(YYYY-MM-DD, 비워두면 전체)', font=('맑은 고딕', 8),
                 bg=c['bg'], fg=c['txt_dim']).pack(side='left', padx=(4, 10))
        for lbl, kw in [('오늘', {'today': True}), ('이번주', {'this_week': True}),
                        ('이번달', {'this_month': True}), ('전체', {'clear': True})]:
            b = tk.Button(date_bar, text=lbl, font=('맑은 고딕', 8),
                          bg=c['btn_neutral_bg'], fg=c['txt_g'], relief='flat', cursor='hand2',
                          padx=6, pady=1,
                          command=lambda kw=kw: self._set_date_range(**kw))
            b.pack(side='left', padx=2)

        # ── 버튼 행 ───────────────────────────────────────────────────────────
        bar2 = tk.Frame(parent, bg=c['bg'])
        bar2.pack(fill='x', padx=6, pady=(0, 2))

        btn_rst = tk.Button(bar2, text='  필터 초기화', font=('맑은 고딕', 9),
                  relief='flat', cursor='hand2', compound='left',
                  command=self._reset_list_filters)
        btn_rst.configure(image=_ph('reset', 18))
        _setup_hover(btn_rst, '#2a3540', '#3a4f60')
        btn_rst.pack(side='left', padx=4, pady=2)

        btn_all = tk.Button(bar2, text='전체 선택/해제', font=('맑은 고딕', 9),
                            relief='flat', cursor='hand2',
                            command=self._toggle_all_checks)
        _setup_hover(btn_all, '#1a2a3a', '#2a4060')
        btn_all.pack(side='left', padx=4, pady=2)

        self.lbl_sel_count = tk.Label(bar2, text='선택: 0건',
                                       bg=c['bg'], fg=ACC_YEL,
                                       font=('맑은 고딕', 9))
        self.lbl_sel_count.pack(side='left', padx=8)

        btn_report = tk.Button(bar2, text='  검사성적서 생성', font=('맑은 고딕', 10, 'bold'),
                               relief='flat', cursor='hand2', compound='left',
                               command=self._show_report_dialog)
        btn_report.configure(image=_ph('printer', 22))
        _setup_hover(btn_report, '#1a3a6b', '#2a5aab')
        btn_report.pack(side='right', padx=4, pady=2)

        btn_xl = tk.Button(bar2, text='  Excel 빠른 저장', font=('맑은 고딕', 10),
                  relief='flat', cursor='hand2', compound='left',
                  command=self._export_excel)
        btn_xl.configure(image=_ph('excel', 20))
        _setup_hover(btn_xl, '#1a5c2e', '#0f8040')
        btn_xl.pack(side='right', padx=4, pady=2)

        btn_bkp = tk.Button(bar2, text='  이력 백업', font=('맑은 고딕', 10),
                  relief='flat', cursor='hand2', compound='left',
                  command=self._backup_now)
        btn_bkp.configure(image=_ph('disk', 20))
        _setup_hover(btn_bkp, '#1a3a5c', '#2a5a8c')
        btn_bkp.pack(side='right', padx=4, pady=2)

        btn_rst_bkp = tk.Button(bar2, text='  백업 복원', font=('맑은 고딕', 10),
                  relief='flat', cursor='hand2', compound='left',
                  command=self._restore_from_backup)
        btn_rst_bkp.configure(image=_ph('folder', 20))
        _setup_hover(btn_rst_bkp, '#2a2a5c', '#3a3a8c')
        btn_rst_bkp.pack(side='right', padx=4, pady=2)

        btn_clr = tk.Button(bar2, text='  이력 삭제', font=('맑은 고딕', 10),
                  relief='flat', cursor='hand2', compound='left',
                  command=self._clear)
        btn_clr.configure(image=_ph('xmark', 20))
        _setup_hover(btn_clr, '#3d1010', '#5c2020')
        btn_clr.pack(side='right', padx=4, pady=2)

        # ── 페이지네이션 바 (하단 고정 — Treeview보다 먼저 pack 해야 자리를 차지함) ──
        pg_bar = tk.Frame(parent, bg=c['bg'])
        pg_bar.pack(side='bottom', fill='x', padx=6, pady=(2, 4))
        self.btn_page_prev = tk.Button(
            pg_bar, text='◀ 이전', font=('맑은 고딕', 9),
            bg=c['btn_neutral_bg'], fg=c['txt'], relief='flat', cursor='hand2',
            padx=8, pady=2, command=self._go_prev_page)
        self.btn_page_prev.pack(side='left', padx=2)
        self.btn_page_next = tk.Button(
            pg_bar, text='다음 ▶', font=('맑은 고딕', 9),
            bg=c['btn_neutral_bg'], fg=c['txt'], relief='flat', cursor='hand2',
            padx=8, pady=2, command=self._go_next_page)
        self.btn_page_next.pack(side='left', padx=2)
        self.lbl_page = tk.Label(pg_bar, text='', font=('맑은 고딕', 9),
                                  bg=c['bg'], fg=c['txt_g'])
        self.lbl_page.pack(side='left', padx=10)
        tk.Label(pg_bar, text=f'(한 페이지 {self._page_size}건씩, 최신순)',
                 font=('맑은 고딕', 8), bg=c['bg'], fg=c['txt_dim']).pack(side='left')

        # ── Treeview (체크박스 ✓ 컬럼 추가) ──────────────────────────────────
        cols   = ('✓', '번호', '날짜', '시각', 'LOT No.', '작업자', '일련번호',
                  '모델', '홀막힘', '불량률', '판정', '불량사유')
        widths = {'✓': 30, '번호': 45, '날짜': 100, '시각': 75, 'LOT No.': 120,
                  '작업자': 80, '일련번호': 110, '모델': 65, '홀막힘': 65, '불량률': 72, '판정': 65,
                  '불량사유': 220}
        self.list_tree = self._make_treeview(parent, cols, widths)
        self.list_tree.bind('<ButtonRelease-1>', self._on_list_click)
        self._refresh_list()

    def _apply_filters(self):
        """필터 값이 바뀌었을 때: 1페이지로 되돌리고 다시 그린다."""
        self._page = 1
        self._refresh_list()

    def _go_prev_page(self):
        if self._page > 1:
            self._page -= 1
            self._refresh_list()

    def _go_next_page(self):
        total = len(self._filtered_list())
        total_pages = max(1, -(-total // self._page_size))
        if self._page < total_pages:
            self._page += 1
            self._refresh_list()

    def _filtered_list(self):
        m      = self.fv_model.get()
        lo     = self.fv_lot.get()
        op     = self.fv_op.get()
        vd     = self.fv_verdict.get()
        d_from = self.fv_date_from.get().strip()
        d_to   = self.fv_date_to.get().strip()
        sn     = self.fv_serial.get().strip().lower()
        return [h for h in self._history
                if (m  == '전체' or h.get('model', '')    == m)
                and (lo == '전체' or h.get('lot', '')      == lo)
                and (op == '전체' or h.get('operator', '') == op)
                and (vd == '전체' or h.get('verdict', '')  == vd)
                and (not d_from or h.get('date', '') >= d_from)
                and (not d_to   or h.get('date', '') <= d_to)
                and (not sn     or sn in h.get('serial', '').lower())]

    def _set_date_range(self, today=False, this_week=False, this_month=False, clear=False):
        d = datetime.date.today()
        if clear:
            self.fv_date_from.set('')
            self.fv_date_to.set('')
        elif today:
            s = d.strftime('%Y-%m-%d')
            self.fv_date_from.set(s)
            self.fv_date_to.set(s)
        elif this_week:
            start = d - datetime.timedelta(days=d.weekday())
            self.fv_date_from.set(start.strftime('%Y-%m-%d'))
            self.fv_date_to.set(d.strftime('%Y-%m-%d'))
        elif this_month:
            start = d.replace(day=1)
            self.fv_date_from.set(start.strftime('%Y-%m-%d'))
            self.fv_date_to.set(d.strftime('%Y-%m-%d'))
        self._apply_filters()

    @staticmethod
    def _row_key(h):
        return h.get('serial', '') + h.get('date', '') + h.get('time', '')

    def _refresh_list(self, *_):
        for row in self.list_tree.get_children():
            self.list_tree.delete(row)
        self._iid_to_entry.clear()

        filtered    = self._filtered_list()
        total       = len(filtered)
        total_pages = max(1, -(-total // self._page_size))   # ceil
        self._page  = max(1, min(self._page, total_pages))
        start       = (self._page - 1) * self._page_size
        page_rows   = filtered[start:start + self._page_size]

        for i, h in enumerate(page_rows, start + 1):
            rate    = h.get('rate', 0)
            verdict = h.get('verdict', '')
            tag     = 'fail' if verdict == '불량' else ('warn' if rate >= DEFECT_LIMIT * 0.75 else 'ok')
            checked = self._row_key(h) in self._checked_keys
            iid = self.list_tree.insert('', 'end', tags=(tag,), values=(
                '☑' if checked else '☐',
                i, h.get('date',''), h.get('time',''),
                h.get('lot',''), h.get('operator',''), h.get('serial',''),
                h.get('model',''), h.get('count',''),
                f'{rate:.2f}%', verdict, h.get('defect_reason', ''),
            ))
            self._iid_to_entry[iid] = h
        self._update_sel_label()
        if hasattr(self, 'lbl_page'):
            self.lbl_page.configure(
                text=f'{self._page} / {total_pages} 페이지  (총 {total:,}건)')
            self.btn_page_prev.configure(state='normal' if self._page > 1 else 'disabled')
            self.btn_page_next.configure(
                state='normal' if self._page < total_pages else 'disabled')

    def _on_list_click(self, event):
        if self.list_tree.identify_region(event.x, event.y) != 'cell':
            return
        if self.list_tree.identify_column(event.x) != '#1':
            return
        iid = self.list_tree.identify_row(event.y)
        if iid:
            self._toggle_check(iid)

    def _toggle_check(self, iid):
        h = self._iid_to_entry.get(iid)
        if h is None:
            return
        key  = self._row_key(h)
        vals = list(self.list_tree.item(iid, 'values'))
        if key in self._checked_keys:
            self._checked_keys.discard(key)
            vals[0] = '☐'
        else:
            self._checked_keys.add(key)
            vals[0] = '☑'
        self.list_tree.item(iid, values=vals)
        self._update_sel_label()

    def _toggle_all_checks(self):
        """전체 선택/해제는 현재 페이지에 보이는 행만 대상으로 한다."""
        all_iids  = self.list_tree.get_children()
        page_keys = [self._row_key(self._iid_to_entry[iid]) for iid in all_iids]
        all_checked = bool(page_keys) and all(k in self._checked_keys for k in page_keys)
        if all_checked:
            self._checked_keys.difference_update(page_keys)
            mark = '☐'
        else:
            self._checked_keys.update(page_keys)
            mark = '☑'
        for iid in all_iids:
            vals    = list(self.list_tree.item(iid, 'values'))
            vals[0] = mark
            self.list_tree.item(iid, values=vals)
        self._update_sel_label()

    def _update_sel_label(self):
        if hasattr(self, 'lbl_sel_count'):
            self.lbl_sel_count.configure(
                text=f'선택: {len(self._checked_keys)}건')

    def _reset_list_filters(self):
        for v in (self.fv_model, self.fv_lot, self.fv_op, self.fv_verdict):
            v.set('전체')
        self.fv_date_from.set('')
        self.fv_date_to.set('')
        self.fv_serial.set('')
        self._apply_filters()

    # ── 검사성적서 출력 옵션 다이얼로그 ───────────────────────────────────────

    def _show_report_dialog(self):
        all_rows = self._filtered_list()
        ok_rows  = [h for h in all_rows if h.get('verdict','') == '양호']
        ng_rows  = [h for h in all_rows if h.get('verdict','') == '불량']
        # 체크 선택은 페이지를 넘나들며 누적되므로, 현재 필터 결과 전체에서 찾는다
        # (화면에 지금 보이는 페이지만이 아니라 다른 페이지의 선택분도 포함됨).
        sel_rows = [h for h in all_rows if self._row_key(h) in self._checked_keys]

        c = self._c
        dlg = tk.Toplevel(self)
        dlg.title('검사성적서 출력 옵션')
        dlg.configure(bg=c['card'])
        dlg.geometry('440x340')
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self)

        # 헤더
        hdr_f = tk.Frame(dlg, bg=c['header_bg'])
        hdr_f.pack(fill='x')
        tk.Label(hdr_f, text='  검사성적서 출력 범위 선택',
                 font=('맑은 고딕', 13, 'bold'),
                 bg=c['header_bg'], fg=c['header_fg']).pack(anchor='w', pady=10, padx=10)

        # 라디오 버튼
        mode_var = tk.StringVar(value='전체')
        options  = [
            ('전체', f'전체 출력              ({len(all_rows):,}건)'),
            ('OK',   f'OK(양호) 제품만 출력   ({len(ok_rows):,}건)'),
            ('NG',   f'NG(불량) 제품만 출력   ({len(ng_rows):,}건)'),
            ('선택', f'선택한 제품만 출력     ({len(sel_rows):,}건 선택됨)'),
        ]
        rb_frame = tk.Frame(dlg, bg=c['card'])
        rb_frame.pack(fill='x', padx=20, pady=16)
        for value, text in options:
            rb = tk.Radiobutton(
                rb_frame, text=text, variable=mode_var, value=value,
                font=('맑은 고딕', 11), bg=c['card'], fg=c['txt'],
                activebackground=c['card'], activeforeground=c['txt'],
                selectcolor=c['header_bg'], cursor='hand2')
            rb.pack(anchor='w', pady=5)

        # 경고 라벨
        warn = tk.Label(dlg, text='', bg=c['card'], fg=ACC_YEL,
                        font=('맑은 고딕', 9))
        warn.pack()

        # 결과를 담을 컨테이너 (wait_window 후에 꺼냄)
        result = {'mode': None, 'rows': None}

        def on_confirm():
            m = mode_var.get()
            if m == '전체':   r = all_rows
            elif m == 'OK':   r = ok_rows
            elif m == 'NG':   r = ng_rows
            else:             r = sel_rows
            if not r:
                warn.configure(text=f'⚠ 출력 대상이 없습니다 ({m} 선택)')
                return
            result['mode'] = m
            result['rows'] = r
            dlg.destroy()          # destroy만 하고 여기서 끝

        btn_f = tk.Frame(dlg, bg=c['card'])
        btn_f.pack(side='bottom', fill='x', padx=16, pady=14)
        tk.Button(btn_f, text='취소', font=('맑은 고딕', 10),
                  bg=c['btn_neutral_bg'], fg=c['txt'], relief='flat', width=10,
                  cursor='hand2', command=dlg.destroy).pack(side='left', padx=4)
        btn_ok = tk.Button(btn_f, text='  보고서 생성', font=('맑은 고딕', 11, 'bold'),
                           bg=c['header_bg'], fg=c['header_fg'], relief='flat',
                           compound='left', cursor='hand2',
                           command=on_confirm)
        btn_ok.configure(image=_ph('printer', 22))
        btn_ok.pack(side='right', padx=4)

        dlg.wait_window()   # 다이얼로그가 완전히 닫힐 때까지 대기

        # 다이얼로그가 완전히 사라진 후 보고서 생성 (파일 다이얼로그 충돌 방지)
        if result['mode']:
            self.after(50, lambda: self._generate_report(
                result['mode'], result['rows'], all_rows))

    # ── 검사성적서 생성 (Excel) ────────────────────────────────────────────────

    def _generate_report(self, mode: str, rows: list, all_rows: list):
        if not rows:
            messagebox.showwarning('경고', '출력할 데이터가 없습니다.')
            return

        ts        = datetime.datetime.now().strftime('%Y%m%d')
        lot       = rows[0].get('lot',      '미지정')
        operator  = rows[0].get('operator', '미지정')
        insp_date = rows[0].get('date',     f'{ts[:4]}-{ts[4:6]}-{ts[6:]}')
        label_map = {'전체': '전체', 'OK': 'OK', 'NG': 'NG', '선택': '선택'}
        label     = label_map.get(mode, mode)
        safe_lot  = lot.replace('/', '-').replace('\\', '-')
        fname     = f'멤브레인홀막힘_검사_{ts}_{safe_lot}.xlsx'

        filepath = filedialog.asksaveasfilename(
            initialfile=fname,
            initialdir=os.path.expanduser('~/Desktop'),
            defaultextension='.xlsx',
            filetypes=[('Excel 통합문서 (*.xlsx)', '*.xlsx'), ('모든 파일', '*.*')])
        if not filepath:
            return
        if not filepath.lower().endswith('.xlsx'):
            filepath += '.xlsx'

        # ── 판정 기준: <3%=합격, 3~4%=양품, ≥4%=부적격 ─────────────────
        PASS_LIMIT  = 3.0            # 성적서 전용 3단계(합격/양품/부적격) 중간 기준
        FAIL_LIMIT  = DEFECT_LIMIT   # 화면 실시간 판정 기준과 항상 동일하게 연동
        TOTAL_HOLES = 300

        def _new_verdict(rate):
            if rate >= FAIL_LIMIT:  return '부적격'
            if rate >= PASS_LIMIT:  return '양품'
            return '합격'

        verdicts   = [_new_verdict(h.get('rate', 0)) for h in rows]
        total      = len(rows)
        pass_cnt   = verdicts.count('합격')
        good_cnt   = verdicts.count('양품')
        reject_cnt = verdicts.count('부적격')
        pass_rate  = round(pass_cnt / total * 100, 1) if total else 0.0

        if reject_cnt > 0:
            final_verdict = '부적격'
            final_note    = f'부적격 {reject_cnt}개'
        elif good_cnt > 0:
            final_verdict = '조건부 합격'
            final_note    = f'재검사 권장 - 양품 {good_cnt}개'
        else:
            final_verdict = '합격'
            final_note    = '전 제품 합격 판정'

        # ── 워크북 ──────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '검사보고서'

        ws.page_setup.orientation  = 'landscape'
        ws.page_setup.paperSize    = 9          # A4
        ws.page_margins.top        = 0.59       # ≈15 mm
        ws.page_margins.bottom     = 0.59
        ws.page_margins.left       = 0.47       # ≈12 mm
        ws.page_margins.right      = 0.47
        ws.page_setup.fitToPage    = True
        ws.page_setup.fitToWidth   = 1
        ws.page_setup.fitToHeight  = 0

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12

        # ── 스타일 헬퍼 ─────────────────────────────────────────────────
        def _fill(r, g, b):
            return PatternFill(patternType='solid',
                               fgColor='{:02X}{:02X}{:02X}'.format(r, g, b))

        def _bdr(style='thin', color='CCCCCC'):
            s = Side(style=style, color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        # ── 색상 체계 : 문서 전체에서 "같은 의미 = 같은 색"으로 통일 ──────
        # (기존 버전은 부적격을 표 안에서 CC0000/A5040D 두 가지, 양품을 E67C0F/8A5A00
        #  두 가지 다른 hex로 섞어 써서 같은 의미인데 색이 미묘하게 달랐음)
        NAVY = '0D3B66'                                    # 브랜드 메인 — 제목/표 헤더
        GOOD_FG,  GOOD_BG  = '1B5E20', (200, 230, 201)      # 합격 — 초록
        WARN_FG,  WARN_BG  = '8A5A00', (255, 243, 224)      # 양품 — 호박색
        CRIT_FG,  CRIT_BG  = 'A5040D', (255, 205, 210)      # 부적격 — 빨강
        VERDICT_COLORS = {'합격': (GOOD_FG, GOOD_BG),
                           '양품': (WARN_FG, WARN_BG),
                           '부적격': (CRIT_FG, CRIT_BG)}

        ctr        = Alignment(horizontal='center', vertical='center')
        med_blue   = _bdr('medium', '0D6EFD')
        med_orange = _bdr('medium', 'FF9800')
        dark_bdr   = _bdr('medium', NAVY)
        info_bdr   = _bdr('thin',   '999999')
        dat_bdr    = _bdr('thin',   NAVY)
        row_bdr    = _bdr('thin',   'DDDDDD')
        crit_edge  = Side(style='thick', color=CRIT_FG)    # 부적격 행 좌측 강조선(스캔용)
        warn_edge  = Side(style='thick', color=WARN_FG)    # 양품 행 좌측 강조선(스캔용)

        def _cell(row, col, value='', size=10, bold=False,
                  fg='000000', bg=None, border=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(name='Arial', size=size, bold=bold, color=fg)
            c.alignment = ctr
            if bg:     c.fill   = bg
            if border: c.border = border
            return c

        def _mg(row, c1, c2, value='', size=10, bold=False,
                fg='000000', bg=None, border=None):
            """c1:c2 병합 — 4면 윤곽선 완전 적용
            · anchor(c1): 전체 border
            · 중간 열: top+bottom (상하 윤곽선)
            · 마지막 열(c2): right+top+bottom (우측 윤곽선)
            """
            ws.merge_cells(start_row=row, end_row=row,
                           start_column=c1, end_column=c2)
            anchor = ws.cell(row=row, column=c1)
            anchor.value  = value
            anchor.font   = Font(name='Arial', size=size, bold=bold, color=fg)
            anchor.alignment = ctr
            if border:
                anchor.border = border
                for col in range(c1 + 1, c2):          # 중간 열: 상·하
                    ws.cell(row=row, column=col).border = Border(
                        top=border.top, bottom=border.bottom)
                if c2 > c1:                             # 마지막 열: 우·상·하
                    ws.cell(row=row, column=c2).border = Border(
                        right=border.right, top=border.top, bottom=border.bottom)
            if bg:
                for col in range(c1, c2 + 1):
                    ws.cell(row=row, column=col).fill = bg

        # ┌── Row 1: 제목 (A1:F1 병합) ──────────────────────────────────┐
        ws.row_dimensions[1].height = 40
        title_bg = PatternFill(patternType='solid', fgColor=NAVY)
        _mg(1, 1, 6, '멤브레인 홀막힘 검사 보고서',
            size=16, bold=True, fg='FFFFFF', bg=title_bg, border=dark_bdr)

        # ── Row 2: 공백 ─────────────────────────────────────────────────
        ws.row_dimensions[2].height = 6

        # ┌── Row 3: 정보 헤더 (E3:F3 병합) ─────────────────────────────┐
        ws.row_dimensions[3].height = 22
        info_hdr_bg = PatternFill(patternType='solid', fgColor='3A4A5E')
        for ci, lbl in enumerate(['검사일', '검사자', 'Lot 번호', '총 검사수량'], 1):
            _cell(3, ci, lbl, bold=True, fg='FFFFFF',
                  bg=info_hdr_bg, border=info_bdr)
        _mg(3, 5, 6, '비고', bold=True, fg='FFFFFF',
            bg=info_hdr_bg, border=info_bdr)

        # ┌── Row 4: 정보 값 (E4:F4 병합) ───────────────────────────────┐
        ws.row_dimensions[4].height = 24
        info_val_bg = _fill(232, 244, 248)
        for ci, v in enumerate([insp_date, _xl_safe(operator), _xl_safe(lot), f'{total} EA'], 1):
            _cell(4, ci, v, fg='000000', bg=info_val_bg, border=info_bdr)
        note_fg = {'합격': GOOD_FG, '조건부 합격': WARN_FG, '부적격': CRIT_FG}.get(
            final_verdict, '000000')
        _mg(4, 5, 6, final_note, bold=True, fg=note_fg, bg=info_val_bg, border=info_bdr)

        # ┌── Row 5: 판정 기준 명시 (A5:F5 병합) ─────────────────────────┐
        ws.row_dimensions[5].height = 20
        policy_bg  = _fill(255, 249, 224)
        policy_bdr = _bdr('thin', 'D8B300')
        _mg(5, 1, 6,
            f'판정 기준   :   합격 < {PASS_LIMIT:.1f}%      ㅣ      '
            f'양품 {PASS_LIMIT:.1f}% ~ {FAIL_LIMIT:.1f}%      ㅣ      '
            f'부적격 ≥ {FAIL_LIMIT:.1f}%      '
            f'(막힌비율 = 홀막힘 수 ÷ 총 파이버 수 × 100)',
            size=9, bold=True, fg=WARN_FG, bg=policy_bg, border=policy_bdr)

        # ── Row 6: 공백 ─────────────────────────────────────────────────
        ws.row_dimensions[6].height = 8

        # ┌── Row 7: 데이터 헤더 ─────────────────────────────────────────┐
        hdr_row = 7
        ws.row_dimensions[hdr_row].height = 26
        dat_hdr_bg = PatternFill(patternType='solid', fgColor=NAVY)
        for ci, h in enumerate(['Lot 번호', '일련번호', '검사자',
                                 '홀막힘 (EA)', '막힌비율 (%)', '판정'], 1):
            _cell(hdr_row, ci, h, size=10.5, bold=True, fg='FFFFFF',
                  bg=dat_hdr_bg, border=dat_bdr)

        # ┌── 데이터 행 (Row 8~) ─────────────────────────────────────────┐
        data_start = hdr_row + 1
        data_end   = data_start + total - 1

        for i, (h, vd) in enumerate(zip(rows, verdicts)):
            rn     = data_start + i
            rate   = h.get('rate', 0)
            cnt    = h.get('count', 0)
            fibers = h.get('fibers') or TOTAL_HOLES
            ws.row_dimensions[rn].height = 20

            v_fg, v_bg_rgb = VERDICT_COLORS[vd]
            v_bg = _fill(*v_bg_rgb)
            # 합격 행은 옅은 지브라 줄무늬만, 양품/부적격 행은 행 전체를 옅게
            # 물들이고 좌측에 굵은 강조선을 넣어 스크롤 없이 스캔만으로도
            # 문제 행이 바로 눈에 띄게 한다("핵심 데이터 한눈에" 요구사항).
            if vd == '부적격':
                row_bg   = _fill(255, 241, 241)
                left_edge = crit_edge
            elif vd == '양품':
                row_bg   = _fill(255, 250, 235)
                left_edge = warn_edge
            else:
                row_bg   = _fill(248, 248, 248) if (i % 2 == 0) else _fill(255, 255, 255)
                left_edge = row_bdr.left

            row_border      = row_bdr
            first_col_border = Border(left=left_edge, right=row_bdr.right,
                                       top=row_bdr.top, bottom=row_bdr.bottom)

            for ci, val in enumerate([_xl_safe(h.get('lot','')), _xl_safe(h.get('serial','')),
                                       _xl_safe(h.get('operator',''))], 1):
                _cell(rn, ci, val, fg='000000', bg=row_bg,
                      border=first_col_border if ci == 1 else row_border)

            c = ws.cell(row=rn, column=4, value=cnt)
            c.font = Font(name='Arial', size=10, bold=(vd != '합격'), color=v_fg)
            c.fill = row_bg; c.alignment = ctr; c.border = row_border

            c = ws.cell(row=rn, column=5,
                        value=f'=ROUND((D{rn}/{fibers})*100,2)')
            c.font = Font(name='Arial', size=10, color=v_fg)
            c.fill = row_bg; c.alignment = ctr; c.border = row_border
            c.number_format = '0.00'

            c = ws.cell(row=rn, column=6,
                        value=(f'=IF(E{rn}>={FAIL_LIMIT},"부적격",'
                               f'IF(E{rn}>={PASS_LIMIT},"양품","합격"))'))
            c.font = Font(name='Arial', size=10, bold=True, color=v_fg)
            c.fill = v_bg; c.alignment = ctr; c.border = row_border

        # ── 공백 ────────────────────────────────────────────────────────
        ws.row_dimensions[data_end + 1].height = 10

        # ┌── 불량률 추이 차트 ────────────────────────────────────────────┐
        # 개체별 막힌비율(%) 을 꺾은선으로, 불량 기준선을 점선으로 겹쳐 그려
        # 한눈에 합격/부적격 흐름을 볼 수 있게 한다.
        chart_row      = data_end + 2
        chart_row_span = 16   # 차트가 차지하는 대략적인 행 수(여유 포함)
        if total >= 2:
            limit_col = 8   # H열 — 인쇄영역(A:F) 밖, 차트 전용 데이터라 숨김 처리
            ws.cell(row=hdr_row, column=limit_col, value='불량기준(%)')
            for i in range(total):
                rn = data_start + i
                c = ws.cell(row=rn, column=limit_col, value=FAIL_LIMIT)
                c.number_format = '0.0'
            ws.column_dimensions[
                ws.cell(row=hdr_row, column=limit_col).column_letter].hidden = True

            chart = LineChart()
            chart.title  = '검사 개체별 막힌비율(%) 추이'
            chart.style  = 2
            chart.y_axis.title  = '막힌비율 (%)'
            chart.x_axis.title  = '일련번호'
            chart.height = 7.5
            chart.width  = 17
            chart.x_axis.delete = False
            chart.y_axis.delete = False

            chart.add_data(Reference(ws, min_col=5, min_row=hdr_row, max_row=data_end),
                            titles_from_data=True)
            chart.add_data(Reference(ws, min_col=limit_col, min_row=hdr_row, max_row=data_end),
                            titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=2, min_row=data_start, max_row=data_end))

            s_rate = chart.series[0]
            s_rate.graphicalProperties.line.solidFill = '0D6EFD'
            s_rate.graphicalProperties.line.width = 22000
            s_rate.marker.symbol = 'circle'
            s_rate.marker.size   = 6
            s_rate.smooth = False

            s_limit = chart.series[1]
            s_limit.graphicalProperties.line.solidFill = 'E53935'
            s_limit.graphicalProperties.line.width = 16000
            s_limit.graphicalProperties.line.dashStyle = 'dash'
            s_limit.marker.symbol = 'none'
            s_limit.smooth = False

            ws.add_chart(chart, f'A{chart_row}')
            stat_r1 = chart_row + chart_row_span
        else:
            stat_r1 = data_end + 2

        # ┌── 통계 3박스 (각 2열씩 병합) ────────────────────────────────┐
        stat_r2 = stat_r1 + 1
        ws.row_dimensions[stat_r1].height = 22
        ws.row_dimensions[stat_r2].height = 44

        stat_bg      = _fill(225, 245, 254)
        crit_stat_bg = _fill(255, 231, 231)
        f_rng   = f'F{data_start}:F{data_end}'
        d_rng   = f'D{data_start}:D{data_end}'

        # 부적격 수는 0건이면 다른 지표와 같은 파란 톤(평온함), 1건이라도
        # 있으면 빨강으로 전환 — 보고서를 열자마자 문제 유무가 색으로 즉시 판별됨.
        reject_accent = CRIT_FG if reject_cnt > 0 else '0D6EFD'
        reject_bg     = crit_stat_bg if reject_cnt > 0 else stat_bg
        reject_border = _bdr('medium', CRIT_FG) if reject_cnt > 0 else med_blue

        stats = [
            ('합격률 (%)',
             f'=ROUND(COUNTIF({f_rng},"합격")/COUNTA({f_rng})*100,1)',
             1, 2, '0.0', '0D6EFD', stat_bg, med_blue),
            ('평균 홀막힘 (EA)',
             f'=ROUND(AVERAGE({d_rng}),2)',
             3, 4, '0.00', '0D6EFD', stat_bg, med_blue),
            ('부적격 수 (EA)',
             f'=COUNTIF({f_rng},"부적격")',
             5, 6, '0', reject_accent, reject_bg, reject_border),
        ]
        for lbl, formula, c1, c2, num_fmt, accent, box_bg, box_bdr in stats:
            _mg(stat_r1, c1, c2, lbl,
                bold=True, fg=accent, bg=box_bg, border=box_bdr)
            ws.merge_cells(start_row=stat_r2, end_row=stat_r2,
                           start_column=c1, end_column=c2)
            vc = ws.cell(row=stat_r2, column=c1, value=formula)
            vc.font = Font(name='Arial', size=24, bold=True, color=accent)
            vc.alignment = ctr
            vc.border = box_bdr
            vc.number_format = num_fmt
            if c2 > c1:
                ws.cell(row=stat_r2, column=c2).border = Border(
                    right=box_bdr.right, top=box_bdr.top, bottom=box_bdr.bottom)
            for col in range(c1, c2 + 1):
                ws.cell(row=stat_r2, column=col).fill = box_bg

        # ── 공백 ────────────────────────────────────────────────────────
        ws.row_dimensions[stat_r2 + 1].height = 8

        # ┌── 최종 판정 (A:F 병합 2행) ──────────────────────────────────┐
        vrd_r1 = stat_r2 + 2
        vrd_r2 = stat_r2 + 3
        ws.row_dimensions[vrd_r1].height = 26
        ws.row_dimensions[vrd_r2].height = 40

        # 최종판정 박스 색상 = 실제 심각도(합격/조건부 합격/부적격)를 그대로 반영.
        # 기존엔 결과와 무관하게 항상 주황색 한 가지였음 — "부적격"이 나와도
        # "합격"과 똑같은 색으로 보여 한눈에 심각도를 알기 어려웠던 문제 수정.
        vrd_palette = {
            '합격':        (GOOD_FG, _fill(224, 242, 225)),
            '조건부 합격': (WARN_FG, _fill(255, 243, 224)),
            '부적격':      (CRIT_FG, _fill(255, 224, 224)),
        }
        vrd_fg, vrd_bg = vrd_palette.get(final_verdict, (WARN_FG, _fill(255, 248, 225)))
        vrd_border = _bdr('medium', vrd_fg)
        _mg(vrd_r1, 1, 6, f'최종 판정  :  {final_verdict}',
            size=14, bold=True, fg=vrd_fg, bg=vrd_bg, border=vrd_border)
        _mg(vrd_r2, 1, 6,
            (f'총 검사수량 : {total} EA'
             f'        합격 : {pass_cnt} EA'
             f'        양품 : {good_cnt} EA'
             f'        부적격 : {reject_cnt} EA'),
            size=13, bold=True, fg=vrd_fg, bg=vrd_bg, border=vrd_border)

        ws.print_area = f'A1:F{vrd_r2}'

        # ── 저장 ────────────────────────────────────────────────────────
        try:
            wb.save(filepath)
        except PermissionError:
            messagebox.showerror('저장 오류',
                f'파일이 이미 열려 있습니다. Excel을 닫고 다시 시도하세요.\n\n{filepath}')
            return
        except Exception as exc:
            messagebox.showerror('저장 오류',
                f'파일 저장 중 오류가 발생했습니다:\n\n{exc}')
            return

        messagebox.showinfo(
            '검사보고서 저장 완료',
            f'출력 구분 : {label}\n'
            f'총 검사수량 : {total:,} EA\n'
            f'합격        : {pass_cnt:,} EA  ({pass_rate:.1f}%)\n'
            f'양품        : {good_cnt:,} EA\n'
            f'부적격      : {reject_cnt:,} EA\n'
            f'종합판정    : {final_verdict}\n\n'
            f'저장 위치 :\n{filepath}')

    # ── Tab 2: 모델별 조회 (LOT + 작업자 필터) ───────────────────────────────

    def _build_model_stats(self, parent):
        from collections import defaultdict
        c = self._c

        fv_lot = tk.StringVar(value='전체')
        fv_op  = tk.StringVar(value='전체')

        bar = tk.Frame(parent, bg=c['bg'])
        bar.pack(fill='x', padx=6, pady=(6, 2))
        for label, var, values, w in [
            ('LOT:',    fv_lot, self._unique('lot'),      14),
            ('작업자:', fv_op,  self._unique('operator'), 10),
        ]:
            tk.Label(bar, text=label, font=('맑은 고딕', 10),
                     bg=c['bg'], fg=c['txt_g']).pack(side='left', padx=(6, 2))
            cb = ttk.Combobox(bar, textvariable=var, values=values,
                               width=w, state='readonly', font=('맑은 고딕', 10))
            cb.pack(side='left', padx=(0, 6))

        cols   = ('모델', '검사 수', '총 홀막힘', '불량 수', '불량률')
        widths = {'모델': 90, '검사 수': 90, '총 홀막힘': 110, '불량 수': 90, '불량률': 90}
        tv = self._make_treeview(parent, cols, widths, 'ModelStat.Treeview')

        def refresh(*_):
            lo = fv_lot.get(); op = fv_op.get()
            data = [h for h in self._history
                    if (lo == '전체' or h.get('lot', '') == lo)
                    and (op == '전체' or h.get('operator', '') == op)]
            stats = defaultdict(lambda: {'count': 0, 'holes': 0, 'bad': 0})
            for h in data:
                m = h.get('model', '기타')
                stats[m]['count'] += 1
                stats[m]['holes'] += h.get('count', 0)
                if h.get('verdict') == '불량':
                    stats[m]['bad'] += 1
            for row in tv.get_children():
                tv.delete(row)
            for model, d in sorted(stats.items()):
                rate = d['bad'] / d['count'] * 100 if d['count'] else 0
                tag  = 'fail' if rate >= DEFECT_LIMIT else 'ok'
                tv.insert('', 'end', tags=(tag,),
                           values=(model, d['count'], d['holes'], d['bad'], f'{rate:.1f}%'))

        for cb_widget in bar.winfo_children():
            if isinstance(cb_widget, ttk.Combobox):
                cb_widget.bind('<<ComboboxSelected>>', refresh)
        refresh()

    # ── Tab 3: LOT별 조회 (날짜 + 작업자 필터) ───────────────────────────────

    def _build_lot_stats(self, parent):
        from collections import defaultdict
        c = self._c

        fv_date = tk.StringVar(value='전체')
        fv_op   = tk.StringVar(value='전체')

        bar = tk.Frame(parent, bg=c['bg'])
        bar.pack(fill='x', padx=6, pady=(6, 2))
        dates = ['전체'] + sorted(
            {h.get('date', '') for h in self._history if h.get('date', '')}, reverse=True)
        for label, var, values, w in [
            ('날짜:',   fv_date, dates,                   14),
            ('작업자:', fv_op,   self._unique('operator'), 10),
        ]:
            tk.Label(bar, text=label, font=('맑은 고딕', 10),
                     bg=c['bg'], fg=c['txt_g']).pack(side='left', padx=(6, 2))
            cb = ttk.Combobox(bar, textvariable=var, values=values,
                               width=w, state='readonly', font=('맑은 고딕', 10))
            cb.pack(side='left', padx=(0, 6))

        cols   = ('날짜', 'LOT번호', '검사수량', '양품', '불량', '불량률', '모델', '작업자')
        widths = {'날짜': 100, 'LOT번호': 130, '검사수량': 80, '양품': 65,
                  '불량': 65, '불량률': 80, '모델': 100, '작업자': 90}
        tv = self._make_treeview(parent, cols, widths, 'LotStat.Treeview')

        def refresh(*_):
            date_f = fv_date.get(); op_f = fv_op.get()
            for row in tv.get_children():
                tv.delete(row)
            groups = defaultdict(
                lambda: {'total': 0, 'good': 0, 'bad': 0, 'models': set(), 'ops': set()})
            for h in self._history:
                if date_f != '전체' and h.get('date', '') != date_f:
                    continue
                if op_f != '전체' and h.get('operator', '') != op_f:
                    continue
                key = (h.get('date', ''), h.get('lot', '-'))
                groups[key]['total'] += 1
                groups[key]['models'].add(h.get('model', ''))
                groups[key]['ops'].add(h.get('operator', ''))
                if h.get('verdict') == '불량':
                    groups[key]['bad'] += 1
                else:
                    groups[key]['good'] += 1
            for (date, lot), g in sorted(groups.items()):
                rate = g['bad'] / g['total'] * 100 if g['total'] else 0
                tag  = 'fail' if g['bad'] > 0 else 'ok'
                tv.insert('', 'end', tags=(tag,),
                           values=(date, lot, g['total'], g['good'], g['bad'],
                                   f'{rate:.1f}%',
                                   '/'.join(sorted(g['models'])),
                                   '/'.join(sorted(g['ops']))))

        for cb_widget in bar.winfo_children():
            if isinstance(cb_widget, ttk.Combobox):
                cb_widget.bind('<<ComboboxSelected>>', refresh)
        refresh()

    # ── Tab 4: 작업자별 조회 (날짜 + LOT + 모델 필터) ────────────────────────

    def _build_operator_stats(self, parent):
        from collections import defaultdict
        c = self._c

        fv_date  = tk.StringVar(value='전체')
        fv_lot   = tk.StringVar(value='전체')
        fv_model = tk.StringVar(value='전체')

        bar = tk.Frame(parent, bg=c['bg'])
        bar.pack(fill='x', padx=6, pady=(6, 2))
        dates = ['전체'] + sorted(
            {h.get('date', '') for h in self._history if h.get('date', '')}, reverse=True)
        for label, var, values, w in [
            ('날짜:',  fv_date,  dates,                    14),
            ('LOT:',   fv_lot,   self._unique('lot'),       14),
            ('모델:',  fv_model, ['전체'] + list(MODELS.keys()), 10),
        ]:
            tk.Label(bar, text=label, font=('맑은 고딕', 10),
                     bg=c['bg'], fg=c['txt_g']).pack(side='left', padx=(6, 2))
            cb = ttk.Combobox(bar, textvariable=var, values=values,
                               width=w, state='readonly', font=('맑은 고딕', 10))
            cb.pack(side='left', padx=(0, 6))

        cols   = ('작업자', '검사 수', '총 홀막힘', '불량 수', '불량률', '사용 LOT', '사용 모델')
        widths = {'작업자': 100, '검사 수': 80, '총 홀막힘': 100,
                  '불량 수': 80, '불량률': 80, '사용 LOT': 160, '사용 모델': 160}
        tv = self._make_treeview(parent, cols, widths, 'OpStat.Treeview')

        def refresh(*_):
            d_f = fv_date.get(); lo_f = fv_lot.get(); m_f = fv_model.get()
            data = [h for h in self._history
                    if (d_f == '전체' or h.get('date', '') == d_f)
                    and (lo_f == '전체' or h.get('lot', '') == lo_f)
                    and (m_f == '전체' or h.get('model', '') == m_f)]
            stats = defaultdict(
                lambda: {'count': 0, 'holes': 0, 'bad': 0, 'lots': set(), 'models': set()})
            for h in data:
                op = h.get('operator', '') or '(미입력)'
                stats[op]['count']  += 1
                stats[op]['holes']  += h.get('count', 0)
                stats[op]['lots'].add(h.get('lot', ''))
                stats[op]['models'].add(h.get('model', ''))
                if h.get('verdict') == '불량':
                    stats[op]['bad'] += 1
            for row in tv.get_children():
                tv.delete(row)
            for op, d in sorted(stats.items()):
                rate = d['bad'] / d['count'] * 100 if d['count'] else 0
                tag  = 'fail' if rate >= DEFECT_LIMIT else 'ok'
                tv.insert('', 'end', tags=(tag,),
                           values=(op, d['count'], d['holes'], d['bad'],
                                   f'{rate:.1f}%',
                                   ', '.join(sorted(d['lots'])),
                                   ', '.join(sorted(d['models']))))

        for cb_widget in bar.winfo_children():
            if isinstance(cb_widget, ttk.Combobox):
                cb_widget.bind('<<ComboboxSelected>>', refresh)
        refresh()

    # ── Tab 5: 불량률 추이 (모델 + LOT + 작업자 + 기간 필터) ─────────────────

    def _build_trend(self, parent):
        import datetime as _dt
        c = self._c

        ctrl = tk.Frame(parent, bg=c['bg'])
        ctrl.pack(fill='x', padx=8, pady=(6, 2))

        trend_model = tk.StringVar(value='전체')
        trend_lot   = tk.StringVar(value='전체')
        trend_op    = tk.StringVar(value='전체')
        trend_range = tk.StringVar(value='전체')

        for label, var, values, w in [
            ('모델:',   trend_model, ['전체'] + list(MODELS.keys()),   9),
            ('LOT:',    trend_lot,   self._unique('lot'),              12),
            ('작업자:', trend_op,    self._unique('operator'),          9),
            ('기간:',   trend_range, ['전체', '오늘', '최근 7일', '최근 30일', '최근 50건'], 10),
        ]:
            tk.Label(ctrl, text=label, bg=c['bg'], fg=c['txt_g'],
                     font=('맑은 고딕', 9)).pack(side='left', padx=(6, 1))
            cb = ttk.Combobox(ctrl, textvariable=var, values=values,
                               width=w, state='readonly', font=('맑은 고딕', 9))
            cb.pack(side='left', padx=(0, 4))

        lbl_stat = tk.Label(ctrl, text='', bg=c['bg'], fg=c['canvas_avg'], font=('맑은 고딕', 9))
        lbl_stat.pack(side='left', padx=8)

        canvas = tk.Canvas(parent, bg=c['canvas_bg'], highlightthickness=0)
        canvas.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        tip = tk.Label(parent, text='', bg=c['tip_bg'], fg=c['tip_fg'],
                       font=('맑은 고딕', 9), relief='solid', bd=1,
                       highlightbackground=c['tip_border'],
                       justify='left', padx=8, pady=4)

        def get_data():
            data = self._history[:]
            m  = trend_model.get()
            lo = trend_lot.get()
            op = trend_op.get()
            r  = trend_range.get()
            if m  != '전체': data = [h for h in data if h.get('model', '')    == m]
            if lo != '전체': data = [h for h in data if h.get('lot', '')      == lo]
            if op != '전체': data = [h for h in data if h.get('operator', '') == op]
            today = _dt.date.today().strftime('%Y-%m-%d')
            if r == '오늘':
                data = [h for h in data if h.get('date', '') == today]
            elif r == '최근 7일':
                cut = (_dt.date.today() - _dt.timedelta(days=7)).strftime('%Y-%m-%d')
                data = [h for h in data if h.get('date', '') >= cut]
            elif r == '최근 30일':
                cut = (_dt.date.today() - _dt.timedelta(days=30)).strftime('%Y-%m-%d')
                data = [h for h in data if h.get('date', '') >= cut]
            elif r == '최근 50건':
                data = data[-50:]
            return data

        pts_ref = []

        def draw(event=None):
            canvas.delete('all')
            tip.place_forget()
            pts_ref.clear()
            data = get_data()
            n = len(data)
            W = max(canvas.winfo_width(),  500)
            H = max(canvas.winfo_height(), 320)
            ML, MR, MT, MB = 58, 80, 30, 52

            if not data:
                canvas.create_text(W // 2, H // 2, text='데이터 없음',
                                   fill=c['txt_g'], font=('맑은 고딕', 14))
                lbl_stat.configure(text='')
                return

            rates = [h.get('rate', 0.0) for h in data]
            mx = max(max(rates) * 1.25, DEFECT_LIMIT * 2)
            GW = W - ML - MR
            GH = H - MT - MB

            def tx(i): return ML + int(i * GW / max(n - 1, 1)) if n > 1 else ML + GW // 2
            def ty(r): return MT + GH - int(r / mx * GH)

            yl = ty(DEFECT_LIMIT)
            canvas.create_rectangle(ML, MT, ML+GW, yl,    fill=c['canvas_fail_zone'], outline='')
            canvas.create_rectangle(ML, yl, ML+GW, MT+GH, fill=c['canvas_ok_zone'],   outline='')

            for k in range(7):
                rv = mx * k / 6
                yg = ty(rv)
                canvas.create_line(ML, yg, ML+GW, yg, fill=c['canvas_grid'], dash=(3, 5))
                canvas.create_text(ML - 5, yg, text=f'{rv:.1f}%',
                                   fill=c['canvas_label'], font=('맑은 고딕', 8), anchor='e')

            canvas.create_line(ML, yl, ML+GW, yl, fill=c['canvas_limit'], dash=(8, 4), width=1)
            canvas.create_text(ML+GW + 4, yl, text=f'기준\n{DEFECT_LIMIT}%',
                               fill=c['canvas_limit'], font=('맑은 고딕', 8), anchor='w')

            avg = sum(rates) / n
            ya  = ty(avg)
            canvas.create_line(ML, ya, ML+GW, ya, fill=c['canvas_avg'], dash=(5, 3), width=1)
            canvas.create_text(ML + 4, ya - 9, text=f'평균 {avg:.2f}%',
                               fill=c['canvas_avg'], font=('맑은 고딕', 8), anchor='w')

            canvas.create_line(ML, MT, ML, MT+GH, fill=c['canvas_axis'], width=1)
            canvas.create_line(ML, MT+GH, ML+GW, MT+GH, fill=c['canvas_axis'], width=1)

            step = max(1, n // 12)
            for i in range(0, n, step):
                x = tx(i)
                h = data[i]
                label = f"{h.get('date','')[-5:]}\n{h.get('time','')[:5]}"
                canvas.create_line(x, MT+GH, x, MT+GH+4, fill=c['canvas_axis'])
                canvas.create_text(x, MT+GH+6, text=label,
                                   fill=c['canvas_label'], font=('맑은 고딕', 7), anchor='n')

            coords = [(tx(i), ty(r)) for i, r in enumerate(rates)]
            if len(coords) >= 2:
                fill_pts = [ML, MT+GH] + [pt for xy in coords for pt in xy] + [coords[-1][0], MT+GH]
                canvas.create_polygon(fill_pts, fill=c['canvas_fill'], outline='')
                for i in range(len(coords) - 1):
                    col = c['canvas_bad'] if rates[i] >= DEFECT_LIMIT else c['canvas_good']
                    canvas.create_line(coords[i][0], coords[i][1],
                                       coords[i+1][0], coords[i+1][1],
                                       fill=col, width=2)

            r_dot = 4 if n <= 80 else 3
            for i, (x, y) in enumerate(coords):
                col = c['canvas_bad'] if rates[i] >= DEFECT_LIMIT else c['canvas_good']
                canvas.create_oval(x-r_dot, y-r_dot, x+r_dot, y+r_dot,
                                    fill=col, outline=c['tip_border'], width=1)
                pts_ref.append((x, y, i))

            canvas.create_text(14, MT + GH//2, text='불\n량\n률\n(%)',
                               fill=c['canvas_label'], font=('맑은 고딕', 8), anchor='center')
            bad_cnt = sum(1 for r in rates if r >= DEFECT_LIMIT)
            canvas.create_text(ML + GW//2, MT - 12,
                               text=f'불량률 추이  (총 {n}건  |  불량 {bad_cnt}건  |  최대 {max(rates):.2f}%)',
                               fill=c['canvas_title'], font=('맑은 고딕', 10, 'bold'), anchor='s')
            lbl_stat.configure(
                text=f'총 {n}건  평균 {avg:.2f}%  최대 {max(rates):.2f}%  불량 {bad_cnt}건')

        def on_motion(ev):
            if not pts_ref:
                return
            best_i, best_d = -1, float('inf')
            for x, y, idx in pts_ref:
                d = ((ev.x - x)**2 + (ev.y - y)**2) ** 0.5
                if d < best_d:
                    best_d, best_i = d, idx
            if best_i >= 0 and best_d < 24:
                h = get_data()[best_i]
                tip.configure(text=(
                    f"날짜: {h.get('date','-')}  시각: {h.get('time','-')}\n"
                    f"LOT: {h.get('lot','-')}  작업자: {h.get('operator','-')}  SN: {h.get('serial','-')}\n"
                    f"모델: {h.get('model','-')}  홀막힘: {h.get('count',0)}개  "
                    f"불량률: {h.get('rate',0):.2f}%  판정: {h.get('verdict','-')}"
                ))
                cx = canvas.winfo_rootx() - parent.winfo_rootx()
                cy = canvas.winfo_rooty() - parent.winfo_rooty()
                tip.place(x=cx + ev.x + 12, y=cy + ev.y - 60)
                tip.lift()
            else:
                tip.place_forget()

        canvas.bind('<Motion>', on_motion)
        canvas.bind('<Leave>',  lambda e: tip.place_forget())
        canvas.bind('<Configure>', draw)
        for cb_widget in ctrl.winfo_children():
            if isinstance(cb_widget, ttk.Combobox):
                cb_widget.bind('<<ComboboxSelected>>', lambda e: draw())
        canvas.after(120, draw)

    # ── Excel 내보내기 (현재 목록 필터 기준) ─────────────────────────────────

    def _export_excel(self):
        rows = self._filtered_list()
        if not rows:
            messagebox.showinfo('알림', '내보낼 이력이 없습니다.')
            return

        ts       = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = filedialog.asksaveasfilename(
            initialfile=f'홀막힘_이력_{ts}.xlsx',
            initialdir=os.path.expanduser('~/Desktop'),
            defaultextension='.xlsx',
            filetypes=[('Excel 통합문서 (*.xlsx)', '*.xlsx'), ('모든 파일', '*.*')])
        if not filepath:
            return
        if not filepath.lower().endswith('.xlsx'):
            filepath += '.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '검사 이력'
        wb.properties.title   = '멤브레인 홀막힘 검사 이력'
        wb.properties.creator = '홀막힘 카운터 v5.3'

        n_cols     = 11
        thin       = Side(style='thin',   color='C9D4E0')
        med        = Side(style='medium', color='0D2D5E')
        bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_border = Border(left=thin, right=thin, top=med, bottom=med)
        ctr        = Alignment(horizontal='center', vertical='center')
        left_al    = Alignment(horizontal='left',   vertical='center')
        title_fill = PatternFill('solid', fgColor='0A1E35')
        hdr_fill   = PatternFill('solid', fgColor='0D2D5E')
        band_fill  = PatternFill('solid', fgColor='F2F6FB')
        fail_fill  = PatternFill('solid', fgColor='FDEAEA')
        sum_fill   = PatternFill('solid', fgColor='E1ECF7')
        no_fill    = PatternFill(fill_type=None)

        # ── Row 1: 제목 ──────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tcell = ws.cell(1, 1, '멤브레인 홀막힘 검사 이력 리포트')
        tcell.font      = Font(name='맑은 고딕', size=15, bold=True, color='FFFFFF')
        tcell.fill      = title_fill
        tcell.alignment = ctr
        ws.row_dimensions[1].height = 30

        # ── Row 2: 메타 정보(생성일시 · 조회조건 · 총 건수) ───────────
        meta = (f'생성일시: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                f'      모델: {self.fv_model.get()}'
                f'      LOT: {self.fv_lot.get()}'
                f'      작업자: {self.fv_op.get()}'
                f'      판정: {self.fv_verdict.get()}'
                f'      총 {len(rows):,}건')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        mcell = ws.cell(2, 1, meta)
        mcell.font      = Font(name='맑은 고딕', size=9, color='5A6B80')
        mcell.alignment = left_al
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 6   # 여백

        # ── Row 4: 헤더 ──────────────────────────────────────────────
        header_row = 4
        headers = ('번호', '날짜', '시각', 'LOT No.', '작업자', '일련번호',
                   '모델', '홀막힘', '불량률', '판정', '불량사유')
        col_w   = (7, 13, 10, 18, 12, 16, 9, 9, 10, 9, 14)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(header_row, c, h)
            cell.font      = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
            cell.fill      = hdr_fill
            cell.alignment = ctr
            cell.border    = hdr_border
        ws.row_dimensions[header_row].height = 24

        # ── 데이터 행 (줄무늬 배경 + 불량 강조) ────────────────────────
        data_start = header_row + 1
        bad_cnt = 0
        for i, entry in enumerate(rows):
            ri      = data_start + i
            rate    = entry.get('rate', 0)
            verdict = entry.get('verdict', '')
            if verdict == '불량':
                bad_cnt += 1
                fill = fail_fill
            else:
                fill = band_fill if i % 2 == 1 else no_fill
            vals = (i + 1, entry.get('date', ''), entry.get('time', ''),
                     _xl_safe(entry.get('lot', '')), _xl_safe(entry.get('operator', '')),
                     _xl_safe(entry.get('serial', '')), entry.get('model', ''),
                     entry.get('count', ''), f'{rate:.2f}%', verdict,
                     _xl_safe(entry.get('defect_reason', '')))
            for c, v in enumerate(vals, 1):
                is_bad_col = c in (9, 10) and verdict == '불량'
                cell = ws.cell(ri, c, v)
                cell.font      = Font(name='맑은 고딕', size=10,
                                       bold=is_bad_col, color='C00000' if is_bad_col else '1A2530')
                cell.alignment = ctr
                cell.border    = bdr
                cell.fill      = fill
        data_end = data_start + len(rows) - 1

        # ── 요약 통계 행 ─────────────────────────────────────────────
        sum_row  = data_end + 2
        total    = len(rows)
        good_cnt = total - bad_cnt
        bad_rate = round(bad_cnt / total * 100, 1) if total else 0.0
        ws.row_dimensions[sum_row].height = 26
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=n_cols)
        scell = ws.cell(sum_row, 1,
            f'총 검사수량 : {total:,} EA        양호 : {good_cnt:,} EA        '
            f'불량 : {bad_cnt:,} EA        불량률 : {bad_rate}%')
        scell.font      = Font(name='맑은 고딕', size=11, bold=True, color='0D2D5E')
        scell.fill      = sum_fill
        scell.alignment = ctr
        scell.border    = Border(top=med, bottom=med, left=med, right=med)

        # ── 열 너비 자동조정 ─────────────────────────────────────────
        for ci in range(1, n_cols + 1):
            col_letter = ws.cell(header_row, ci).column_letter
            max_len = 0
            for row_cells in ws.iter_rows(min_row=header_row, max_row=data_end,
                                           min_col=ci, max_col=ci):
                for cell in row_cells:
                    if cell.value is not None:
                        txt = str(cell.value)
                        char_len = sum(2 if ord(ch) > 127 else 1 for ch in txt)
                        max_len = max(max_len, char_len)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, col_w[ci-1]), 42)

        # ── 인쇄/탐색 편의 (틀고정, 자동필터, 인쇄 설정) ────────────────
        ws.freeze_panes = f'A{data_start}'
        ws.auto_filter.ref = f'A{header_row}:{ws.cell(header_row, n_cols).column_letter}{data_end}'
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation  = 'landscape'
        ws.page_setup.paperSize    = 9          # A4
        ws.page_setup.fitToWidth   = 1
        ws.page_setup.fitToHeight  = 0
        ws.page_setup.fitToPage    = True
        ws.print_title_rows        = f'{header_row}:{header_row}'

        try:
            wb.save(filepath)
        except PermissionError:
            messagebox.showerror(
                '저장 오류',
                f'파일이 이미 열려 있습니다. Excel을 닫고 다시 시도하세요.\n\n{filepath}')
            return
        except Exception as ex:
            messagebox.showerror('저장 오류', f'파일 저장 중 오류가 발생했습니다:\n\n{ex}')
            return

        messagebox.showinfo('완료', f'이력 내보내기 완료! ({len(rows)}건)\n{filepath}')

    # ── 이력 백업 (수동) ──────────────────────────────────────────────────────

    def _backup_now(self):
        path = _backup_history_db(reason='수동')
        if path:
            messagebox.showinfo(
                '백업 완료',
                '이력 데이터베이스를 백업했습니다.\n'
                '(누를 때마다 새 파일로 저장되며 기존 백업은 덮어쓰지 않습니다)\n\n'
                f'{path}')
        else:
            messagebox.showerror(
                '백업 실패', '백업할 이력 데이터베이스가 없거나 백업 중 오류가 발생했습니다.')

    # ── 백업 복원 ─────────────────────────────────────────────────────────────

    def _restore_from_backup(self):
        if not os.path.isdir(HISTORY_BACKUP_DIR):
            messagebox.showinfo('알림', '백업 파일이 없습니다.')
            return
        files = sorted(
            (f for f in os.listdir(HISTORY_BACKUP_DIR)
             if f.startswith('홀막힘_이력_') and f.endswith('.db')),
            key=lambda f: os.path.getmtime(os.path.join(HISTORY_BACKUP_DIR, f)),
            reverse=True)
        if not files:
            messagebox.showinfo('알림', '백업 파일이 없습니다.')
            return

        c = self._c
        hdr_accent_bg = '#e6e3f7' if self._theme == 'light' else '#2a2a5c'
        hdr_accent_fg = '#2a2360' if self._theme == 'light' else TXT_W

        dlg = tk.Toplevel(self)
        dlg.title('백업에서 복원')
        dlg.configure(bg=c['card'])
        dlg.geometry('560x420')
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self)

        hdr = tk.Frame(dlg, bg=hdr_accent_bg)
        hdr.pack(fill='x')
        tk.Label(hdr, text='  백업 파일에서 이력 복원',
                 font=('맑은 고딕', 13, 'bold'),
                 bg=hdr_accent_bg, fg=hdr_accent_fg).pack(anchor='w', pady=10, padx=10)

        tk.Label(dlg, text='복원할 백업을 선택하세요 (최신순). 복원 전 현재 이력은 자동 백업됩니다.',
                 font=('맑은 고딕', 9), bg=c['card'], fg=c['txt_g'],
                 wraplength=520, justify='left').pack(anchor='w', padx=14, pady=(10, 4))

        list_frame = tk.Frame(dlg, bg=c['card'])
        list_frame.pack(fill='both', expand=True, padx=14, pady=4)
        sb = tk.Scrollbar(list_frame, orient='vertical')
        lb = tk.Listbox(list_frame, font=('맑은 고딕', 10), bg=c['entry_bg'], fg=c['entry_fg'],
                         selectbackground=c['tree_sel'], selectforeground=c['txt'],
                         relief='flat', yscrollcommand=sb.set, activestyle='none')
        sb.configure(command=lb.yview)
        sb.pack(side='right', fill='y')
        lb.pack(side='left', fill='both', expand=True)

        for f in files:
            path  = os.path.join(HISTORY_BACKUP_DIR, f)
            mtime = datetime.datetime.fromtimestamp(os.path.getmtime(path))
            size_kb = os.path.getsize(path) / 1024
            lb.insert('end', f'{mtime.strftime("%Y-%m-%d %H:%M:%S")}   {f}   ({size_kb:.0f}KB)')
        lb.selection_set(0)

        def on_confirm():
            sel = lb.curselection()
            if not sel:
                return
            chosen = files[sel[0]]
            if not messagebox.askyesno(
                    '복원 확인',
                    f'선택한 백업으로 이력을 복원하시겠습니까?\n{chosen}\n\n'
                    '현재 이력은 복원 전 자동으로 백업됩니다.', icon='warning'):
                return
            src = os.path.join(HISTORY_BACKUP_DIR, chosen)
            _backup_history_db(reason='복원전')
            try:
                shutil.copy2(src, HISTORY_DB_FILE)
            except Exception as ex:
                messagebox.showerror('복원 실패', f'백업 파일을 복원하지 못했습니다.\n\n상세 정보: {ex}')
                return
            self._history = self._load()
            self._refresh_list()
            if hasattr(self._parent, '_load_total_products'):
                self._parent._load_total_products()
            if hasattr(self._parent, '_refresh_suggestion_cache'):
                self._parent._refresh_suggestion_cache()
            dlg.destroy()
            messagebox.showinfo('복원 완료', f'이력을 복원했습니다.\n{chosen}')

        btn_f = tk.Frame(dlg, bg=c['card'])
        btn_f.pack(side='bottom', fill='x', padx=14, pady=12)
        tk.Button(btn_f, text='취소', font=('맑은 고딕', 10),
                  bg=c['btn_neutral_bg'], fg=c['txt'], relief='flat', width=10,
                  cursor='hand2', command=dlg.destroy).pack(side='left', padx=4)
        tk.Button(btn_f, text='  이 백업으로 복원', font=('맑은 고딕', 11, 'bold'),
                  bg=hdr_accent_bg, fg=hdr_accent_fg, relief='flat', cursor='hand2',
                  command=on_confirm).pack(side='right', padx=4)

    # ── 이력 전체 삭제 ────────────────────────────────────────────────────────

    def _clear(self):
        if not messagebox.askyesno(
                '전체 삭제',
                '이력을 전체 삭제하시겠습니까?\n삭제 전 현재 이력이 자동으로 백업됩니다.',
                icon='warning'):
            return
        backup_path = _backup_history_db(reason='삭제전')
        try:
            _db_delete_all()
        except Exception as ex:
            messagebox.showerror(
                '삭제 실패',
                f'이력 데이터베이스 삭제 중 오류가 발생했습니다.\n\n상세 정보: {ex}')
            return
        self._history = []
        self._refresh_list()
        if hasattr(self._parent, '_load_total_products'):
            self._parent._load_total_products()
        if backup_path:
            messagebox.showinfo(
                '삭제 완료',
                f'이력이 삭제되었습니다.\n삭제 전 백업 파일: {os.path.basename(backup_path)}')


class HoleCounter(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('멤브레인 홀막힘 카운터  v5.3')
        self.configure(bg=BG)
        self.resizable(True, True)
        self.minsize(900, 600)

        self.selected_model  = tk.StringVar(value='')
        self.fiber_count     = tk.IntVar(value=0)
        self.count           = tk.IntVar(value=0)
        self.lot_no          = tk.StringVar(value='')
        self.operator        = tk.StringVar(value='')
        self.serial_no       = tk.StringVar(value='')   # 제품일련번호
        self._lot_suggestions      = []   # LOT 자동완성 후보 (이력에서 추출)
        self._operator_suggestions = []   # 작업자 자동완성 후보
        self._lot_recent      = []   # LOT 직전 입력 3건 (최신순, 드롭다운용)
        self._serial_recent   = []   # 일련번호 직전 입력 3건
        self._operator_roster = []   # 작업자 고정 목록 (관리 버튼으로 추가/수정/삭제)
        self.count_key        = 'space'   # 홀막힘 카운트 단축키 (변경 가능)
        self.step            = tk.IntVar(value=1)
        self.total_products  = tk.IntVar(value=0)
        self.log_list        = []
        self.flash_active    = False
        self.running         = True
        self._frame          = None
        self._frame_lock     = threading.Lock()
        self._model_selected = False
        self._alarm_done     = False
        self._last_saved_path = None   # 직전 저장 파일 경로
        self._last_saved_history_id = None   # 직전 저장된 이력 DB 행 id (저장취소용)

        self.current_theme  = 'dark'
        self._fullscreen    = False
        self.alarm_enabled  = True
        self.logo_photo     = None
        self._hint_active     = {}   # var id → True if showing hint placeholder
        self._summary_since   = None  # 현황판 리셋 기준 시각
        self._session_log     = []   # 세션 중 완료된 제품 목록 (현황판 전용)
        self.cam_brightness = tk.IntVar(value=0)
        self.cam_contrast   = tk.IntVar(value=100)
        self._log_visible   = False  # 카운팅 로그 기본 접힘

        # ── 치수 측정 ─────────────────────────────────────────
        self._measure_mode    = False
        self._measure_pts     = []   # [(nx, ny), ...] 정규화 좌표 (0~1)
        self._cal_mode        = False
        self._cal_pts         = []
        self._px_per_mm       = 0.0
        self._meas_result_var = None  # _build_dimension_panel 에서 생성
        self._cal_ref_var     = None
        self._cal_status_var  = None
        self._deviation_val   = 0.0  # 마지막 측정된 이탈 거리 (mm)
        self._cal_store       = self._load_cal_file()
        self._focus_score        = 0.0
        self._focus_baseline     = 0.0
        self._px_per_mm_baseline = 0.0
        self._auto_adjust        = False
        self._measured_px_var    = None   # 측정된 px 표시 (UI 생성 후 할당)
        self._cal_result_var     = None   # 계산된 px/mm 표시 (UI 생성 후 할당)
        self._magnifier_on       = False  # 돋보기 ON/OFF
        self._mouse_nx           = -1.0   # 캔버스 마우스 정규화 x (-1=캔버스 밖)
        self._mouse_ny           = -1.0   # 캔버스 마우스 정규화 y
        self._cam_mode           = 'direct'   # 'direct' | 'dino_window'
        self._cam_gen            = 0   # 카메라 재연결 세대 번호 — 이전 _cam_loop 스레드 종료 신호
        self._exposure_locked    = False   # DNX64 SDK로 노출 고정했는지
        self._led_on             = False   # DNX64 SDK로 LED를 켜뒀는지
        # LED "켜짐 유지" 백그라운드 스레드가 보내는 낡은 신호를 무시하기 위한
        # 세대 번호. OFF를 누르면 즉시(블로킹 없이) 번호를 올려서, 스레드가
        # 뒤늦게 보내려는 "켜져라" 신호를 스스로 무시하게 만든다.
        self._led_generation = 0
        self._locked_brightness  = None    # DNX64 하드웨어 밝기 고정값(설정파일에서 로드)
        self._locked_contrast    = None    # DNX64 하드웨어 대비 고정값(설정파일에서 로드)
        self._cutting_side       = None    # 제품 컷팅 이미지 캡처 — 'R' | 'L' | None
        self._cutting_count      = None    # 제품 컷팅 이미지 캡처 — 선택된 컷팅 횟수
        # 제품 컷팅 이미지 캡처 전용 입력란 — 위쪽 검사정보(LOT/일련번호)와 별개로
        # 헷갈리지 않게 독립 관리. 최근 입력값을 드롭다운으로 보여줌.
        self._cut_lot_var    = tk.StringVar(value='')
        self._cut_serial_var = tk.StringVar(value='')
        self._cut_blade_var  = tk.StringVar(value='')
        self._cut_lot_recent    = []
        self._cut_serial_recent = []
        self._cut_blade_recent  = []
        self._btn_cam_mode       = None
        # ── 드래그 캘리브레이션 라인 ──────────────────────────
        self._cal_line_p1   = (0.04, 0.88)   # 정규화 좌표
        self._cal_line_p2   = (0.27, 0.88)
        self._line_dragging = None   # 'p1' | 'p2' | 'line' | None
        self._line_drag_ref = (0.0, 0.0, 0.0, 0.0)  # (mx,my,p1x,p1y) or (mx,my,midx,midy)
        self._line_active   = None   # 방향키 미세조정 대상 ('p1' | 'p2' | 'line' | None)

        _migrate_json_to_db_if_needed()
        # 이력 DB 자동 백업 — 지금까지는 사용자가 수동으로 "백업" 버튼을 누를
        # 때만 백업이 생겼음. 앱을 켤 때마다 자동으로 한 번 백업해둬서, 파일이
        # 사라지거나 손상돼도 최소한 "지난 실행 시점"으로는 복구할 수 있게 함.
        # 실패해도(디스크 오류 등) 조용히 넘어감 — 시작 자체를 막으면 안 되므로.
        _backup_history_db(reason='자동_시작시')
        self._load_total_products()
        self._load_config()
        self._refresh_suggestion_cache()
        self._build_ui()
        self._bind_keys()
        self.lot_no.trace_add('write',    lambda *_: self._save_config())
        self.operator.trace_add('write',  lambda *_: self._save_config())
        self.serial_no.trace_add('write', lambda *_: self._save_config())
        self._start_camera()
        self._update_preview()
        threading.Thread(target=self._apply_locked_quality, daemon=True).start()
        self.protocol('WM_DELETE_WINDOW', self._on_close)

    # ── UI 구성 ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── 헤더 ─────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg='#1a3a6b', height=52)
        hdr.pack(fill='x', side='top')
        try:
            if os.path.exists(LOGO_PATH):
                _li = Image.open(LOGO_PATH)
                _lh = 38
                _lw = int(_li.width * _lh / _li.height)
                _li = _li.resize((_lw, _lh), Image.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(_li)
                tk.Label(hdr, image=self.logo_photo, bg='#1a3a6b'
                         ).pack(side='left', padx=10, pady=7)
        except Exception:
            pass
        tk.Label(hdr, text='  멤브레인 홀막힘 카운팅 시스템  v5.3  |  불량 기준 4%',
                 font=('맑은 고딕', 13, 'bold'), bg='#1a3a6b', fg=TXT_W
                 ).pack(side='left', pady=10)

        self.lbl_time = tk.Label(hdr, text='', font=('맑은 고딕', 10),
                                  bg='#1a3a6b', fg='#aaccee')
        self.lbl_time.pack(side='right', padx=12)
        self._tick_clock()

        self.btn_fs = tk.Button(
            hdr, text=' 전체화면', font=('맑은 고딕', 10),
            relief='flat', cursor='hand2',
            command=self._toggle_fullscreen)
        _setup_hover(self.btn_fs, '#0d2d5e', '#1a4a8a')
        self.btn_fs.pack(side='right', padx=4, pady=8)

        self.btn_theme = tk.Button(
            hdr, text=' 라이트 ', font=('맑은 고딕', 10),
            relief='flat', cursor='hand2',
            command=self._toggle_theme)
        _setup_hover(self.btn_theme, '#0d2d5e', '#1a4a8a')
        self.btn_theme.pack(side='right', padx=4, pady=8)

        self.btn_alarm = tk.Button(
            hdr, text=' 알람  ', font=('맑은 고딕', 10),
            relief='flat', cursor='hand2',
            command=self._toggle_alarm)
        _setup_hover(self.btn_alarm, '#0d2d5e', '#1a4a8a')
        self.btn_alarm.pack(side='right', padx=4, pady=8)

        # 모델 드롭다운 (헤더 우측)
        self._build_model_dropdown(hdr)

        # ── 하단 고정: 로그 토글 바 + 로그 패널 컨테이너 ─────────────────────
        self._bottom_container = tk.Frame(self, bg='#21262d')
        self._bottom_container.pack(fill='x', side='bottom')

        log_toggle_bar = tk.Frame(self._bottom_container, bg='#21262d', height=24)
        log_toggle_bar.pack(fill='x', side='bottom')
        self.btn_log_toggle = tk.Button(
            log_toggle_bar,
            text='▶  카운팅 로그  (클릭하여 펼치기)',
            font=('맑은 고딕', 8, 'bold'),
            bg='#21262d', fg=TXT_G,
            relief='flat', cursor='hand2',
            command=self._toggle_log)
        self.btn_log_toggle.pack(fill='x', ipady=2)

        self._log_frame = tk.Frame(self._bottom_container, bg=BG, height=130)
        self._log_frame.pack_propagate(False)
        self._build_log_panel(self._log_frame)
        # 기본 접힘 — 처음엔 pack 하지 않음

        # ── 메인 영역: 가로 PanedWindow (카메라 좌 | 컨트롤 우) ──────────────
        main = tk.Frame(self, bg=BG)
        main.pack(fill='both', expand=True, padx=4, pady=(4, 0))

        self.main_paned = tk.PanedWindow(main, orient='horizontal',
                                          bg='#2c3e50', sashwidth=5,
                                          sashrelief='raised', sashpad=2)
        self.main_paned.pack(fill='both', expand=True)

        # ── 좌측: 라이브 뷰 (70~80%) ─────────────────────────────────────────
        cam_frame = tk.Frame(self.main_paned, bg=BG)
        self.main_paned.add(cam_frame, minsize=500)
        self._build_camera_panel(cam_frame)

        # ── 우측: 컨트롤 패널 (스크롤 가능) ─────────────────────────────────
        ctrl_wrap = tk.Frame(self.main_paned, bg=BG)
        self.main_paned.add(ctrl_wrap, minsize=260, width=310)

        self._ctrl_canvas = tk.Canvas(ctrl_wrap, bg=BG, highlightthickness=0)
        ctrl_sb = tk.Scrollbar(ctrl_wrap, orient='vertical', command=self._ctrl_canvas.yview)
        self._ctrl_canvas.configure(yscrollcommand=ctrl_sb.set)
        self._ctrl_canvas.pack(side='left', fill='both', expand=True)
        ctrl_sb.pack(side='right', fill='y')

        ctrl_inner = tk.Frame(self._ctrl_canvas, bg=BG)
        self._ctrl_win_id = self._ctrl_canvas.create_window((0, 0), window=ctrl_inner, anchor='nw')
        ctrl_inner.bind('<Configure>',
                        lambda e: self._ctrl_canvas.configure(scrollregion=self._ctrl_canvas.bbox('all')))
        self._ctrl_canvas.bind('<Configure>',
                         lambda e: self._ctrl_canvas.itemconfig(self._ctrl_win_id, width=e.width))
        self._ctrl_canvas.bind('<MouseWheel>', self._ctrl_scroll)

        self._build_control_panel(ctrl_inner)
        self._build_dimension_panel(ctrl_inner)
        self._build_summary_panel(ctrl_inner)
        self._bind_ctrl_scroll(ctrl_inner)

        # 초기 sash: 우측 패널 310px, 나머지는 카메라
        self.after(100, lambda: self.main_paned.sash_place(0,
            max(500, self.winfo_width() - 314), 0))
        self.after(180, self._apply_icons)

    def _build_model_dropdown(self, parent):
        """헤더에 모델 선택 드롭다운을 배치한다.
        설명 문구 없이, 드롭다운 자체를 노란 테두리로 감싸 바로 눈에 띄게 하고
        선택 후에는 초록 테두리로 바뀌어 완료 상태를 알려준다."""
        self._model_dd_wrap = tk.Frame(parent, bg=ACC_YEL, padx=2, pady=2)
        self._model_dd_wrap.pack(side='right', padx=(0, 4), pady=6)

        s = ttk.Style()
        s.theme_use('clam')
        s.configure('ModelHeader.TCombobox',
                     fieldbackground='#ffffff', background='#ffffff',
                     foreground='#0d1e2e', arrowsize=20,
                     padding=(8, 6))
        s.map('ModelHeader.TCombobox',
               fieldbackground=[('readonly', '#ffffff')],
               foreground=[('readonly', '#0d1e2e')])

        self._model_dd_var = tk.StringVar(value='모델 선택')
        model_cb = ttk.Combobox(
            self._model_dd_wrap,
            textvariable=self._model_dd_var,
            values=list(MODELS.keys()),
            width=13, state='readonly',
            style='ModelHeader.TCombobox',
            font=('맑은 고딕', 13, 'bold'))
        model_cb.pack()
        model_cb.bind('<<ComboboxSelected>>',
                      lambda e: self._select_model(self._model_dd_var.get()))

    def _build_camera_panel(self, parent):
        frame = tk.Frame(parent, bg=CARD_BG)
        frame.pack(fill='both', expand=True)

        top = tk.Frame(frame, bg='#1a3a6b')
        top.pack(fill='x')
        tk.Label(top, text='  현미경 라이브 뷰', font=('맑은 고딕', 11, 'bold'),
                 bg='#1a3a6b', fg=TXT_W).pack(side='left', padx=6, pady=4)
        self.lbl_cam_status = tk.Label(top, text='● 연결 중...',
                                        font=('맑은 고딕', 10), bg='#1a3a6b', fg=ACC_YEL)
        self.lbl_cam_status.pack(side='right', padx=10)

        # 하단 컨트롤 — DNX64 SDK 버튼 전용 줄 (먼저 pack해야 맨 아래에 위치)
        bot2 = tk.Frame(frame, bg='#111820')
        bot2.pack(side='bottom', fill='x', padx=4, pady=(0, 2))

        self._btn_exposure_lock = tk.Button(
            bot2, text='🔒 노출 고정',
            font=('맑은 고딕', 8), bg='#21262d', fg=TXT_G,
            relief='flat', cursor='hand2', padx=6, pady=1,
            command=self._toggle_exposure_lock)
        self._btn_exposure_lock.pack(side='left', padx=4)

        self._btn_led = tk.Button(
            bot2, text='💡 LED',
            font=('맑은 고딕', 8), bg='#21262d', fg=TXT_G,
            relief='flat', cursor='hand2', padx=6, pady=1,
            command=self._toggle_led)
        self._btn_led.pack(side='left', padx=4)

        self._btn_quality_lock = tk.Button(
            bot2, text='🎨 화질 고정',
            font=('맑은 고딕', 8), bg='#21262d', fg=TXT_G,
            relief='flat', cursor='hand2', padx=6, pady=1,
            command=self._lock_quality_now)
        self._btn_quality_lock.pack(side='left', padx=4)

        # 하단 컨트롤 — 한 줄로 압축 (bottom 먼저 배치해야 캔버스가 나머지 차지)
        bot = tk.Frame(frame, bg='#111820')
        bot.pack(side='bottom', fill='x', padx=4, pady=2)

        # 카메라 포트
        tk.Label(bot, text='포트:', font=('맑은 고딕', 8),
                 bg='#111820', fg='#ffffff').pack(side='left', padx=(4, 0))
        self.cam_idx = tk.IntVar(value=0)
        for i in range(4):
            tk.Radiobutton(bot, text=str(i), variable=self.cam_idx, value=i,
                           command=self._reconnect_cam,
                           font=('맑은 고딕', 8), bg='#111820', fg=TXT_W,
                           selectcolor='#2c5f9e',
                           activebackground='#111820').pack(side='left', padx=2)

        tk.Label(bot, text='│', bg='#111820', fg='#444').pack(side='left', padx=4)

        self._btn_cam_mode = tk.Button(
            bot, text='🖥 DinoCapture 창 캡처',
            font=('맑은 고딕', 8), bg='#21262d', fg=TXT_G,
            relief='flat', cursor='hand2', padx=6, pady=1,
            command=self._toggle_cam_mode)
        self._btn_cam_mode.pack(side='left', padx=4)

        tk.Label(bot, text='│', bg='#111820', fg='#444').pack(side='left', padx=4)

        # 밝기/대비 슬라이더 — 가로로 한 줄
        for lbl, var, lo, hi, dflt in [
            ('밝기', self.cam_brightness, -100, 100,   0),
            ('대비', self.cam_contrast,    50,  200, 100),
        ]:
            tk.Label(bot, text=lbl, font=('맑은 고딕', 8),
                     bg='#111820', fg='#ffffff').pack(side='left')
            tk.Scale(bot, from_=lo, to=hi, orient='horizontal', variable=var,
                     bg='#111820', fg=TXT_W, troughcolor='#21262d',
                     highlightthickness=0, showvalue=False, length=90,
                     width=8, sliderlength=12,
                     ).pack(side='left', padx=2)
            tk.Button(bot, text='↺', font=('맑은 고딕', 7),
                      bg='#21262d', fg=TXT_W, relief='flat', cursor='hand2',
                      padx=2, pady=0,
                      command=lambda v=var, d=dflt: v.set(d)
                      ).pack(side='left', padx=(0, 6))

        # 캡처/저장 버튼
        self._cam_save_btn = tk.Button(bot, text='  저장', font=('맑은 고딕', 9),
                               bg='#1a5c2e', fg='#fff', activebackground=ACC_GRN,
                               relief='flat', cursor='hand2', padx=6, pady=1,
                               compound='left',
                               command=self._save_excel)
        self._cam_save_btn.pack(side='right', padx=2)
        self._cap_btn = tk.Button(bot, text='  화면 캡처', font=('맑은 고딕', 9),
                             bg='#1a3a5c', fg='#fff', activebackground='#2c5f9e',
                             relief='flat', cursor='hand2', padx=6, pady=1,
                             compound='left',
                             command=self._capture_screen)
        self._cap_btn.pack(side='right', padx=2)

        # 캔버스: 나머지 공간 전체 차지
        self.canvas = tk.Canvas(frame, bg='#000', cursor='crosshair', highlightthickness=0)
        self.canvas.pack(fill='both', expand=True)
        self.canvas.bind('<ButtonPress-1>',   self._on_canvas_press)
        self.canvas.bind('<B1-Motion>',       self._on_canvas_drag)
        self.canvas.bind('<ButtonRelease-1>', self._on_canvas_release)
        self.canvas.bind('<Motion>',          self._on_canvas_hover)
        self.canvas.bind('<Leave>',
                         lambda e: setattr(self, '_mouse_nx', -1.0))
        # 캘리브레이션 라인 방향키 미세조정 (1px, Shift+방향키 5px)
        self.canvas.bind('<Left>',        lambda e: self._nudge_line(-1, 0))
        self.canvas.bind('<Right>',       lambda e: self._nudge_line(1, 0))
        self.canvas.bind('<Up>',          lambda e: self._nudge_line(0, -1))
        self.canvas.bind('<Down>',        lambda e: self._nudge_line(0, 1))
        self.canvas.bind('<Shift-Left>',  lambda e: self._nudge_line(-5, 0))
        self.canvas.bind('<Shift-Right>', lambda e: self._nudge_line(5, 0))
        self.canvas.bind('<Shift-Up>',    lambda e: self._nudge_line(0, -5))
        self.canvas.bind('<Shift-Down>',  lambda e: self._nudge_line(0, 5))

    def _build_control_panel(self, parent):
        # ── 현재 모델 표시 ────────────────────────────────────────────────────
        mf = tk.Frame(parent, bg='#111820',
                       highlightbackground='#2c5f9e', highlightthickness=1)
        mf.pack(fill='x', pady=(0, 2))
        self.lbl_model_info = tk.Label(
            mf, text='← 모델을 선택하세요',
            font=('맑은 고딕', 8), bg='#111820', fg=ACC_YEL,
            wraplength=290, justify='left')
        self.lbl_model_info.pack(anchor='w', padx=8, pady=4)

        # ── 검사 정보 ─────────────────────────────────────────────────────────
        info = tk.Frame(parent, bg=CARD_BG,
                         highlightbackground='#30363d', highlightthickness=1)
        info.pack(fill='x')

        tk.Label(info, text='검사 정보', font=('맑은 고딕', 9, 'bold'),
                 bg=CARD_BG, fg='#ffffff').pack(anchor='w', padx=8, pady=(3, 2))

        grid = tk.Frame(info, bg=CARD_BG)
        grid.pack(fill='x', padx=6, pady=(0, 4))
        grid.columnconfigure(1, weight=1)
        grid.columnconfigure(3, weight=1)

        fields = [
            ('LOT No.',    self.lot_no,   'LOT-2026-001', 0, 0),
            ('작업자',      self.operator, '작업자 이름',   0, 2),
            ('일련번호',    self.serial_no, 'SN-001',      1, 0),
        ]
        _kind_map = {'LOT No.': 'lot', '작업자': 'operator', '일련번호': 'serial'}
        self._info_entries = []   # 테마 전환 시 Entry 색상 갱신용
        for lbl, var, ph, r, c in fields:
            if lbl == '작업자':
                lf = tk.Frame(grid, bg=CARD_BG)
                lf.grid(row=r, column=c, sticky='w', padx=(3 if c > 0 else 0, 2), pady=1)
                tk.Label(lf, text=lbl, font=('맑은 고딕', 8),
                         bg=CARD_BG, fg='#c8ddf0').pack(side='left')
                btn_roster = tk.Button(
                    lf, text='관리', font=('맑은 고딕', 7),
                    bg='#21262d', fg='#8fb8e0', relief='flat', cursor='hand2',
                    padx=3, pady=0, command=self._open_operator_roster_dialog)
                btn_roster.pack(side='left', padx=(4, 0))
                _setup_hover(btn_roster, '#21262d', '#2c5f9e')
                self._btn_roster = btn_roster
            else:
                tk.Label(grid, text=lbl, font=('맑은 고딕', 8),
                         bg=CARD_BG, fg='#c8ddf0', anchor='w'
                         ).grid(row=r, column=c, sticky='w', padx=(3 if c > 0 else 0, 2), pady=1)
            has_real = bool(var.get())
            e = tk.Entry(grid, textvariable=var,
                         font=('맑은 고딕', 9), bg='#21262d',
                         fg=TXT_W if has_real else '#666666',
                         insertbackground=TXT_W, relief='flat', bd=2,
                         justify='left',
                         highlightthickness=2, highlightbackground='#30363d',
                         highlightcolor=ACC_BLU)
            if r == 1:
                e.grid(row=r, column=c+1, columnspan=3, sticky='ew', pady=1)
            else:
                e.grid(row=r, column=c+1, sticky='ew', pady=1)
            if not has_real:
                var.set(ph)
                self._hint_active[id(var)] = True
            self._info_entries.append((e, var, ph))
            _kind = _kind_map.get(lbl)
            if _kind:
                self._setup_autocomplete(e, var, _kind)

            # 카운트 키(기본 SPACE, 변경 가능)는 검사 정보란(LOT/작업자/일련번호)에
            # 커서가 있어도 그 문자가 입력되지 않고 바로 카운트되도록 _bind_keys() ~
            # _apply_count_key() 에서 이 위젯들(self._info_entries)에 가로채기를 건다.

            def on_focus_in(ev, _e=e, _v=var):
                t = getattr(self, 'current_theme', 'dark')
                real_fg = TXT_W if t == 'dark' else LT_TXT
                real_bg = '#21262d' if t == 'dark' else '#ffffff'
                if self._hint_active.get(id(_v)):
                    _v.set('')
                    self._hint_active[id(_v)] = False
                _e.configure(fg=real_fg, bg=real_bg, insertbackground=real_fg)

            def on_focus_out(ev, _e=e, _v=var, _p=ph, _k=_kind):
                t = getattr(self, 'current_theme', 'dark')
                hint_fg = '#888888' if t == 'dark' else '#7a8fa6'
                real_fg = TXT_W if t == 'dark' else LT_TXT
                real_bg = '#21262d' if t == 'dark' else '#ffffff'
                if not _v.get() or self._hint_active.get(id(_v)):
                    _v.set(_p)
                    _e.configure(fg=hint_fg, bg=real_bg)
                    self._hint_active[id(_v)] = True
                else:
                    _e.configure(fg=real_fg, bg=real_bg)
                    if _k:
                        self._push_recent(_k, _v.get())

            e.bind('<FocusIn>',  on_focus_in, add='+')
            e.bind('<FocusOut>', on_focus_out, add='+')

        # ── 총 검사 제품 수 ────────────────────────────────────────────────────
        pr = tk.Frame(parent, bg='#0a2010',
                       highlightbackground='#2e8b57', highlightthickness=1)
        pr.pack(fill='x', padx=8, pady=(3, 4))
        tk.Label(pr, text='  총 검사 제품 수', font=('맑은 고딕', 9, 'bold'),
                 bg='#0a2010', fg='#4caf50').pack(side='left', pady=3)
        tk.Label(pr, textvariable=self.total_products,
                 font=('맑은 고딕', 13, 'bold'), bg='#0a2010', fg='#4caf50'
                 ).pack(side='left', padx=3)
        tk.Label(pr, text='개', font=('맑은 고딕', 9),
                 bg='#0a2010', fg='#4caf50').pack(side='left')
        self.btn_history = tk.Button(pr, text='이력 조회 →', font=('맑은 고딕', 9, 'bold'),
                  bg='#1a5c2e', fg='#fff', relief='flat', cursor='hand2',
                  command=self._open_history)
        self.btn_history.pack(side='right', padx=4, pady=3)

        # ── 카운트 표시 ───────────────────────────────────────────────────────
        cnt = tk.Frame(parent, bg=CARD_BG,
                        highlightbackground='#30363d', highlightthickness=1)
        cnt.pack(fill='x', pady=(4, 0))
        tk.Label(cnt, text='홀막힘 수량', font=('맑은 고딕', 9, 'bold'),
                 bg=CARD_BG, fg='#ffffff').pack(pady=(3, 0))

        self.lbl_count = tk.Label(cnt, textvariable=self.count,
                                   font=('맑은 고딕', 44, 'bold'),
                                   bg=CARD_BG, fg=ACC_RED)
        self.lbl_count.pack()

        self.lbl_rate_big = tk.Label(cnt, text='불량률: —',
                                      font=('맑은 고딕', 14, 'bold'),
                                      bg=CARD_BG, fg=TXT_G)
        self.lbl_rate_big.pack(pady=(0, 1))

        self.lbl_verdict = tk.Label(cnt, text='',
                                     font=('맑은 고딕', 12, 'bold'),
                                     bg=CARD_BG, fg=TXT_G)
        self.lbl_verdict.pack(pady=(0, 2))

        self.lbl_fiber_info = tk.Label(cnt, text='모델을 먼저 선택하세요',
                                        font=('맑은 고딕', 9),
                                        bg=CARD_BG, fg=TXT_W,
                                        wraplength=340)
        self.lbl_fiber_info.pack(pady=(0, 1))

        self.lbl_flash = tk.Label(cnt, text='',
                                   font=('맑은 고딕', 10, 'bold'),
                                   bg=CARD_BG, fg=ACC_YEL,
                                   wraplength=340)
        self.lbl_flash.pack(pady=(0, 4))

        # ── 카운팅 단위 ───────────────────────────────────────────────────────
        sc = tk.Frame(parent, bg=CARD_BG,
                       highlightbackground='#2c5f9e', highlightthickness=1)
        sc.pack(fill='x', pady=(4, 0))
        tk.Label(sc, text='카운팅 단위  ( 1=1개  2=10개  3=20개 )',
                 font=('맑은 고딕', 8, 'bold'), bg=CARD_BG, fg='#ffffff'
                 ).pack(anchor='w', padx=8, pady=(4, 2))

        sr = tk.Frame(sc, bg=CARD_BG)
        sr.pack(fill='x', padx=6, pady=(0, 6))
        self.step_btns = {}
        for val, label in ((1, '1개씩'), (10, '10개씩'), (20, '20개씩')):
            b = tk.Button(sr, text=label, font=('맑은 고딕', 10, 'bold'),
                          bg='#2c5f9e' if val == 1 else '#21262d', fg='#fff',
                          relief='flat', cursor='hand2', pady=4,
                          command=lambda v=val: self._set_step(v))
            b.pack(side='left', expand=True, fill='x', padx=2)
            self.step_btns[val] = b

        # ── 마우스 클릭 조작 ──────────────────────────────────────────────────
        cc = tk.Frame(parent, bg=CARD_BG,
                       highlightbackground='#30363d', highlightthickness=1)
        cc.pack(fill='x', pady=(4, 0))
        tk.Label(cc, text='마우스 클릭 조작', font=('맑은 고딕', 9, 'bold'),
                 bg=CARD_BG, fg='#ffffff').pack(anchor='w', padx=8, pady=(4, 2))

        self.btn_click_count = tk.Button(
            cc,
            text=f'● 홀막힘  카운트\n( 클릭  또는  {self._count_key_label(self.count_key)} )',
            font=('맑은 고딕', 11, 'bold'),
            bg=ACC_RED, fg='#fff',
            activebackground='#ff6666',
            relief='flat', cursor='hand2', pady=8,
            command=self._count_hole)
        self.btn_click_count.pack(fill='x', padx=8, pady=(0, 4))
        self.btn_click_count.bind('<ButtonPress-1>',   lambda e: self.btn_click_count.configure(bg='#ff6666'))
        self.btn_click_count.bind('<ButtonRelease-1>', lambda e: self.btn_click_count.configure(bg=ACC_RED))

        key_row = tk.Frame(cc, bg=CARD_BG)
        key_row.pack(fill='x', padx=8, pady=(0, 4))
        self.lbl_count_key = tk.Label(
            key_row, text=f'카운트 키: {self._count_key_label(self.count_key)}',
            font=('맑은 고딕', 8), bg=CARD_BG, fg=TXT_G)
        self.lbl_count_key.pack(side='left')
        btn_key_change = tk.Button(
            key_row, text='변경', font=('맑은 고딕', 8),
            bg='#21262d', fg='#8fb8e0', relief='flat', cursor='hand2',
            padx=6, pady=0, command=self._open_count_key_dialog)
        btn_key_change.pack(side='right')
        _setup_hover(btn_key_change, '#21262d', '#2c5f9e')

        r2 = tk.Frame(cc, bg=CARD_BG)
        r2.pack(fill='x', padx=8, pady=(0, 4))

        self.btn_undo = tk.Button(r2,
                                   text='↩  마지막 취소\n( Z 키 )',
                                   font=('맑은 고딕', 10, 'bold'),
                                   bg='#7a5c00', fg='#fff',
                                   activebackground=ACC_YEL,
                                   relief='flat', cursor='hand2', pady=6,
                                   command=self._undo)
        self.btn_undo.pack(side='left', expand=True, fill='x', padx=(0, 3))
        self.btn_undo.bind('<ButtonPress-1>',   lambda e: self.btn_undo.configure(bg=ACC_YEL))
        self.btn_undo.bind('<ButtonRelease-1>', lambda e: self.btn_undo.configure(bg='#7a5c00'))

        self.btn_save = tk.Button(r2,
                                   text='💾  결과 저장\n( S 키 )',
                                   font=('맑은 고딕', 10, 'bold'),
                                   bg='#1a5c2e', fg='#fff',
                                   activebackground=ACC_GRN,
                                   relief='flat', cursor='hand2', pady=6,
                                   command=self._save_excel)
        self.btn_save.pack(side='left', expand=True, fill='x', padx=(3, 0))
        self.btn_save.bind('<ButtonPress-1>',   lambda e: self.btn_save.configure(bg=ACC_GRN))
        self.btn_save.bind('<ButtonRelease-1>', lambda e: self.btn_save.configure(bg='#1a5c2e'))

        # 다음 제품 버튼
        next_row = tk.Frame(cc, bg=CARD_BG)
        next_row.pack(fill='x', padx=8, pady=(0, 3))

        self.btn_next = tk.Button(next_row,
                             text='▶  다음 제품  ( 초기화 )',
                             font=('맑은 고딕', 10, 'bold'),
                             bg='#1a3a6b', fg='#fff',
                             activebackground='#2c5f9e',
                             relief='flat', cursor='hand2',
                             command=self._next_product)
        self.btn_next.pack(side='left', fill='x', expand=True, ipady=3)
        self.btn_next.bind('<ButtonPress-1>',   lambda e: self.btn_next.configure(bg='#2c5f9e'))
        self.btn_next.bind('<ButtonRelease-1>', lambda e: self.btn_next.configure(bg='#1a3a6b'))

        self.btn_reset = tk.Button(cc, text='전체 초기화  ( R 키 )',
                  font=('맑은 고딕', 9),
                  bg='#2a2a2a', fg=TXT_W,
                  relief='flat', cursor='hand2', pady=4,
                  command=self._reset)
        self.btn_reset.pack(fill='x', padx=8, pady=(0, 6))

        # ── 키보드 단축키 ─────────────────────────────────────────────────────
        kc = tk.Frame(parent, bg=CARD_BG,
                       highlightbackground='#30363d', highlightthickness=1)
        kc.pack(fill='x', pady=(4, 0))
        tk.Label(kc, text='키보드 단축키', font=('맑은 고딕', 9, 'bold'),
                 bg=CARD_BG, fg='#ffffff').pack(anchor='w', padx=8, pady=(4, 2))

        shortcuts = [
            ('count_key', self._count_key_label(getattr(self, 'count_key', 'space')),
                           '홀막힘 카운트',        ACC_RED),
            ('1/2/3',      '1/2/3',   '1개/10개/20개 단위',   '#4a9fd4'),
            ('Z',          'Z',       '마지막 취소',            ACC_YEL),
            ('S',          'S',       '저장',                  ACC_GRN),
            ('R',          'R',       '초기화',                TXT_W),
        ]
        for tag, key, desc, col in shortcuts:
            row = tk.Frame(kc, bg=CARD_BG)
            row.pack(fill='x', padx=6, pady=0)
            lbl_key = tk.Label(row, text=key, font=('Consolas', 8, 'bold'),
                     bg='#21262d', fg=col,
                     relief='solid', bd=1, padx=3, pady=1, width=10
                     )
            lbl_key.pack(side='left')
            if tag == 'count_key':
                self._lbl_shortcut_key = lbl_key
            tk.Label(row, text=desc, font=('맑은 고딕', 8),
                     bg=CARD_BG, fg='#ffffff').pack(side='left', padx=5)

        tk.Frame(kc, bg=CARD_BG, height=3).pack()

    def _toggle_log(self):
        if self._log_visible:
            self._log_frame.pack_forget()
            self.btn_log_toggle.configure(text='▶  카운팅 로그  (클릭하여 펼치기)')
        else:
            self._log_frame.pack(fill='x', side='bottom',
                                  in_=self._bottom_container)
            self.btn_log_toggle.configure(text='▼  카운팅 로그  (클릭하여 접기)')
        self._log_visible = not self._log_visible

    def _build_log_panel(self, parent=None):
        if parent is None:
            parent = self
        lf = tk.Frame(parent, bg=CARD_BG,
                       highlightbackground='#30363d', highlightthickness=1)
        lf.pack(fill='both', expand=True, padx=4, pady=(2, 2))

        hdr = tk.Frame(lf, bg='#1a3a6b')
        hdr.pack(fill='x')
        tk.Label(hdr, text='  카운팅 로그', font=('맑은 고딕', 10, 'bold'),
                 bg='#1a3a6b', fg=TXT_W).pack(side='left', padx=8, pady=4)

        cols = ('번호', '시각', '모델', '단위', '홀막힘 수', '불량률', '판정')
        ls = ttk.Style()
        ls.configure('Log.Treeview',
                      background='#ffffff', foreground='#111111',
                      rowheight=22, fieldbackground='#ffffff',
                      font=('맑은 고딕', 10))
        ls.configure('Log.Treeview.Heading',
                      background='#dde3ea', foreground='#1a3a6b',
                      font=('맑은 고딕', 10, 'bold'))
        ls.map('Log.Treeview', background=[('selected', '#b0c8e8')])
        self.tree = ttk.Treeview(lf, columns=cols, show='headings',
                                  style='Log.Treeview')

        widths = {'번호': 45, '시각': 90, '모델': 65, '단위': 65,
                  '홀막힘 수': 80, '불량률': 80, '판정': 65}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=widths[c], anchor='center')

        self.tree.tag_configure('fail', background='#ffe0e0', foreground='#cc0000')
        self.tree.tag_configure('warn', background='#fff8dc', foreground='#8a6000')
        self.tree.tag_configure('ok',   background='#e8f5e9', foreground='#1a5c2e')

        log_sb = tk.Scrollbar(lf, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=log_sb.set)
        log_sb.pack(side='right', fill='y')
        self.tree.pack(fill='both', expand=True)

    def _build_summary_panel(self, parent):
        sf = tk.Frame(parent, bg=CARD_BG,
                      highlightbackground='#2c5f9e', highlightthickness=1)
        sf.pack(fill='x', padx=0, pady=(6, 4))

        hdr = tk.Frame(sf, bg='#0d2d5e')
        hdr.pack(fill='x')
        tk.Label(hdr, text='  📊 금일 LOT별 종합 현황',
                 font=('맑은 고딕', 10, 'bold'),
                 bg='#0d2d5e', fg=TXT_W).pack(side='left', padx=8, pady=4)
        tk.Button(hdr, text='🗑 초기화',
                  font=('맑은 고딕', 9), bg='#5c1a1a', fg=TXT_W,
                  relief='flat', cursor='hand2',
                  command=self._reset_summary).pack(side='right', padx=4, pady=3)
        tk.Button(hdr, text='↺ 새로고침',
                  font=('맑은 고딕', 9), bg='#1a3a6b', fg=TXT_W,
                  relief='flat', cursor='hand2',
                  command=self._refresh_summary).pack(side='right', padx=4, pady=3)

        cols = ('날짜', 'LOT번호', '일련번호', '홀막힘수', '양품', '불량', '불량률')
        ss = ttk.Style()
        ss.configure('Sum.Treeview',
                      background='#ffffff', foreground='#111111',
                      rowheight=22, fieldbackground='#ffffff',
                      font=('맑은 고딕', 10))
        ss.configure('Sum.Treeview.Heading',
                      background='#dde3ea', foreground='#1a3a6b',
                      font=('맑은 고딕', 10, 'bold'))
        ss.map('Sum.Treeview', background=[('selected', '#b0c8e8')])
        self.summary_tree = ttk.Treeview(sf, columns=cols, show='headings', height=5,
                                          style='Sum.Treeview')
        widths = {'날짜': 85, 'LOT번호': 100, '일련번호': 90,
                  '홀막힘수': 65, '양품': 50, '불량': 50, '불량률': 65}
        for c in cols:
            self.summary_tree.heading(c, text=c)
            self.summary_tree.column(c, width=widths[c], anchor='center')
        self.summary_tree.tag_configure('fail', background='#ffe0e0', foreground='#cc0000')
        self.summary_tree.tag_configure('ok',   background='#e8f5e9', foreground='#1a5c2e')
        self.summary_tree.pack(fill='x')

        self._refresh_summary()

    def _reset_summary(self):
        if not self._session_log:
            return
        if not messagebox.askyesno('초기화 확인',
                                    f'금일 현황 {len(self._session_log)}건을 초기화하시겠습니까?',
                                    icon='warning'):
            return
        self._session_log.clear()
        self._refresh_summary()

    def _refresh_summary(self):
        for row in self.summary_tree.get_children():
            self.summary_tree.delete(row)
        # 세션 메모리 기반으로 표시 (다음 제품 누를 때마다 쌓임)
        total_good = total_bad = 0
        for entry in self._session_log:
            lot     = entry.get('lot', '') or '-'
            serial  = entry.get('serial', '') or '-'
            verdict = entry.get('verdict', '양호')
            date    = entry.get('date', '')
            rate    = entry.get('rate', 0.0)
            good    = 0 if verdict == '불량' else 1
            bad     = 1 if verdict == '불량' else 0
            total_good += good
            total_bad  += bad
            tag = 'fail' if verdict == '불량' else 'ok'
            hole_cnt = entry.get('count', 0)
            self.summary_tree.insert('', 'end', tags=(tag,),
                values=(date, lot, serial, hole_cnt, good, bad, f'{rate:.2f}%'))

    # ── HMI 아이콘 일괄 적용 ─────────────────────────────────────────────────

    def _apply_icons(self):
        """모든 주요 버튼에 기능별 색상 HMI 아이콘 + Hover 효과 적용"""
        global _ICON_CACHE
        _ICON_CACHE.clear()          # 여기서 clear → 즉시 재생성하므로 GC 문제 없음
        t  = self.current_theme
        SZ = 22

        def _set(btn, name, text, bg_norm, bg_hover, sz=SZ,
                 fg_n='#ffffff', fg_h='#ffffff'):
            try:
                ico = _ph(name, sz, t)
                btn._icon_ref = ico  # GC 방지: 버튼 객체에 참조 보관
                btn.configure(image=ico, compound='left', text=text,
                              state='normal')
                _setup_hover(btn, bg_norm, bg_hover,
                             fg_norm=fg_n, fg_hover=fg_h)
            except Exception:
                btn.configure(state='normal')

        # ── 헤더 버튼 — 다크/라이트 명확히 대비되는 색상 ─────────────────
        alarm_on = getattr(self, 'alarm_enabled', True)
        if t == 'dark':
            hdr_norm   = '#0d2d5e';  hdr_hov   = '#1a4a8a'; hdr_fg = '#ffffff'
            alm_off_n  = '#3d1010';  alm_off_h = '#5c2020'; alm_off_fg = '#ffffff'
        else:
            # 라이트 모드: 헤더 bg='#1565c0', 버튼은 흰색으로 대비
            hdr_norm   = '#ffffff';  hdr_hov   = '#e3f2fd'; hdr_fg = '#1a3a6b'
            alm_off_n  = '#ffcdd2';  alm_off_h = '#ef9a9a'; alm_off_fg = '#7f0000'

        _set(self.btn_alarm,
             'bell' if alarm_on else 'bell_off',
             ' 알람  ' if alarm_on else ' 알람끔',
             hdr_norm  if alarm_on else alm_off_n,
             hdr_hov   if alarm_on else alm_off_h,
             fg_n=hdr_fg if alarm_on else alm_off_fg,
             fg_h=hdr_fg if alarm_on else alm_off_fg)

        _set(self.btn_theme,
             'sun'  if t == 'dark' else 'moon',
             ' 라이트 ' if t == 'dark' else ' 다크  ',
             hdr_norm, hdr_hov, fg_n=hdr_fg, fg_h=hdr_fg)

        fs_on = getattr(self, '_fullscreen', False)
        _set(self.btn_fs,
             'windowed' if fs_on else 'fullscreen',
             ' 창 모드 ' if fs_on else ' 전체화면',
             hdr_norm, hdr_hov, fg_n=hdr_fg, fg_h=hdr_fg)

        # ── 카메라 패널 버튼 (테마별 색상) ──────────────────────────────────
        cap_n  = '#0d2d5e' if t == 'dark' else '#1565c0'
        cap_h  = '#1255a0' if t == 'dark' else '#1976d2'
        save_n = '#1a5c2e' if t == 'dark' else '#2e7d32'
        save_h = '#0f8040' if t == 'dark' else '#388e3c'
        if hasattr(self, '_cap_btn'):
            _set(self._cap_btn,      'camera', '  화면 캡처', cap_n,  cap_h)
        if hasattr(self, '_cam_save_btn'):
            _set(self._cam_save_btn, 'disk',   '  저장',      save_n, save_h)

        # ── 컨트롤 패널 버튼 (테마별 색상) ──────────────────────────────────
        his_n  = '#1a5c2e' if t == 'dark' else '#1b5e20'
        his_h  = '#0f8040' if t == 'dark' else '#2e7d32'
        undo_n = '#7a4500' if t == 'dark' else '#bf360c'
        undo_h = '#b06000' if t == 'dark' else '#e64a19'
        nxt_n  = '#1a3a6b' if t == 'dark' else '#1565c0'
        nxt_h  = '#1a5aab' if t == 'dark' else '#1976d2'
        rst_n  = '#2a3540' if t == 'dark' else '#546e7a'
        rst_h  = '#3a4f60' if t == 'dark' else '#607d8b'
        if hasattr(self, 'btn_history'):
            _set(self.btn_history, 'folder', '  이력 조회 ', his_n, his_h)
        if hasattr(self, 'btn_undo'):
            _set(self.btn_undo,    'undo',   '  마지막 취소', undo_n, undo_h)
        if hasattr(self, 'btn_save'):
            _set(self.btn_save,    'disk',   '  결과 저장',   save_n, save_h)
        if hasattr(self, 'btn_next'):
            _set(self.btn_next,    'next',   '  다음 제품  ▶', nxt_n, nxt_h)
        if hasattr(self, 'btn_reset'):
            _set(self.btn_reset,   'reset',  '  전체 초기화  ( R 키 )',
                 rst_n, rst_h)

    # ── 테마 / 전체화면 ──────────────────────────────────────────────────────

    def _toggle_theme(self):
        global _CURRENT_THEME
        self.current_theme = 'light' if self.current_theme == 'dark' else 'dark'
        _CURRENT_THEME = self.current_theme          # 아이콘 시스템에 전파
        # _ICON_CACHE 는 _apply_icons 안에서 clear → 바로 재생성하여 GC 방지
        cmap = _D2L if self.current_theme == 'light' else _L2D
        self._retheme(self, cmap)
        if hasattr(self, '_s3_panel'):
            self._apply_cut_capture_theme()
        if self.current_theme == 'light':
            self._apply_light_button_styles()        # 내부에서 _apply_icons 호출
        else:
            self._apply_dark_button_styles()

    def _refresh_sdk_button_colors(self):
        """LED/노출고정/화질고정 버튼 색을 현재 테마+ON/OFF 상태에 맞게 다시 칠한다.
        테마 전환 시 및 각 토글 함수에서 상태 바뀔 때 호출됨 — 라이트↔다크 왕복 시
        일부 버튼이 이전 테마 색 그대로 남아 튀어 보이던 문제 수정(2026-07-16)."""
        if self.current_theme == 'light':
            on_bg, on_hov, off_bg, off_hov, off_fg = \
                '#1b5e20', '#2e7d32', '#455a64', '#546e7a', '#ffffff'
        else:
            on_bg, on_hov, off_bg, off_hov, off_fg = \
                '#1a5c2e', '#0f8040', '#21262d', '#30363d', TXT_G
        for btn_name, on_state in (
                ('_btn_led', self._led_on),
                ('_btn_exposure_lock', self._exposure_locked),
                ('_btn_quality_lock', self._locked_brightness is not None)):
            btn = getattr(self, btn_name, None)
            if btn is None:
                continue
            if on_state:
                _setup_hover(btn, on_bg, on_hov)
            else:
                _setup_hover(btn, off_bg, off_hov, fg_norm=off_fg)
        # 컷팅면 R/L 버튼 — 선택 상태 유지하며 테마 전환에도 색이 안 틀어지게
        if hasattr(self, '_btn_cut_r'):
            sel = self._cutting_side
            self._btn_cut_r.configure(
                bg=ACC_RED if sel == 'R' else '#21262d',
                fg='#fff' if sel == 'R' else ACC_RED)
            self._btn_cut_l.configure(
                bg=ACC_BLU if sel == 'L' else '#21262d',
                fg='#fff' if sel == 'L' else ACC_BLU)

    def _apply_light_button_styles(self):
        for v, btn in self.step_btns.items():
            if v == self.step.get():
                _setup_hover(btn, '#0d47a1', '#1565c0')
            else:
                _setup_hover(btn, '#455a64', '#546e7a')
        _setup_hover(self.btn_click_count, '#b71c1c', '#e53935')
        _setup_hover(self.btn_undo,  '#bf360c', '#e64a19')
        _setup_hover(self.btn_save,  '#1b5e20', '#2e7d32')
        _setup_hover(self.btn_next,    '#0d47a1', '#1565c0')
        _setup_hover(self.btn_reset,   '#37474f', '#455a64')
        _setup_hover(self.btn_history, '#1b5e20', '#2e7d32')
        self._refresh_sdk_button_colors()
        # LOT/작업자/일련번호 Entry 라이트 모드 색상
        if hasattr(self, '_info_entries'):
            for e, var, ph in self._info_entries:
                is_hint = self._hint_active.get(id(var), False)
                e.configure(
                    bg='#ffffff', insertbackground=LT_TXT,
                    fg='#7a8fa6' if is_hint else LT_TXT,
                    relief='flat',
                    highlightbackground='#a9bdd4', highlightcolor=ACC_BLU)
        self.after(30, self._apply_icons)

    def _apply_dark_button_styles(self):
        cur = self.step.get()
        for v, btn in self.step_btns.items():
            if v == cur:
                _setup_hover(btn, '#2c5f9e', '#3a78c2')
            else:
                _setup_hover(btn, '#21262d', '#30363d')
        _setup_hover(self.btn_click_count, ACC_RED, '#ff4040')
        _setup_hover(self.btn_undo,  '#7a5c00', '#b08800')
        _setup_hover(self.btn_save,  '#1a5c2e', '#0f8040')
        _setup_hover(self.btn_next,    '#1a3a6b', '#1a5aab')
        _setup_hover(self.btn_reset,   '#2a3540', '#3a4f60', fg_norm=TXT_W)
        _setup_hover(self.btn_history, '#1a5c2e', '#0f8040')
        self._refresh_sdk_button_colors()
        # LOT/작업자/일련번호 Entry 다크 모드 색상
        if hasattr(self, '_info_entries'):
            for e, var, ph in self._info_entries:
                is_hint = self._hint_active.get(id(var), False)
                e.configure(
                    bg='#21262d', insertbackground=TXT_W,
                    fg='#888888' if is_hint else TXT_W,
                    relief='flat',
                    highlightbackground='#30363d', highlightcolor=ACC_BLU)
        self.after(30, self._apply_icons)

    # 이 위젯들은 _apply_dark_button_styles/_apply_light_button_styles/
    # _refresh_sdk_button_colors 가 상태(ON/OFF 등)에 맞춰 항상 명시적으로
    # 색을 다시 칠하므로, 여기서 컬러맵으로 자동 변환하면 안 된다. _D2L 안에
    # 여러 다크색이 같은 라이트색으로 겹쳐 매핑돼 있어(예: #1a5c2e/#00c853/
    # #2e8b57 전부 #1b5e20) 역변환(_L2D)이 모호해지고, 라이트→다크 왕복 시
    # 엉뚱한(형광) 색으로 복원되는 버그가 있었음(2026-07-16 확인).
    _RETHEME_EXCLUDE_ATTRS = (
        'btn_save', 'btn_history', 'btn_undo', 'btn_next',
        'btn_reset', 'btn_click_count', '_btn_led', '_btn_exposure_lock',
        '_btn_quality_lock', '_cam_save_btn', '_cap_btn', 'btn_alarm',
        'btn_theme', 'btn_fs', '_btn_cut_r', '_btn_cut_l', '_s3_panel',
    )

    def _retheme(self, widget, cmap):
        for attr in self._RETHEME_EXCLUDE_ATTRS:
            if widget is getattr(self, attr, None):
                return
        if hasattr(self, 'step_btns') and widget in self.step_btns.values():
            return
        for prop in ('background', 'foreground', 'activebackground',
                     'activeforeground', 'highlightbackground', 'selectcolor'):
            try:
                val = str(widget.cget(prop)).lower()
                if val in cmap:
                    widget.configure(**{prop: cmap[val]})
            except tk.TclError:
                pass
        for child in widget.winfo_children():
            self._retheme(child, cmap)

    def _apply_cut_capture_theme(self):
        """컷팅 캡처 패널(③) 색상을 현재 테마(다크/라이트)에 맞게 직접 적용.
        이 패널은 _RETHEME_EXCLUDE_ATTRS에 들어있어 전역 _retheme 자동변환을
        타지 않으므로, 테마가 바뀔 때마다(그리고 최초 빌드 시) 반드시 이 함수를
        호출해야 함 — 그렇지 않으면 라이트 모드에서 어두운 배경/글씨가 그대로
        남아 안 보이는 문제가 재발한다."""
        c = CUT_CAPTURE_COLORS[self.current_theme]

        self._s3_panel.configure(bg=c['panel_bg'], highlightbackground=c['border'])
        self._cut_title_lbl.configure(bg=c['panel_bg'], fg=c['title_fg'])
        self._cut_subtitle_lbl.configure(bg=c['panel_bg'], fg=c['subtitle_fg'])

        for _row in self._cut_field_rows:
            _row.configure(bg=c['panel_bg'])
        for _flbl in self._cut_field_labels:
            _flbl.configure(bg=c['panel_bg'], fg=c['label_fg'])

        cs = ttk.Style()
        cs.configure('CutCapture.TCombobox',
                     fieldbackground=c['combo_bg'], background=c['combo_bg'],
                     foreground=c['combo_fg'], arrowsize=14, padding=(5, 4),
                     bordercolor=c['combo_border'], lightcolor=c['combo_border'],
                     darkcolor=c['combo_border'], insertcolor=c['combo_insert'])
        cs.map('CutCapture.TCombobox',
               fieldbackground=[('readonly', c['combo_bg'])],
               foreground=[('readonly', c['combo_fg'])])

        self._cut_side_row.configure(bg=c['panel_bg'])
        self._cut_count_hdr_lbl.configure(bg=c['panel_bg'], fg=c['label_fg'])
        self._cut_grid.configure(bg=c['panel_bg'])
        for cnt, btn in self._cut_count_btns.items():
            if cnt == self._cutting_count:
                btn.configure(bg=c['count_sel_bg'], fg=c['count_sel_fg'])
            else:
                btn.configure(bg=c['count_unsel_bg'], fg=c['count_unsel_fg'])

        self._cut_custom_row.configure(bg=c['panel_bg'])
        self._cut_custom_lbl.configure(bg=c['panel_bg'], fg=c['label_fg'])
        self._cut_count_entry.configure(bg=c['entry_bg'], fg=c['entry_fg'],
                                         insertbackground=c['entry_insert'])
        self._cut_unit_lbl.configure(bg=c['panel_bg'], fg=c['label_fg'])
        self._cut_apply_btn.configure(bg=c['apply_bg'], fg=c['apply_fg'])

        self._cut_status_lbl.configure(bg=c['panel_bg'], fg=c['status_fg'])
        self._cut_capture_btn.configure(bg=c['capture_bg'], fg=c['capture_fg'])

    def _ctrl_scroll(self, event):
        self._ctrl_canvas.yview_scroll(-1 if event.delta > 0 else 1, 'units')

    def _bind_ctrl_scroll(self, widget):
        widget.bind('<MouseWheel>', self._ctrl_scroll)
        for child in widget.winfo_children():
            self._bind_ctrl_scroll(child)

    def _toggle_fullscreen(self):
        self._fullscreen = not self._fullscreen
        self.attributes('-fullscreen', self._fullscreen)
        ico = 'windowed'  if self._fullscreen else 'fullscreen'
        txt = ' 창 모드 ' if self._fullscreen else ' 전체화면'
        self.btn_fs.configure(image=_ph(ico, 22), compound='left', text=txt)
        _setup_hover(self.btn_fs, '#0d2d5e', '#1a4a8a')
        if self._fullscreen:
            sw = self.winfo_screenwidth()
            self.after(150, lambda: self.main_paned.sash_place(0, sw - 316, 0))
        else:
            self.after(150, lambda: self.main_paned.sash_place(0,
                max(500, self.winfo_width() - 316), 0))

    def _exit_fullscreen(self):
        if self._fullscreen:
            self._toggle_fullscreen()

    def _next_product(self):
        if not self.log_list:
            self.lbl_flash.configure(text='카운팅 데이터가 없습니다', fg=ACC_YEL)
            self.after(1800, lambda: self.lbl_flash.configure(text=''))
            return
        model   = self.selected_model.get()
        fibers  = self.fiber_count.get()
        cnt     = self.count.get()
        rate    = cnt / fibers * 100 if fibers > 0 else 0
        verdict = '불량' if rate >= DEFECT_LIMIT else '양호'
        defect_reason = self._ask_defect_reason() if verdict == '불량' else ''

        h = self._hint_active
        lot    = '' if h.get(id(self.lot_no))    else self.lot_no.get()
        op     = '' if h.get(id(self.operator))  else self.operator.get()
        serial = '' if h.get(id(self.serial_no)) else self.serial_no.get()

        missing = []
        if not lot:
            missing.append('LOT 번호')
        if not serial:
            missing.append('일련번호')
        if missing:
            if not messagebox.askyesno(
                    '입력값 확인',
                    f'{", ".join(missing)}이(가) 비어 있습니다.\n'
                    '이 상태로 저장하면 나중에 이력에서 찾기 어려울 수 있습니다.\n\n'
                    '그래도 저장하시겠습니까?', icon='warning'):
                return

        import datetime as _dt
        current_entry = {
            'date':    _dt.datetime.now().strftime('%Y-%m-%d'),
            'time':    _dt.datetime.now().strftime('%H:%M:%S'),
            'lot':     lot,
            'serial':  serial,
            'verdict': verdict,
            'rate':    round(rate, 2),
            'count':   cnt,
            'model':   model,
            'defect_reason': defect_reason,
        }

        # ── 중복 일련번호 감지 ────────────────────────────────────────────────
        dup_idx = -1
        if serial:
            for i, e in enumerate(self._session_log):
                if e.get('serial') == serial:
                    dup_idx = i
                    break

        if dup_idx >= 0:
            choice = self._show_duplicate_dialog(self._session_log[dup_idx], current_entry)
            if choice == 'cancel':
                return
            elif choice == 'keep_existing':
                # 현재 결과 버림 → 카운터만 초기화
                self._do_reset_counters()
                self.lbl_flash.configure(text='기존 결과를 유지합니다', fg=ACC_YEL)
                self.after(2500, lambda: self.lbl_flash.configure(text=''))
                return
            else:  # 'use_current'
                # 기존 항목 세션 로그에서 제거 + JSON에서도 제거
                old = self._session_log.pop(dup_idx)
                self._remove_from_history(old.get('time', ''), serial)

        # 세션 로그에 추가
        self._session_log.append(current_entry)
        self._refresh_summary()

        # autosave 여부와 무관하게 항상 JSON 이력에 기록
        self._save_history(model, fibers, cnt, rate, verdict, defect_reason)

        self._do_reset_counters()
        color = ACC_GRN if verdict == '양호' else ACC_RED
        if dup_idx >= 0:
            msg = f'✅  현재 결과로 교체 완료  ({serial})'
        else:
            msg = f'✅  {verdict} 기록 완료  →  다음 제품 카운팅을 시작하세요'
        self.lbl_flash.configure(text=msg, fg=color)
        self.after(3000, lambda: self.lbl_flash.configure(text=''))

    def _do_reset_counters(self):
        self.count.set(0)
        self.log_list.clear()
        self._alarm_done = False
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._update_rate()

    def _remove_from_history(self, time_str, serial):
        """이력 DB에서 동일 일련번호+시각 항목 제거"""
        try:
            if _db_delete_by_serial_time(serial, time_str):
                self._load_total_products()
        except Exception:
            pass

    # ── 불량 사유 선택 ────────────────────────────────────────────────────────

    def _ask_defect_reason(self):
        """불량 판정 시 사유 선택 다이얼로그. 반환값: 사유 문자열."""
        result = {'reason': '미입력'}

        dlg = tk.Toplevel(self)
        dlg.title('불량 사유 선택')
        dlg.configure(bg=CARD_BG)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self)

        hdr = tk.Frame(dlg, bg='#7b1a1a')
        hdr.pack(fill='x')
        tk.Label(hdr, text='  ⚠  불량 판정 — 사유를 선택하세요',
                 font=('맑은 고딕', 12, 'bold'),
                 bg='#7b1a1a', fg='#ffffff').pack(anchor='w', pady=10, padx=10)

        var  = tk.StringVar(value=DEFECT_REASONS[0])
        body = tk.Frame(dlg, bg=CARD_BG)
        body.pack(fill='both', padx=20, pady=14)
        for r in DEFECT_REASONS:
            tk.Radiobutton(
                body, text=r, variable=var, value=r,
                font=('맑은 고딕', 11), bg=CARD_BG, fg=TXT_W,
                activebackground=CARD_BG, activeforeground=TXT_W,
                selectcolor='#5c1a1a', cursor='hand2').pack(anchor='w', pady=4)

        tk.Label(body, text='비고 (선택 입력):', font=('맑은 고딕', 9),
                 bg=CARD_BG, fg=TXT_G).pack(anchor='w', pady=(10, 2))
        note_var = tk.StringVar(value='')
        tk.Entry(body, textvariable=note_var, font=('맑은 고딕', 10),
                  bg='#21262d', fg=TXT_W, insertbackground=TXT_W,
                  relief='flat').pack(fill='x', ipady=3)

        def on_confirm():
            reason = var.get()
            note   = note_var.get().strip()
            result['reason'] = f'{reason} ({note})' if note else reason
            dlg.destroy()

        btn_f = tk.Frame(dlg, bg=CARD_BG)
        btn_f.pack(side='bottom', fill='x', padx=16, pady=14)
        tk.Button(btn_f, text='확인', font=('맑은 고딕', 11, 'bold'),
                  bg='#7b1a1a', fg='#ffffff', relief='flat', cursor='hand2',
                  width=12, command=on_confirm).pack(side='right', padx=4)

        dlg.protocol('WM_DELETE_WINDOW', on_confirm)
        dlg.wait_window()
        return result['reason']

    def _show_duplicate_dialog(self, existing, current):
        """중복 일련번호 비교 선택 다이얼로그. 반환값: 'keep_existing' / 'use_current' / 'cancel'"""
        result = ['cancel']

        # 이 팝업은 라이트/다크 어느 쪽에서 떠도 항상 다크 고정 배색이었음
        # (컷팅캡처 패널과 마찬가지로 자동 리테마 대상 밖에 있는 위젯이라
        #  직접 색을 골라줘야 함) — 생성 시점에 현재 테마를 보고 팔레트를 선택.
        if self.current_theme == 'light':
            DLG_BG, WARN_FG, DESC_FG = '#f4f6fa', '#8a5a00', '#5a6b80'
            HDR_BG,  ODD_BG, EVN_BG = '#d7e3f5', '#eef2f9', '#e3e9f5'
            HDR_FG_MAIN, HDR_FG_LBL = '#0d3b66', '#5a6b80'
            LBL_FG, VAL_FG = '#4a6080', '#16202e'
            LESS_FG, MORE_FG = '#1b7a1b', '#c62828'
            HINT_FG = '#1b7a1b'
            BTN_KEEP = dict(bg='#c8d8ef', fg='#0d3b66', activebackground='#b0c8ea')
            BTN_USE  = dict(bg='#c8e6c9', fg='#1b5e20', activebackground='#a8d8aa')
            BTN_CNCL = dict(bg='#e0e0e0', fg='#555555', activebackground='#cfcfcf')
        else:
            DLG_BG, WARN_FG, DESC_FG = '#1a1f2e', '#f0c040', '#a0b0c0'
            HDR_BG,  ODD_BG, EVN_BG = '#0d2044', '#1e2840', '#232e48'
            HDR_FG_MAIN, HDR_FG_LBL = '#cce0ff', '#7a8fa6'
            LBL_FG, VAL_FG = '#8aaccc', '#e8eef8'
            LESS_FG, MORE_FG = '#50d050', '#e05050'
            HINT_FG = '#a0e0a0'
            BTN_KEEP = dict(bg='#2c4a6e', fg='#cce0ff', activebackground='#3a6090')
            BTN_USE  = dict(bg='#1a5c2e', fg='#ccffcc', activebackground='#2a8a45')
            BTN_CNCL = dict(bg='#3a3a3a', fg='#aaaaaa', activebackground='#555555')

        dlg = tk.Toplevel(self)
        dlg.title(f'중복 일련번호 감지 — {current.get("serial", "")}')
        dlg.configure(bg=DLG_BG)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self)

        serial = current.get('serial', '-')
        tk.Label(dlg, text=f'⚠  일련번호  "{serial}"  이(가) 이미 검사되었습니다.',
                 font=('맑은 고딕', 11, 'bold'), bg=DLG_BG, fg=WARN_FG
                 ).pack(padx=20, pady=(14, 6))
        tk.Label(dlg, text='저장할 결과를 선택하세요. 선택하지 않은 결과는 삭제됩니다.',
                 font=('맑은 고딕', 9), bg=DLG_BG, fg=DESC_FG
                 ).pack(padx=20, pady=(0, 10))

        # ── 비교 테이블 ──────────────────────────────────────────────────────
        tbl = tk.Frame(dlg, bg=DLG_BG)
        tbl.pack(padx=20, pady=(0, 12), fill='x')

        def col_badge(val_e, val_c):
            """홀막힘 수를 비교해 색상 반환 (기존, 현재)"""
            if val_e < val_c:
                return LESS_FG, MORE_FG
            elif val_e > val_c:
                return MORE_FG, LESS_FG
            return VAL_FG, VAL_FG

        cnt_e = existing.get('count', 0)
        cnt_c = current.get('count', 0)
        rate_e = existing.get('rate', 0.0)
        rate_c = current.get('rate', 0.0)

        col_e, col_c = col_badge(cnt_e, cnt_c)

        rows_data = [
            ('', '기존 검사 결과', '현재 검사 결과'),
            ('검사 일시',
             f"{existing.get('date','-')} {existing.get('time','-')}",
             f"{current.get('date','-')}  {current.get('time','-')}"),
            ('LOT',       existing.get('lot', '-'),   current.get('lot', '-')),
            ('모델',      existing.get('model', '-'), current.get('model', '-')),
            ('홀막힘 수', f"{cnt_e}개",               f"{cnt_c}개"),
            ('불량률',    f"{rate_e:.2f}%",           f"{rate_c:.2f}%"),
            ('판정',      existing.get('verdict', '-'), current.get('verdict', '-')),
        ]

        for r, row in enumerate(rows_data):
            bg = HDR_BG if r == 0 else (ODD_BG if r % 2 else EVN_BG)
            for c, cell in enumerate(row):
                if r == 0:
                    fg = HDR_FG_MAIN if c > 0 else HDR_FG_LBL
                    font = ('맑은 고딕', 10, 'bold')
                elif c == 0:
                    fg, font = LBL_FG, ('맑은 고딕', 9)
                elif row[0] == '홀막힘 수':
                    fg  = col_e if c == 1 else col_c
                    font = ('맑은 고딕', 10, 'bold')
                elif row[0] == '불량률':
                    fg  = col_e if c == 1 else col_c
                    font = ('맑은 고딕', 9, 'bold')
                elif row[0] == '판정':
                    v   = existing.get('verdict') if c == 1 else current.get('verdict')
                    fg  = MORE_FG if v == '불량' else LESS_FG
                    font = ('맑은 고딕', 9, 'bold')
                else:
                    fg, font = VAL_FG, ('맑은 고딕', 9)

                pad_x = (12, 6) if c == 0 else (6, 12)
                tk.Label(tbl, text=cell, bg=bg, fg=fg, font=font,
                         anchor='w', padx=pad_x[0], pady=5,
                         width=26 if c > 0 else 10
                         ).grid(row=r, column=c, sticky='ew', padx=1, pady=1)

        # 홀막힘 적은 쪽 안내
        hint = ''
        if cnt_e < cnt_c:
            hint = f'◀ 기존 결과가 홀막힘이 {cnt_c - cnt_e}개 적습니다'
        elif cnt_c < cnt_e:
            hint = f'현재 결과가 홀막힘이 {cnt_e - cnt_c}개 적습니다 ▶'
        else:
            hint = '양쪽 홀막힘 수가 동일합니다'
        tk.Label(dlg, text=hint, font=('맑은 고딕', 9, 'bold'),
                 bg=DLG_BG, fg=HINT_FG).pack(pady=(0, 12))

        # ── 버튼 ─────────────────────────────────────────────────────────────
        btn_row = tk.Frame(dlg, bg=DLG_BG)
        btn_row.pack(padx=20, pady=(0, 16))

        def pick(choice):
            result[0] = choice
            dlg.destroy()

        tk.Button(btn_row, text='기존 결과 유지',
                  font=('맑은 고딕', 10, 'bold'), width=14,
                  bg=BTN_KEEP['bg'], fg=BTN_KEEP['fg'],
                  activebackground=BTN_KEEP['activebackground'],
                  relief='flat', cursor='hand2', pady=6,
                  command=lambda: pick('keep_existing')
                  ).pack(side='left', padx=6)

        tk.Button(btn_row, text='현재 결과 저장',
                  font=('맑은 고딕', 10, 'bold'), width=14,
                  bg=BTN_USE['bg'], fg=BTN_USE['fg'],
                  activebackground=BTN_USE['activebackground'],
                  relief='flat', cursor='hand2', pady=6,
                  command=lambda: pick('use_current')
                  ).pack(side='left', padx=6)

        tk.Button(btn_row, text='취소',
                  font=('맑은 고딕', 10), width=8,
                  bg=BTN_CNCL['bg'], fg=BTN_CNCL['fg'],
                  activebackground=BTN_CNCL['activebackground'],
                  relief='flat', cursor='hand2', pady=6,
                  command=lambda: pick('cancel')
                  ).pack(side='left', padx=6)

        dlg.update_idletasks()
        w = dlg.winfo_width()
        h = dlg.winfo_height()
        x = self.winfo_x() + (self.winfo_width()  - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f'+{x}+{y}')

        dlg.wait_window()
        return result[0]

    def _toggle_alarm(self):
        self.alarm_enabled = not self.alarm_enabled
        if self.alarm_enabled:
            self.btn_alarm.configure(image=_ph('bell', 22), compound='left', text=' 알람  ')
            _setup_hover(self.btn_alarm, '#0d2d5e', '#1a4a8a')
        else:
            self.btn_alarm.configure(image=_ph('bell_off', 22), compound='left', text=' 알람끔')
            _setup_hover(self.btn_alarm, '#3d1010', '#5c2020')

    # ── 시계 ──────────────────────────────────────────────────────────────────

    def _tick_clock(self):
        self.lbl_time.configure(text=datetime.datetime.now().strftime('%Y-%m-%d  %H:%M:%S'))
        self.after(1000, self._tick_clock)

    # ── 키 바인딩 ─────────────────────────────────────────────────────────────

    def _bind_keys(self):
        self.bind_all('<F11>',        lambda e: self._toggle_fullscreen())
        self.bind_all('<Escape>',     lambda e: self._exit_fullscreen())
        self._apply_count_key(getattr(self, 'count_key', 'space'))
        self.bind_all('<KeyPress-z>', self._on_key_guard(self._undo))
        self.bind_all('<KeyPress-Z>', self._on_key_guard(self._undo))
        self.bind_all('<KeyPress-r>', self._on_key_guard(self._reset))
        self.bind_all('<KeyPress-R>', self._on_key_guard(self._reset))
        self.bind_all('<KeyPress-s>', self._on_key_guard(self._save_excel))
        self.bind_all('<KeyPress-S>', self._on_key_guard(self._save_excel))
        self.bind_all('<KeyPress-1>', self._on_key_guard(lambda: self._set_step(1)))
        self.bind_all('<KeyPress-2>', self._on_key_guard(lambda: self._set_step(10)))
        self.bind_all('<KeyPress-3>', self._on_key_guard(lambda: self._set_step(20)))

    def _is_entry_focused(self):
        """실제 텍스트 입력 위젯(Entry/Text)에 포커스가 있으면 True.
        Combobox는 카운트 키 입력을 막지 않도록 제외."""
        w = self.focus_get()
        return isinstance(w, (tk.Entry, tk.Text, ttk.Entry))

    def _on_key_guard(self, fn):
        """Entry 포커스 중에는 단축키를 막는 래퍼"""
        def handler(event):
            if not self._is_entry_focused():
                fn()
        return handler

    # ── 카운트 키(단축키) 설정 ───────────────────────────────────────────────
    # 기본값은 SPACE 이지만, [변경] 버튼으로 원하는 키보드 키로 바꿀 수 있다.

    _COUNT_KEY_LABELS = {
        'space': 'SPACE', 'Return': 'ENTER', 'Tab': 'TAB',
        'Up': '↑', 'Down': '↓', 'Left': '←', 'Right': '→',
        'BackSpace': 'BACKSPACE', 'Delete': 'DELETE',
    }
    _COUNT_KEY_RESERVED = {
        'z': 'Z (실행취소)',    'Z': 'Z (실행취소)',
        'r': 'R (전체초기화)',  'R': 'R (전체초기화)',
        's': 'S (엑셀저장)',    'S': 'S (엑셀저장)',
        '1': '1 (1개씩 단위)', '2': '2 (10개씩 단위)', '3': '3 (20개씩 단위)',
        'F11': 'F11 (전체화면)', 'Escape': 'Escape (전체화면 종료)',
    }
    _COUNT_KEY_BLOCKED_MODIFIERS = {
        'Shift_L', 'Shift_R', 'Control_L', 'Control_R',
        'Alt_L', 'Alt_R', 'Caps_Lock', 'Num_Lock',
    }

    def _count_key_label(self, keysym):
        return self._COUNT_KEY_LABELS.get(keysym, keysym.upper())

    def _on_count_key(self, event):
        if not self._is_entry_focused():
            self._count_hole()

    def _on_count_key_entry(self, event):
        self._count_hole()
        return 'break'

    def _count_key_target_entries(self):
        """카운트 키를 누르면 문자 입력 대신 바로 카운트되어야 하는, 메인
        화면에 있는 모든 입력칸(Entry) — 검사 정보란 3칸 + 캘리브레이션
        기준 길이칸. (이력창/팝업 다이얼로그의 입력칸은 제외 — 거기서는
        정상적으로 문자를 입력할 수 있어야 한다.)"""
        entries = [e for e, _v, _p in getattr(self, '_info_entries', [])]
        cal_entry = getattr(self, '_cal_entry', None)
        if cal_entry is not None:
            entries.append(cal_entry)
        return entries

    def _apply_count_key(self, keysym):
        """카운트 키를 keysym 으로 (재)바인딩한다. 전역 바인딩과,
        메인 화면 입력칸들(_count_key_target_entries)에서 문자 입력 대신
        바로 카운트되게 하는 위젯별 가로채기를 함께 갱신한다."""
        old = getattr(self, 'count_key', None)
        targets = self._count_key_target_entries()
        if old and old != keysym:
            try:
                self.unbind_all(f'<KeyPress-{old}>')
            except Exception:
                pass
            for e in targets:
                try:
                    e.unbind(f'<KeyPress-{old}>')
                except Exception:
                    pass
        self.count_key = keysym
        self.bind_all(f'<KeyPress-{keysym}>', self._on_count_key)
        for e in targets:
            e.bind(f'<KeyPress-{keysym}>', self._on_count_key_entry)
        self._save_config()
        if hasattr(self, 'lbl_count_key'):
            self.lbl_count_key.configure(text=f'카운트 키: {self._count_key_label(keysym)}')
        if hasattr(self, '_lbl_shortcut_key'):
            self._lbl_shortcut_key.configure(text=self._count_key_label(keysym))
        self._update_count_btn_text()

    def _open_count_key_dialog(self):
        dlg = tk.Toplevel(self)
        dlg.title('카운트 키 변경')
        dlg.configure(bg=CARD_BG)
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)

        tk.Label(dlg, text='카운트로 사용할 키를 눌러주세요',
                 font=('맑은 고딕', 11, 'bold'), bg=CARD_BG, fg='#ffffff'
                 ).pack(padx=24, pady=(18, 6))
        tk.Label(dlg, text='(마우스 클릭 말고 키보드 키를 누르세요 — 취소: Esc)',
                 font=('맑은 고딕', 9), bg=CARD_BG, fg=TXT_G
                 ).pack(padx=24, pady=(0, 14))

        def on_key(ev):
            if ev.keysym == 'Escape':
                dlg.destroy()
                return
            if ev.keysym in self._COUNT_KEY_BLOCKED_MODIFIERS:
                return
            if ev.keysym in self._COUNT_KEY_RESERVED:
                messagebox.showwarning(
                    '사용 불가',
                    f'"{self._COUNT_KEY_RESERVED[ev.keysym]}" 은(는) 이미 다른 '
                    '단축키로 사용 중입니다.\n다른 키를 눌러주세요.', parent=dlg)
                return
            self._apply_count_key(ev.keysym)
            dlg.destroy()
            self.lbl_flash.configure(
                text=f'✅  카운트 키가 [{self._count_key_label(ev.keysym)}] (으)로 변경되었습니다',
                fg='#4a9fd4')
            self.after(2500, lambda: self.lbl_flash.configure(text=''))

        dlg.bind('<Key>', on_key)

        def reset_to_default():
            self._apply_count_key('space')
            dlg.destroy()
            self.lbl_flash.configure(
                text='✅  카운트 키가 [SPACE] (으)로 초기화되었습니다', fg='#4a9fd4')
            self.after(2500, lambda: self.lbl_flash.configure(text=''))

        btn_row = tk.Frame(dlg, bg=CARD_BG)
        btn_row.pack(pady=(0, 18))
        tk.Button(btn_row, text='기본값(SPACE)으로', font=('맑은 고딕', 9, 'bold'),
                  bg='#2a3540', fg='#ffffff', relief='flat', cursor='hand2',
                  padx=8, pady=4, command=reset_to_default
                  ).pack(side='left', padx=(24, 6))
        tk.Button(btn_row, text='취소', font=('맑은 고딕', 9),
                  bg='#21262d', fg='#ffffff', relief='flat', cursor='hand2',
                  padx=8, pady=4, command=dlg.destroy).pack(side='left', padx=(0, 24))

        dlg.focus_set()


    # ── 단위 설정 ─────────────────────────────────────────────────────────────

    def _set_step(self, val):
        self.step.set(val)
        if self.current_theme == 'light':
            for v, btn in self.step_btns.items():
                btn.configure(bg='#0d47a1' if v == val else '#546e7a', fg='#ffffff')
        else:
            for v, btn in self.step_btns.items():
                btn.configure(bg='#2c5f9e' if v == val else '#21262d', fg='#fff')
        label = {1: '1개씩', 10: '10개씩', 20: '20개씩'}.get(val, f'{val}개씩')
        self._update_count_btn_text()
        self.lbl_flash.configure(text=f'단위: {label} 설정됨', fg='#4a9fd4')
        self.after(1200, lambda: self.lbl_flash.configure(text=''))

    def _update_count_btn_text(self):
        label = {1: '1개씩', 10: '10개씩', 20: '20개씩'}.get(
            self.step.get(), f'{self.step.get()}개씩')
        key_label = self._count_key_label(getattr(self, 'count_key', 'space'))
        if hasattr(self, 'btn_click_count'):
            self.btn_click_count.configure(
                text=f'● 홀막힘  카운트  [ {label} ]\n( 클릭  또는  {key_label} )')

    # ── 카운트 ────────────────────────────────────────────────────────────────

    def _count_hole(self, *_):
        if not self._model_selected:
            self.lbl_flash.configure(text='먼저 모델을 선택하세요!', fg=ACC_YEL)
            self.after(2000, lambda: self.lbl_flash.configure(text=''))
            return
        if not self.lot_no.get().strip() or not self.operator.get().strip():
            self.lbl_flash.configure(text='LOT No.와 작업자를 먼저 입력하세요!', fg=ACC_YEL)
            self.after(2000, lambda: self.lbl_flash.configure(text=''))
            return

        step      = self.step.get()
        new_val   = self.count.get() + step
        self.count.set(new_val)

        fibers = self.fiber_count.get()
        rate   = new_val / fibers * 100
        verdict = '불량' if rate >= DEFECT_LIMIT else '양호'
        now    = datetime.datetime.now().strftime('%H:%M:%S')
        label  = {1: '1개씩', 10: '10개씩', 20: '20개씩'}.get(step, f'{step}개씩')

        self.log_list.append({
            'time': now, 'model': self.selected_model.get(),
            'count': new_val, 'step': step,
            'fibers': fibers, 'rate': rate, 'verdict': verdict,
        })
        self._add_log_row(len(self.log_list), now, self.selected_model.get(),
                          label, new_val, rate, verdict)
        self._update_rate()

        if rate >= DEFECT_LIMIT:
            self._flash_red(new_val, fibers, rate)
            if not self._alarm_done and self.alarm_enabled:
                self._alarm_done = True
                threading.Thread(target=self._play_alarm, daemon=True).start()

    def _undo(self, *_):
        if self.count.get() <= 0 or not self.log_list:
            return
        last = self.log_list.pop()
        self.count.set(max(0, self.count.get() - last.get('step', 1)))
        children = self.tree.get_children()
        if children:
            self.tree.delete(children[-1])
        step_label = {1: '1개씩', 10: '10개씩', 20: '20개씩'}.get(last.get('step', 1), '?')
        self.lbl_flash.configure(text=f'취소: {step_label}  {last.get("count", 0)}개', fg=ACC_YEL)
        self.after(1500, lambda: self.lbl_flash.configure(text=''))
        self._update_rate()
        fibers = self.fiber_count.get()
        if fibers > 0:
            rate = self.count.get() / fibers * 100
            if rate < DEFECT_LIMIT:
                self._alarm_done = False

    def _reset(self, *_):
        cnt = self.count.get()
        if cnt == 0:
            return
        if not messagebox.askyesno('초기화 확인',
                                    f'카운트 {cnt}건을 초기화하시겠습니까?', icon='warning'):
            return
        self._do_reset()

    def _do_reset(self):
        self.count.set(0)
        self.log_list.clear()
        self._alarm_done = False
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._update_rate()
        self.lbl_flash.configure(text='초기화 완료', fg=TXT_G)
        self.after(1500, lambda: self.lbl_flash.configure(text=''))

    # ── 불량률 업데이트 ───────────────────────────────────────────────────────

    def _update_rate(self):
        cnt    = self.count.get()
        fibers = self.fiber_count.get()

        if cnt == 0 or not self.selected_model.get():
            self.lbl_rate_big.configure(text='불량률: —', fg=TXT_G)
            self.lbl_verdict.configure(text='')
            self.lbl_fiber_info.configure(text='모델을 먼저 선택하세요')
            model = self.selected_model.get()
            if model:
                suffix = 'H' if model.endswith('H') else 'L'
                self.lbl_count.configure(fg='#ff6b6b' if suffix == 'H' else '#5b9bd5')
            return

        rate   = cnt / fibers * 100
        model  = self.selected_model.get()
        suffix = 'H' if model.endswith('H') else 'L'

        # 홀막힘 수 색
        if rate >= DEFECT_LIMIT:
            self.lbl_count.configure(fg=ACC_RED)
        elif rate >= DEFECT_LIMIT * 0.75:
            self.lbl_count.configure(fg='#ff4444')
        else:
            self.lbl_count.configure(fg='#ff6b6b' if suffix == 'H' else '#5b9bd5')

        limit_count = int(DEFECT_LIMIT / 100 * fibers)

        if rate >= DEFECT_LIMIT:
            self.lbl_rate_big.configure(text=f'불량률: {rate:.2f}%', fg=ACC_RED)
            self.lbl_verdict.configure(text='불  량  판  정', fg=ACC_RED)
        elif rate >= DEFECT_LIMIT * 0.75:
            remaining = limit_count - cnt
            self.lbl_rate_big.configure(
                text=f'주의 — {remaining}개 남음', fg='#ff4444')
            self.lbl_verdict.configure(text='양  호', fg='#4caf50')
        else:
            self.lbl_rate_big.configure(text=f'불량률: {rate:.2f}%', fg='#4caf50')
            self.lbl_verdict.configure(text='양  호', fg='#4caf50')

        fibers_disp = f'{fibers:,}'
        self.lbl_fiber_info.configure(
            text=f'  파이버 {model}, {fibers_disp}EA  기준 {limit_count}개')

    # ── 경보 ──────────────────────────────────────────────────────────────────

    def _flash_red(self, cnt, fibers, rate):
        self.flash_active = True
        self.lbl_flash.configure(
            text=f'+{self.step.get()}개  →  누계 {cnt}개  ({rate:.2f}%)',
            fg=ACC_RED)
        self.after(1800, lambda: setattr(self, 'flash_active', False)
                   if not self.flash_active else None)

    def _play_alarm(self):
        for _ in range(3):
            winsound.Beep(1200, 100)
            time.sleep(0.2)
            winsound.Beep(800, 600)

    # ── 로그 행 추가 ─────────────────────────────────────────────────────────

    def _add_log_row(self, num, now, model, label, cnt, rate, verdict):
        tag = 'fail' if rate >= DEFECT_LIMIT else ('warn' if rate >= DEFECT_LIMIT * 0.75 else 'ok')
        self.tree.insert('', 'end', tags=(tag,), values=(
            num, now, model, label, f'{cnt}개', f'{rate:.2f}%', verdict))
        children = self.tree.get_children()
        if children:
            self.tree.see(children[-1])

    # ── 엑셀 저장 ─────────────────────────────────────────────────────────────

    def _save_excel(self, *_):
        if not self.log_list:
            messagebox.showinfo('알림', '저장할 데이터가 없습니다.')
            return

        model   = self.selected_model.get()
        fibers  = self.fiber_count.get()
        cnt     = self.count.get()
        rate    = cnt / fibers * 100 if fibers > 0 else 0
        verdict = '불량' if rate >= DEFECT_LIMIT else '양호'
        defect_reason = self._ask_defect_reason() if verdict == '불량' else ''
        ts      = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        serial  = self.serial_no.get().strip()
        serial_part = f'_{serial}' if serial else ''

        default = os.path.join(os.path.expanduser('~'), 'Desktop',
                               f'홀막힘_{ts}{serial_part}.xlsx')
        filepath = filedialog.asksaveasfilename(
            initialfile=os.path.basename(default),
            initialdir=os.path.dirname(default),
            defaultextension='.xlsx',
            filetypes=[('Excel 통합문서 (*.xlsx)', '*.xlsx'), ('모든 파일', '*.*')])
        if not filepath:
            return
        if not filepath.lower().endswith('.xlsx'):
            filepath += '.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '홀막힘 검사'

        thin  = Side(style='thin', color='CCCCCC')
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr   = Alignment(horizontal='center', vertical='center')
        hdr_fill = PatternFill('solid', fgColor='1A3A6B')

        # 제목 행
        ws.merge_cells('A1:H1')
        c = ws['A1']
        c.value     = '멤브레인 홀막힘 검사 결과'
        c.font      = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
        c.alignment = ctr
        c.fill      = hdr_fill
        ws.row_dimensions[1].height = 28

        # 정보 블록
        info_rows = [
            ('검사 일시', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('LOT No.',   _xl_safe(self.lot_no.get())),
            ('작업자',     _xl_safe(self.operator.get())),
            ('제품일련번호', _xl_safe(serial) if serial else ''),
            ('모델',       model),
            ('총 파이버',  f'{fibers:,}EA'),
            ('홀막힘 수',  f'{cnt}개'),
            ('불량률',     f'{rate:.2f}%'),
            ('불량 기준',  f'{DEFECT_LIMIT}% 이상'),
            ('최종 판정',  verdict),
            ('불량 사유',  _xl_safe(defect_reason) if defect_reason else '-'),
            ('이탈 거리',  f'{self._deviation_val:.2f} mm' if self._deviation_val > 0 else '-'),
        ]
        for r, (k, v) in enumerate(info_rows, 2):
            ws.cell(r, 1, k).font      = Font(name='맑은 고딕', size=10, bold=True, color='AACCEE')
            ws.cell(r, 1, k).fill      = PatternFill('solid', fgColor='1A3A6B')
            ws.cell(r, 1, k).alignment = ctr
            ws.cell(r, 1, k).border    = bdr
            ws.cell(r, 2, v).font      = Font(name='맑은 고딕', size=10)
            ws.cell(r, 2, v).alignment = ctr
            ws.cell(r, 2, v).border    = bdr
            if k == '최종 판정':
                color = 'CC0000' if verdict == '불량' else '006600'
                ws.cell(r, 2).font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
                ws.cell(r, 2).fill = PatternFill('solid', fgColor=color)

        # 로그 헤더
        log_start = len(info_rows) + 3
        log_cols  = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H')
        disp_hdr  = ('순번', '시각', '모델', '홀막힘 수', '단위', '파이버', '불량률', '판정')
        for col, hd in zip(log_cols, disp_hdr):
            c = ws[f'{col}{log_start}']
            c.value     = hd
            c.font      = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
            c.fill      = PatternFill('solid', fgColor='1A3A6B')
            c.alignment = ctr
            c.border    = bdr

        step_map = {1: '1개씩', 10: '10개씩', 20: '20개씩'}
        for i, entry in enumerate(self.log_list, 1):
            r  = log_start + i
            rv = entry.get('rate', 0)
            vd = entry.get('verdict', '')
            row_fill = PatternFill('solid', fgColor='FFE8E8' if vd == '불량' else 'F5F5F5')
            vals = [
                i,
                entry.get('time', ''),
                entry.get('model', ''),
                f"{entry.get('count', 0)}개",
                step_map.get(entry.get('step', 1), '?'),
                f"{entry.get('fibers', 0):,}EA",
                f"{rv:.2f}%",
                vd,
            ]
            for col, val in zip(log_cols, vals):
                cell = ws[f'{col}{r}']
                cell.value     = val
                cell.font      = Font(name='맑은 고딕', size=10)
                cell.alignment = ctr
                cell.border    = bdr
                cell.fill      = row_fill

        col_widths = {'A': 8, 'B': 14, 'C': 10, 'D': 12, 'E': 10,
                      'F': 14, 'G': 12, 'H': 10}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        try:
            wb.save(filepath)
        except PermissionError:
            messagebox.showerror(
                '저장 오류',
                f'파일이 이미 열려 있습니다. Excel을 닫고 다시 시도하세요.\n\n{filepath}')
            return
        except Exception as exc:
            messagebox.showerror('저장 오류',
                f'파일 저장 중 오류가 발생했습니다:\n\n{exc}')
            return

        self._last_saved_path = filepath
        self._last_saved_history_id = self._save_history(
            model, fibers, cnt, rate, verdict, defect_reason)
        self.lbl_flash.configure(text=f'저장 완료!\n{os.path.basename(filepath)}', fg=ACC_GRN)
        self.after(2000, lambda: self.lbl_flash.configure(text=''))
        messagebox.showinfo('저장 완료',
                            f'저장 완료!\n{filepath}\n\n검사 이력에도 자동 기록되었습니다.')

    # ── 파일저장 직전취소 ─────────────────────────────────────────────────────

    def _cancel_last_save(self):
        if not self._last_saved_path:
            messagebox.showinfo('알림', '취소할 저장 파일이 없습니다.')
            return
        fname = os.path.basename(self._last_saved_path)
        if not messagebox.askyesno('저장 취소',
                                    f'마지막 저장 파일을 삭제하시겠습니까?\n{fname}',
                                    icon='warning'):
            return
        try:
            if os.path.exists(self._last_saved_path):
                os.remove(self._last_saved_path)
            # 이력 DB에서도 방금 저장한 그 행만 정확히 제거
            if self._last_saved_history_id is not None:
                _db_delete_by_id(self._last_saved_history_id)
                self._load_total_products()
            self._last_saved_path = None
            self._last_saved_history_id = None
            self.lbl_flash.configure(text='저장 취소 완료', fg=ACC_YEL)
            self.after(1500, lambda: self.lbl_flash.configure(text=''))
        except Exception as ex:
            messagebox.showerror(
                '저장 취소 실패',
                '파일 또는 이력 삭제 중 문제가 발생했습니다.\n'
                'Excel 등에서 관련 파일을 열어두었는지 확인 후 다시 시도하세요.\n\n'
                f'상세 정보: {ex}')

    # ── 화면 캡처 ─────────────────────────────────────────────────────────────

    def _capture_screen(self):
        with self._frame_lock:
            frame = self._frame.copy() if self._frame is not None else None
        if frame is None:
            messagebox.showwarning('캡처 실패', '카메라 화면이 없습니다.')
            return

        h = self._hint_active
        lot    = '' if h.get(id(self.lot_no))    else self.lot_no.get().strip()
        serial = '' if h.get(id(self.serial_no)) else self.serial_no.get().strip()
        model  = self.selected_model.get() or 'NOMODEL'
        now    = datetime.datetime.now()
        ts     = now.strftime('%Y%m%d_%H%M%S')

        # LOT 번호로 폴더 생성: 홀막힘_캡처/LOT-2026-001/
        # (LOT/일련번호는 자유 입력값이므로 경로 조작·Windows 금지문자를 제거해 사용)
        lot_folder = _safe_filename(lot, 'NO_LOT')
        safe_serial = _safe_filename(serial, '') if serial else ''
        save_dir   = os.path.join(CAPTURE_DIR, lot_folder)

        # 파일명: 일련번호.jpg (없으면 타임스탬프)
        if safe_serial:
            base_name = f'{safe_serial}.jpg'
        else:
            base_name = f'캡처_{ts}.jpg'

        fname = os.path.join(save_dir, base_name)

        # 동일 파일명 존재 시 → SN-001_YYYYMMDD_HHMMSS.jpg
        if os.path.exists(fname) and safe_serial:
            base_name = f'{safe_serial}_{ts}.jpg'
            fname     = os.path.join(save_dir, base_name)

        img  = Image.fromarray(frame)
        draw = ImageDraw.Draw(img)
        iw, ih = img.width, img.height
        try:
            font    = ImageFont.truetype('C:\\Windows\\Fonts\\malgunbd.ttf', 20)
            font_md = ImageFont.truetype('C:\\Windows\\Fonts\\malgunbd.ttf', 24)
            font_lg = ImageFont.truetype('C:\\Windows\\Fonts\\malgunbd.ttf', 36)
        except Exception:
            font = font_md = font_lg = ImageFont.load_default()

        # ── 캘리브레이션 라인 오버레이 ────────────────────────
        _px_mm = getattr(self, '_px_per_mm', 0.0)
        _lp1x = int(self._cal_line_p1[0] * iw)
        _lp1y = int(self._cal_line_p1[1] * ih)
        _lp2x = int(self._cal_line_p2[0] * iw)
        _lp2y = int(self._cal_line_p2[1] * ih)
        _lclr = (0, 210, 255) if _px_mm > 0 else (130, 140, 150)
        draw.line([(_lp1x, _lp1y), (_lp2x, _lp2y)], fill=_lclr, width=2)
        _lang = _math.atan2(_lp2y - _lp1y, _lp2x - _lp1x) + _math.pi / 2
        for _ex, _ey in ((_lp1x, _lp1y), (_lp2x, _lp2y)):
            _tx, _ty = int(_math.cos(_lang)*10), int(_math.sin(_lang)*10)
            draw.line([(_ex-_tx, _ey-_ty), (_ex+_tx, _ey+_ty)], fill=_lclr, width=2)
            draw.ellipse([_ex-8, _ey-8, _ex+8, _ey+8], outline=_lclr, width=2)
        if _px_mm > 0:
            _lpx = _math.sqrt((_lp2x-_lp1x)**2 + (_lp2y-_lp1y)**2)
            _lmm = _lpx / _px_mm
            draw.text((min(_lp1x, _lp2x), min(_lp1y, _lp2y) - 30),
                      f'하우징 {_lmm:.2f} mm', fill=_lclr, font=font_md)

        # ── 이탈 측정선 + 수치 오버레이 ──────────────────────
        _m_pts = getattr(self, '_measure_pts', [])
        _dev   = getattr(self, '_deviation_val', 0.0)
        if len(_m_pts) == 2 and _dev > 0:
            _m_clr = (0, 255, 150)
            _mp1x = int(_m_pts[0][0] * iw)
            _mp1y = int(_m_pts[0][1] * ih)
            _mp2x = int(_m_pts[1][0] * iw)
            _mp2y = int(_m_pts[1][1] * ih)

            # 측정선
            draw.line([(_mp1x, _mp1y), (_mp2x, _mp2y)], fill=_m_clr, width=3)
            # 끝점 마커
            for _ex, _ey, _pl in ((_mp1x, _mp1y, 'P1'), (_mp2x, _mp2y, 'P2')):
                draw.ellipse([_ex-10, _ey-10, _ex+10, _ey+10],
                              outline=_m_clr, width=2)
                draw.line([_ex-16, _ey, _ex+16, _ey], fill=_m_clr, width=1)
                draw.line([_ex, _ey-16, _ex, _ey+16], fill=_m_clr, width=1)
                draw.text((_ex+14, _ey-22), _pl, fill=_m_clr, font=font_md)

            # 수치 — 측정선 중앙에 크게 표시
            _cx = (_mp1x + _mp2x) // 2
            _cy = (_mp1y + _mp2y) // 2
            _rtxt = f'{_dev:.2f} mm'
            # 배경 박스 (텍스트 크기 추정)
            _bw = len(_rtxt) * 20 + 16
            draw.rectangle([_cx - _bw//2, _cy - 44, _cx + _bw//2, _cy + 6],
                           fill=(0, 25, 15))
            draw.rectangle([_cx - _bw//2, _cy - 44, _cx + _bw//2, _cy + 6],
                           outline=_m_clr, width=1)
            draw.text((_cx - _bw//2 + 8, _cy - 42),
                      _rtxt, fill=(0, 255, 150), font=font_lg)

        # ── 1mm 스케일 바 (우하단) ───────────────────────────
        if _px_mm > 0:
            _sb2_px, _sb2_mm = _nice_scale(_px_mm, target_px=100)
            _sb2_x0 = iw - SCALE_MARGIN - _sb2_px
            _sb2_x1 = iw - SCALE_MARGIN
            _sb2_y  = ih - SCALE_MARGIN - 10
            draw.line([(_sb2_x0, _sb2_y), (_sb2_x1, _sb2_y)],
                       fill=(0, 180, 220), width=3)
            for _ex in (_sb2_x0, _sb2_x1):
                draw.line([(_ex, _sb2_y-8), (_ex, _sb2_y+8)],
                           fill=(0, 180, 220), width=3)
            _sb2_lbl = (f'{_sb2_mm*1000:.0f}μm' if _sb2_mm < 0.1
                         else f'{_sb2_mm:.2f}mm')
            draw.text((_sb2_x0, _sb2_y - 30),
                      _sb2_lbl, fill=(0, 180, 220), font=font_md)

        # ── 하단 정보 바 ──────────────────────────────────────
        cnt    = self.count.get()
        fibers = self.fiber_count.get()
        rate   = cnt / fibers * 100 if fibers > 0 else 0
        overlay = (
            f'LOT: {lot or "-"}  SN: {serial or "-"}  Model: {model}  '
            f'홀막힘: {cnt}개  불량률: {rate:.2f}%  '
            f'이탈: {_dev:.2f}mm  [{now.strftime("%Y-%m-%d %H:%M:%S")}]'
        )
        draw.rectangle([0, img.height - 36, img.width, img.height],
                       fill=(0, 0, 0))
        draw.text((8, img.height - 30), overlay, fill=(74, 159, 212), font=font)

        try:
            os.makedirs(save_dir, exist_ok=True)
            img.save(fname, quality=95)
        except Exception as ex:
            messagebox.showerror(
                '캡처 저장 실패',
                'LOT/일련번호에 사용할 수 없는 문자가 포함되었거나 '
                '디스크 오류로 화면 캡처를 저장하지 못했습니다.\n\n'
                f'상세 정보: {ex}')
            return

        self.lbl_flash.configure(
            text=f'캡처 저장: {lot_folder}/{os.path.basename(fname)}', fg='#4a9fd4')
        self.after(3000, lambda: self.lbl_flash.configure(text=''))

    # ── 이력 저장 ─────────────────────────────────────────────────────────────

    def _save_history(self, model, fibers, cnt, rate, verdict, defect_reason=''):
        h = self._hint_active
        entry = {
            'date':     datetime.datetime.now().strftime('%Y-%m-%d'),
            'time':     datetime.datetime.now().strftime('%H:%M:%S'),
            'lot':      '' if h.get(id(self.lot_no))    else self.lot_no.get(),
            'operator': '' if h.get(id(self.operator))  else self.operator.get(),
            'serial':   '' if h.get(id(self.serial_no)) else self.serial_no.get(),
            'model':    model,
            'fibers':   fibers,
            'count':    cnt,
            'rate':     round(rate, 2),
            'verdict':    verdict,
            'deviation':  self._deviation_val,
            'defect_reason': defect_reason,
        }
        new_id = None
        try:
            new_id = _db_insert(entry)
        except Exception as ex:
            messagebox.showerror(
                '이력 저장 실패',
                '검사 이력 데이터베이스에 결과를 저장하지 못했습니다.\n'
                '디스크 용량이 부족하거나 파일이 잠겨있는지 확인 후 다시 시도하세요.\n\n'
                f'파일 위치: {HISTORY_DB_FILE}\n상세 정보: {ex}')
        self._load_total_products()
        self._refresh_summary()
        self._refresh_suggestion_cache()
        return new_id

    def _load_total_products(self):
        try:
            self.total_products.set(_db_count_all())
        except Exception:
            self.total_products.set(0)

    def _load_config(self):
        if not os.path.exists(CONFIG_FILE):
            return
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            self.lot_no.set(cfg.get('lot', ''))
            self.operator.set(cfg.get('operator', ''))
            self.serial_no.set(cfg.get('serial', ''))
            self._lot_recent      = list(cfg.get('lot_recent', []))[:3]
            self._serial_recent   = list(cfg.get('serial_recent', []))[:3]
            self._operator_roster = list(cfg.get('operator_roster', []))
            self.count_key        = cfg.get('count_key', 'space')
            self._locked_brightness = cfg.get('locked_brightness')
            self._locked_contrast   = cfg.get('locked_contrast')
            self._cut_lot_recent    = list(cfg.get('cut_lot_recent', []))[:10]
            self._cut_serial_recent = list(cfg.get('cut_serial_recent', []))[:10]
            self._cut_blade_recent  = list(cfg.get('cut_blade_recent', []))[:10]
        except Exception:
            pass

    def _save_config(self):
        try:
            h = self._hint_active
            cfg = {
                'lot':      '' if h.get(id(self.lot_no))    else self.lot_no.get(),
                'operator': '' if h.get(id(self.operator))  else self.operator.get(),
                'serial':   '' if h.get(id(self.serial_no)) else self.serial_no.get(),
                'lot_recent':      self._lot_recent,
                'serial_recent':   self._serial_recent,
                'operator_roster': self._operator_roster,
                'count_key':       getattr(self, 'count_key', 'space'),
                'locked_brightness': self._locked_brightness,
                'locked_contrast':   self._locked_contrast,
                'cut_lot_recent':    self._cut_lot_recent,
                'cut_serial_recent': self._cut_serial_recent,
                'cut_blade_recent':  self._cut_blade_recent,
            }
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # ── LOT/작업자/일련번호 자동완성 ─────────────────────────────────────────

    def _push_recent(self, kind, value):
        """LOT/일련번호 입력값을 '직전 3건' 드롭다운 목록 맨 앞에 기록.
        작업자는 [관리] 버튼으로 직접 관리하는 고정 목록이라 여기서 다루지 않는다."""
        value = (value or '').strip()
        if not value:
            return
        lst = {'lot': self._lot_recent, 'serial': self._serial_recent}.get(kind)
        if lst is None:
            return
        if value in lst:
            lst.remove(value)
        lst.insert(0, value)
        del lst[3:]
        self._save_config()

    def _open_operator_roster_dialog(self):
        """작업자 고정 목록 관리 창: 추가 / 이름 변경 / 삭제."""
        dlg = tk.Toplevel(self)
        dlg.title('작업자 목록 관리')
        dlg.configure(bg=CARD_BG)
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)

        tk.Label(dlg, text='작업자 목록', font=('맑은 고딕', 10, 'bold'),
                 bg=CARD_BG, fg='#ffffff').pack(anchor='w', padx=10, pady=(10, 4))

        body = tk.Frame(dlg, bg=CARD_BG)
        body.pack(padx=10, pady=(0, 6))

        lb = tk.Listbox(body, font=('맑은 고딕', 10), width=22, height=8,
                         bg='#21262d', fg=TXT_W, selectbackground='#2c5f9e',
                         selectforeground='#ffffff', relief='flat',
                         highlightthickness=1, highlightbackground='#3a4f60',
                         activestyle='none', exportselection=False)
        lb.pack(side='left', fill='y')
        sb = tk.Scrollbar(body, command=lb.yview)
        sb.pack(side='left', fill='y')
        lb.configure(yscrollcommand=sb.set)

        def reload_list():
            lb.delete(0, 'end')
            for name in self._operator_roster:
                lb.insert('end', name)

        reload_list()

        entry_row = tk.Frame(dlg, bg=CARD_BG)
        entry_row.pack(fill='x', padx=10, pady=(0, 6))
        name_var = tk.StringVar(value='')
        name_entry = tk.Entry(entry_row, textvariable=name_var, font=('맑은 고딕', 10),
                               bg='#21262d', fg=TXT_W, insertbackground=TXT_W,
                               relief='flat', bd=2,
                               highlightthickness=2, highlightbackground='#30363d',
                               highlightcolor=ACC_BLU)
        name_entry.pack(side='left', fill='x', expand=True, ipady=2)

        def add_name():
            name = name_var.get().strip()
            if not name:
                return
            if name in self._operator_roster:
                messagebox.showinfo('안내', '이미 목록에 있는 이름입니다.', parent=dlg)
                return
            self._operator_roster.append(name)
            self._save_config()
            name_var.set('')
            reload_list()

        def rename_selected():
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo('안내', '변경할 이름을 목록에서 먼저 선택하세요.', parent=dlg)
                return
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showinfo('안내', '위 입력칸에 새 이름을 입력한 뒤 [이름변경]을 누르세요.',
                                     parent=dlg)
                return
            old_name = self._operator_roster[sel[0]]
            if new_name != old_name and new_name in self._operator_roster:
                messagebox.showinfo('안내', '이미 목록에 있는 이름입니다.', parent=dlg)
                return
            self._operator_roster[sel[0]] = new_name
            if self.operator.get() == old_name:
                self.operator.set(new_name)
            self._save_config()
            name_var.set('')
            reload_list()

        def delete_selected():
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo('안내', '삭제할 이름을 목록에서 먼저 선택하세요.', parent=dlg)
                return
            name = self._operator_roster[sel[0]]
            if not messagebox.askyesno('삭제 확인', f'"{name}" 을(를) 목록에서 삭제할까요?',
                                        parent=dlg):
                return
            del self._operator_roster[sel[0]]
            self._save_config()
            reload_list()

        def on_pick(_ev=None):
            sel = lb.curselection()
            if sel:
                name_var.set(lb.get(sel[0]))
        lb.bind('<<ListboxSelect>>', on_pick)

        btn_row = tk.Frame(dlg, bg=CARD_BG)
        btn_row.pack(fill='x', padx=10, pady=(0, 10))
        for text, cmd, bg in (('추가', add_name, '#1a5c2e'),
                               ('이름변경', rename_selected, '#1a3a6b'),
                               ('삭제', delete_selected, '#5c1a1a'),
                               ('닫기', dlg.destroy, '#2a3540')):
            b = tk.Button(btn_row, text=text, font=('맑은 고딕', 9, 'bold'),
                          bg=bg, fg='#ffffff', relief='flat', cursor='hand2',
                          padx=8, pady=3, command=cmd)
            b.pack(side='left', padx=(0, 4))
            _setup_hover(b, bg, bg)

    def _refresh_suggestion_cache(self):
        """이력 DB에서 LOT/작업자 자동완성 후보 및 직전 3건 초기값을 갱신."""
        lots, ops = set(), set()
        try:
            records = _db_query_all(order='id DESC')   # 최신순
        except Exception:
            records = []
        for h in records:
            if h.get('lot'):
                lots.add(h['lot'])
            if h.get('operator'):
                ops.add(h['operator'])
        self._lot_suggestions      = sorted(lots)
        self._operator_suggestions = sorted(ops)

        def _recent_from_history(key):
            seen, out = [], []
            for h in sorted(records, key=lambda r: (r.get('date', ''), r.get('time', '')),
                             reverse=True):
                v = (h.get(key) or '').strip()
                if v and v not in seen:
                    seen.append(v)
                    out.append(v)
                if len(out) >= 3:
                    break
            return out

        # 설정 파일에 저장된 직전 입력값이 아직 없을 때만 이력에서 초기값을 채운다.
        if not self._lot_recent:
            self._lot_recent = _recent_from_history('lot')
        if not self._serial_recent:
            self._serial_recent = _recent_from_history('serial')
        # 작업자 고정 목록도 아직 한 번도 구성된 적 없으면 이력에서 초기값을 채운다.
        if not self._operator_roster:
            self._operator_roster = sorted(ops)

    def _setup_autocomplete(self, entry, var, kind):
        """entry에 '직전 3건' 드롭다운(팝업 리스트)을 붙인다.
        kind: 'lot' / 'operator' / 'serial' — 어떤 최근 입력값 목록을 쓸지 결정."""
        state = {'win': None, 'listbox': None}

        def hide():
            if state['win'] is not None:
                try:
                    state['win'].destroy()
                except Exception:
                    pass
                state['win'] = None
                state['listbox'] = None

        def select(value):
            var.set(value)
            self._hint_active[id(var)] = False
            hide()
            entry.icursor('end')
            entry.focus_set()

        def show(matches):
            hide()
            if not matches:
                return
            win = tk.Toplevel(self)
            win.wm_overrideredirect(True)
            win.attributes('-topmost', True)
            x = entry.winfo_rootx()
            y = entry.winfo_rooty() + entry.winfo_height()
            win.geometry(f'+{x}+{y}')
            frm = tk.Frame(win, bg='#21262d')
            frm.pack()
            lb = tk.Listbox(frm, font=('맑은 고딕', 9), height=min(6, len(matches)),
                             bg='#21262d', fg=TXT_W, selectbackground='#2c5f9e',
                             selectforeground='#ffffff', relief='flat',
                             highlightthickness=1, highlightbackground='#3a4f60',
                             activestyle='none', exportselection=False)
            for m in matches:
                lb.insert('end', m)
            lb.pack(side='left')
            if len(matches) > 6:
                sb = tk.Scrollbar(frm, command=lb.yview)
                sb.pack(side='left', fill='y')
                lb.configure(yscrollcommand=sb.set)

            def on_pick(_ev=None):
                sel = lb.curselection()
                if sel:
                    select(lb.get(sel[0]))
            lb.bind('<ButtonRelease-1>', on_pick)
            lb.bind('<Return>', on_pick)
            state['win']     = win
            state['listbox'] = lb

        def pool():
            if kind == 'operator':
                return self._operator_roster
            return {'lot': self._lot_recent, 'serial': self._serial_recent}.get(kind, [])

        def on_key(ev):
            if ev.keysym == 'Escape':
                hide()
                return
            if ev.keysym == 'Down' and state['listbox'] is not None:
                state['listbox'].focus_set()
                state['listbox'].selection_set(0)
                return
            if ev.keysym in ('Up', 'Return', 'Tab', 'Shift_L', 'Shift_R',
                              'Control_L', 'Control_R', 'space'):
                return
            if self._hint_active.get(id(var)):
                hide()
                return
            typed = var.get().strip()
            p = pool()
            if not typed:
                matches = p
            else:
                matches = [c for c in p if typed.lower() in c.lower() and c != typed]
            show(matches)

        def on_focus_show(ev=None):
            p = pool()
            if p:
                show(p)

        entry.bind('<KeyRelease>', on_key, add='+')
        entry.bind('<FocusIn>', on_focus_show, add='+')
        entry.bind('<FocusOut>', lambda e: self.after(150, hide), add='+')

    # ── 모델 선택 ─────────────────────────────────────────────────────────────

    def _select_model(self, name, fibers=None):
        if fibers is None:
            fibers = MODELS.get(name, 0)
        if self._model_selected and self.count.get() > 0:
            if not messagebox.askyesno('모델 변경',
                                        f'카운트 {self.count.get()}건이 있습니다.\n'
                                        f'[{name}]으로 변경 시 초기화됩니다.'):
                return
            self._do_reset()

        self.selected_model.set(name)
        self.fiber_count.set(fibers)
        self._alarm_done = False
        self._model_selected = True

        # 드롭다운 동기화
        if hasattr(self, '_model_dd_var'):
            self._model_dd_var.set(name)
        if hasattr(self, '_model_dd_wrap'):
            self._model_dd_wrap.configure(bg='#2e8b57')   # 선택 완료 → 초록 테두리

        fibers = MODELS[name]
        limit  = int(DEFECT_LIMIT / 100 * fibers)
        suffix = 'H' if name.endswith('H') else 'L'
        clr    = MODEL_CLR[suffix]['select']

        self.lbl_model_info.configure(
            text=f'▶  {name}  |  파이버: {fibers:,}EA  |  기준: {limit}개 ({DEFECT_LIMIT}%)')
        self._update_rate()
        self.lbl_flash.configure(
            text=f'✅  {name} 선택 ({fibers:,}EA, 기준 {limit}개)  — 카운팅을 시작하세요',
            fg=clr)
        self.after(2500, lambda: self.lbl_flash.configure(text=''))

    # ── 이력 조회 ─────────────────────────────────────────────────────────────

    def _open_history(self):
        HistoryWindow(self)

    # ── 카메라 ────────────────────────────────────────────────────────────────

    def _start_camera(self):
        threading.Thread(target=self._cam_loop, daemon=True).start()

    def _cam_loop(self):
        my_gen = self._cam_gen
        if self._cam_mode == 'dino_window':
            self._cam_loop_window()
            return

        def _still_active():
            return (self.running and self._cam_mode == 'direct'
                    and self._cam_gen == my_gen)

        retry_shown = False
        # 바깥 루프: 카메라가 없거나 끊기면 자동으로 재연결(핸들 재생성)을 계속 시도
        while _still_active():
            cam = cv2.VideoCapture(self.cam_idx.get(), cv2.CAP_DSHOW)
            cam.set(cv2.CAP_PROP_FRAME_WIDTH,  CAM_W)
            cam.set(cv2.CAP_PROP_FRAME_HEIGHT, CAM_H)
            cam.set(cv2.CAP_PROP_FPS, FPS_LIMIT)

            if not cam.isOpened():
                cam.release()
                if not retry_shown:
                    retry_shown = True
                    self.after(0, lambda: self.lbl_cam_status.configure(
                        text='● 카메라 없음 (자동 재연결 시도 중...)', fg=ACC_RED))
                time.sleep(3.0)
                continue

            self.after(0, lambda: self.lbl_cam_status.configure(text='● 연결됨', fg='#4caf50'))
            retry_shown = False

            fail_streak = 0
            disconnected_shown = False
            # self._cam_gen 이 바뀌면(포트 재연결 요청) 이 루프는 즉시 스스로 종료하고
            # cam.release() 한다 — 재연결 시 이전 스레드/카메라 핸들이 누적되는 것을 방지
            while _still_active():
                ret, frame = cam.read()
                if ret:
                    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    with self._frame_lock:
                        self._frame = frame
                    if disconnected_shown:
                        disconnected_shown = False
                        self.after(0, lambda: self.lbl_cam_status.configure(
                            text='● 연결됨', fg='#4caf50'))
                    fail_streak = 0
                else:
                    fail_streak += 1
                    # 약 3초(6회) 연속 프레임 수신 실패 시 알림 후, 내부 루프를 빠져나가
                    # 바깥 루프에서 카메라 핸들을 새로 열어 자동 재연결을 시도
                    if fail_streak >= 6:
                        if not disconnected_shown:
                            disconnected_shown = True
                            self.after(0, lambda: self.lbl_cam_status.configure(
                                text='● 연결 끊김 (자동 재연결 시도 중...)', fg=ACC_RED))
                        break
                    time.sleep(0.5)
            cam.release()
            if disconnected_shown and _still_active():
                time.sleep(2.0)

    def _cam_loop_window(self):
        from PIL import ImageGrab
        self.after(0, lambda: self.lbl_cam_status.configure(
            text='● DinoCapture 창 캡처', fg='#00bcd4'))
        _warned = False
        while self.running and self._cam_mode == 'dino_window':
            rect = self._find_dino_window_rect()
            if rect:
                try:
                    img = ImageGrab.grab(bbox=rect)
                    frame = np.array(img)
                    with self._frame_lock:
                        self._frame = frame
                    if _warned:
                        _warned = False
                        self.after(0, lambda: self.lbl_cam_status.configure(
                            text='● DinoCapture 창 캡처', fg='#00bcd4'))
                except Exception:
                    pass
            else:
                if not _warned:
                    _warned = True
                    self.after(0, lambda: self.lbl_cam_status.configure(
                        text='● DinoCapture 창 없음', fg=ACC_YEL))
                time.sleep(1.0)
                continue
            time.sleep(1.0 / FPS_LIMIT)

    def _find_dino_window_rect(self):
        """DinoCapture 창 화면 좌표 (left, top, right, bottom) 반환. 없으면 None."""
        result = []
        WNDENUMPROC = ctypes.WINFUNCTYPE(
            ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)

        def _cb(hwnd, _):
            if ctypes.windll.user32.IsWindowVisible(hwnd):
                length = ctypes.windll.user32.GetWindowTextLengthW(hwnd)
                buf = ctypes.create_unicode_buffer(length + 1)
                ctypes.windll.user32.GetWindowTextW(hwnd, buf, length + 1)
                title = buf.value.lower()
                if 'dinocapture' in title or ('dino' in title and 'bridge' not in title):
                    rect = ctypes.wintypes.RECT()
                    ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
                    w = rect.right - rect.left
                    h = rect.bottom - rect.top
                    if w > 200 and h > 200:
                        result.append((rect.left, rect.top, rect.right, rect.bottom))
            return True

        ctypes.windll.user32.EnumWindows(WNDENUMPROC(_cb), 0)
        return result[0] if result else None

    def _toggle_cam_mode(self):
        if self._cam_mode == 'direct':
            self._cam_mode = 'dino_window'
            self._btn_cam_mode.configure(bg='#1a5c2e', fg='#fff',
                                          text='🖥 DinoCapture 창 캡처  ON')
            threading.Thread(target=self._cam_loop, daemon=True).start()
        else:
            self._cam_mode = 'direct'
            self._btn_cam_mode.configure(bg='#21262d', fg=TXT_G,
                                          text='🖥 DinoCapture 창 캡처')
            threading.Thread(target=self._cam_loop, daemon=True).start()

    def _toggle_exposure_lock(self):
        """DNX64 SDK로 자동노출을 끄고 현재 노출값+AE타겟값에 고정(또는 해제)한다.
        매 프레임 자동노출이 흔들려서 측정 문턱값이 오락가락하는 것을 방지.
        DinoCapture 없이 카운터 앱 자체 카메라 스트림만으로 동작 확인됨.
        ExposureValue, AETarget 둘 다 Set 후 readback이 실제로 바뀌는 것을
        확인함(2026-07-16).

        실제 DLL 호출은 전부 백그라운드 스레드에서 수행한다 — 이 장비의 USB
        통신이 느리거나 불안정할 때 메인(UI) 스레드에서 동기 호출하면 앱 전체가
        멈춘 것처럼 보이는 문제가 있었다(2026-07-21, LED 건에서 먼저 발견되어
        여기도 같은 방식으로 함께 고침). 버튼 상태는 클릭 즉시(낙관적으로) 바꾸고,
        DLL 호출이 실패하면 그때 되돌린다."""
        dll = _DNX64.get()
        if dll is None:
            self.lbl_flash.configure(text='DNX64 SDK를 사용할 수 없습니다', fg=ACC_YEL)
            self.after(2000, lambda: self.lbl_flash.configure(text=''))
            return
        idx = _DNX64._idx
        if not self._exposure_locked:
            self._exposure_locked = True
            self._btn_exposure_lock.configure(text='🔒 노출 고정 ON')
            self._refresh_sdk_button_colors()

            def _worker():
                try:
                    with _DNX64.lock:
                        current_exp = dll.GetExposureValue(idx)
                        current_ae  = dll.GetAETarget(idx)
                        dll.SetAutoExposure(idx, 0)
                        time.sleep(0.05)
                        dll.SetExposureValue(idx, current_exp)
                        dll.SetAETarget(idx, current_ae)
                except Exception:
                    def _fail():
                        self._exposure_locked = False
                        self._btn_exposure_lock.configure(text='🔒 노출 고정')
                        self._refresh_sdk_button_colors()
                        self.lbl_flash.configure(text='노출 고정 실패', fg=ACC_YEL)
                        self.after(2000, lambda: self.lbl_flash.configure(text=''))
                    self.after(0, _fail)
            threading.Thread(target=_worker, daemon=True).start()
        else:
            self._exposure_locked = False
            self._btn_exposure_lock.configure(text='🔒 노출 고정')
            self._refresh_sdk_button_colors()

            def _worker():
                try:
                    with _DNX64.lock:
                        dll.SetAutoExposure(idx, 1)
                except Exception:
                    pass
            threading.Thread(target=_worker, daemon=True).start()

    def _toggle_led(self):
        """DNX64 SDK로 LED를 켠다/끈다. SetLEDState(1)은 몇 초 후 장치가 자동으로
        꺼버리는 것을 확인해서(원인 미상), 켜져있는 동안 2초마다 재전송(keep-alive)
        하는 백그라운드 스레드로 계속 켜진 상태를 유지한다.

        이 함수는 DLL을 절대 메인(UI) 스레드에서 직접 호출하지 않는다 — 예전엔
        동기 호출이라 이 장비의 USB 응답이 느릴 때 앱 전체가 멈춘 것처럼 보이는
        문제가 있었다(2026-07-21). 버튼 상태는 클릭 즉시(낙관적으로) 바꾸고,
        실제 SetLEDState 호출은 전부 백그라운드 스레드에서 수행한다.

        keep-alive 스레드가 "켜져라" 신호를 보내려는 순간과 사용자의 OFF 클릭이
        겹치면, OFF 신호가 먼저 나간 직후 스레드의 "켜져라" 신호가 뒤늦게 도착해
        실제로는 꺼지지 않는 경합이 있었다(같은 날 발견). "세대 번호(generation)"로
        해결: OFF를 누르면 번호를 올리고, keep-alive는 "보내기 직전, 락을 잡은
        상태에서" 자기 세대 번호가 아직 최신인지 재확인 후에만 전송한다. 두 신호가
        같은 락으로 순서가 매겨지므로 뒤늦게 도착하는 낡은 '켜져라' 신호는 항상
        무시된다."""
        dll = _DNX64.get()
        if dll is None:
            self.lbl_flash.configure(text='DNX64 SDK를 사용할 수 없습니다', fg=ACC_YEL)
            self.after(2000, lambda: self.lbl_flash.configure(text=''))
            return
        idx = _DNX64._idx
        if not self._led_on:
            self._led_generation += 1
            my_gen = self._led_generation
            self._led_on = True
            self._btn_led.configure(text='💡 LED ON')
            self._refresh_sdk_button_colors()

            def _worker():
                try:
                    with _DNX64.lock:
                        dll.SetLEDState(idx, 1)
                except Exception:
                    def _fail():
                        self._led_on = False
                        self._btn_led.configure(text='💡 LED')
                        self._refresh_sdk_button_colors()
                        self.lbl_flash.configure(text='LED 제어 실패', fg=ACC_YEL)
                        self.after(2000, lambda: self.lbl_flash.configure(text=''))
                    self.after(0, _fail)
                    return

                def _keepalive():
                    while getattr(self, 'running', True) and self._led_on:
                        time.sleep(2.0)
                        try:
                            with _DNX64.lock:
                                # 락을 잡은 이 순간에도 여전히 내 세대가 최신일
                                # 때만 전송 — OFF가 그 사이 눌렸다면(세대가
                                # 바뀌었다면) 이 낡은 신호는 조용히 무시된다.
                                if my_gen == self._led_generation:
                                    dll.SetLEDState(idx, 1)
                        except Exception:
                            pass
                threading.Thread(target=_keepalive, daemon=True).start()
            threading.Thread(target=_worker, daemon=True).start()
        else:
            self._led_on = False
            self._led_generation += 1   # keep-alive의 낡은 전송을 즉시 무효화
            self._btn_led.configure(text='💡 LED')
            self._refresh_sdk_button_colors()

            def _worker():
                try:
                    with _DNX64.lock:
                        dll.SetLEDState(idx, 0)
                except Exception:
                    pass
            threading.Thread(target=_worker, daemon=True).start()

    def _lock_quality_now(self):
        """DNX64 SDK로 현재 하드웨어 밝기/대비값(VideoProcAmp)을 그대로 읽어서
        설정파일에 저장해둔다. 다음 실행 시(_apply_locked_quality) 자동으로
        같은 값을 재적용해서, OS/드라이버가 기본값을 다르게 잡거나 다른 프로그램이
        건드려놔도 항상 같은 화질로 검사하게 한다. Set 후 readback이 실제로
        바뀌는 것을 확인함(2026-07-16, VideoProcAmp index 0=밝기).

        DLL 읽기는 백그라운드 스레드에서 하고, 결과를 받은 뒤에만 self.after(0, ...)
        로 메인 스레드에 넘겨 위젯/설정파일을 갱신한다 — 이 장비의 USB 응답이
        느릴 때 메인 스레드가 멈춘 것처럼 보이는 문제를 막기 위함(2026-07-21)."""
        dll = _DNX64.get()
        if dll is None:
            self.lbl_flash.configure(text='DNX64 SDK를 사용할 수 없습니다', fg=ACC_YEL)
            self.after(2000, lambda: self.lbl_flash.configure(text=''))
            return

        def _worker():
            try:
                with _DNX64.lock:
                    brightness = dll.GetVideoProcAmp(0)
                    contrast   = dll.GetVideoProcAmp(1)
            except Exception:
                def _fail():
                    self.lbl_flash.configure(text='화질 고정 실패', fg=ACC_YEL)
                    self.after(2000, lambda: self.lbl_flash.configure(text=''))
                self.after(0, _fail)
                return

            def _done():
                self._locked_brightness = brightness
                self._locked_contrast   = contrast
                self._save_config()
                self._btn_quality_lock.configure(text='🎨 화질 고정됨')
                self._refresh_sdk_button_colors()
                self.lbl_flash.configure(
                    text=f'현재 화질(밝기{brightness}/대비{contrast})로 고정 — 재시작해도 유지됨',
                    fg='#4a9fd4')
                self.after(2500, lambda: self.lbl_flash.configure(text=''))
            self.after(0, _done)

        threading.Thread(target=_worker, daemon=True).start()

    def _apply_locked_quality(self):
        """앱 시작 시 저장된 밝기/대비 고정값이 있으면 하드웨어에 재적용.
        DNX64 로딩이 블로킹이라 백그라운드 스레드에서 호출됨 — 위젯 갱신은
        반드시 self.after(0, ...)로 메인 스레드에 넘겨야 한다."""
        if self._locked_brightness is None and self._locked_contrast is None:
            return
        dll = _DNX64.get()
        if dll is None:
            return
        try:
            with _DNX64.lock:
                if self._locked_brightness is not None:
                    dll.SetVideoProcAmp(0, self._locked_brightness)
                if self._locked_contrast is not None:
                    dll.SetVideoProcAmp(1, self._locked_contrast)
            if hasattr(self, '_btn_quality_lock'):
                self.after(0, lambda: self._btn_quality_lock.configure(text='🎨 화질 고정됨'))
                self.after(0, self._refresh_sdk_button_colors)
        except Exception:
            pass

    def _reconnect_cam(self):
        self.lbl_cam_status.configure(text='● 재연결 중...', fg=ACC_YEL)
        # 세대 번호를 올려 기존 _cam_loop 스레드가 다음 프레임 확인 시점에
        # 스스로 종료하고 자신의 카메라 핸들을 release 하도록 함
        self._cam_gen += 1
        threading.Thread(target=self._cam_loop, daemon=True).start()

    # ── 하우징 자동 배율 감지 ─────────────────────────────────────────────────

    # ── 돋보기 ────────────────────────────────────────────────────────────────
    def _toggle_magnifier(self):
        self._magnifier_on = not self._magnifier_on
        if self._magnifier_on:
            self.btn_magnifier.configure(bg='#1a4a7a', fg='#ffffff', text='돋보기  ON')
        else:
            self.btn_magnifier.configure(bg='#21262d', fg='#8b949e', text='돋보기 OFF')

    def _show_magnifier_popup(self, nx: float, ny: float, label: str = '',
                               auto_close: bool = True):
        """클릭 위치를 2배 확대한 팝업 창 표시. auto_close=False 이면 수동 닫기 전까지 유지."""
        with self._frame_lock:
            frame = self._frame.copy() if self._frame is not None else None
        if frame is None:
            return

        h, w = frame.shape[:2]
        px, py = int(nx * w), int(ny * h)
        half = 90
        x1 = max(0, px - half);  y1 = max(0, py - half)
        x2 = min(w, px + half);  y2 = min(h, py + half)
        crop = frame[y1:y2, x1:x2].copy()

        # 2배 확대
        zoom = cv2.resize(crop, (crop.shape[1] * 2, crop.shape[0] * 2),
                          interpolation=cv2.INTER_LINEAR)
        # 클릭 지점 십자선 (초록)
        cw_px, ch_px = (px - x1) * 2, (py - y1) * 2
        cv2.line(zoom, (cw_px - 28, ch_px), (cw_px + 28, ch_px), (0, 255, 80), 1)
        cv2.line(zoom, (cw_px, ch_px - 28), (cw_px, ch_px + 28), (0, 255, 80), 1)
        cv2.circle(zoom, (cw_px, ch_px), 7, (0, 255, 80), 1)

        # 기존 팝업 닫기
        try:
            if self._magnifier_popup and self._magnifier_popup.winfo_exists():
                self._magnifier_popup.destroy()
        except Exception:
            pass

        popup = tk.Toplevel(self)
        popup.title(f'돋보기 2×  {label}')
        popup.resizable(False, False)
        popup.configure(bg='#111111')
        popup.attributes('-topmost', True)
        self._magnifier_popup = popup

        img   = Image.fromarray(zoom)
        photo = ImageTk.PhotoImage(img)
        lbl_img = tk.Label(popup, image=photo, bg='#000000', bd=0)
        lbl_img.image = photo
        lbl_img.pack()

        _hint = '3초 후 자동 닫힘' if auto_close else 'P2 클릭 시 자동 닫힘'
        tk.Label(popup, text=f'{label}  위치 2× 확대     {_hint}',
                 font=('맑은 고딕', 8), bg='#111111', fg='#00ccff').pack(pady=2)
        tk.Button(popup, text='닫기', command=popup.destroy,
                  font=('맑은 고딕', 8), bg='#333333', fg='#ffffff',
                  relief='flat', padx=8).pack(pady=(0, 4))

        # 클릭 위치 근처에 팝업 배치 (화면 밖 나가지 않도록)
        popup.update_idletasks()
        pw2 = popup.winfo_reqwidth()
        ph2 = popup.winfo_reqheight()
        sx  = self.canvas.winfo_rootx() + int(nx * self.canvas.winfo_width())  + 18
        sy  = self.canvas.winfo_rooty() + int(ny * self.canvas.winfo_height()) + 18
        sw  = self.winfo_screenwidth()
        sh  = self.winfo_screenheight()
        sx  = min(sx, sw - pw2 - 10)
        sy  = min(sy, sh - ph2 - 10)
        popup.geometry(f'+{sx}+{sy}')

        if auto_close:
            popup.after(3000, lambda: popup.destroy() if popup.winfo_exists() else None)

    def _update_preview(self):
        if not self.running:
            return
        with self._frame_lock:
            frame = self._frame.copy() if self._frame is not None else None

        if frame is not None:
            beta  = self.cam_brightness.get()
            alpha = self.cam_contrast.get() / 100.0
            if beta != 0 or alpha != 1.0:
                frame = cv2.convertScaleAbs(frame, alpha=alpha, beta=beta)
            h_f, w_f = frame.shape[:2]
            cnt    = self.count.get()
            model  = self.selected_model.get()
            fibers = self.fiber_count.get()
            rate   = cnt / fibers * 100 if fibers > 0 else 0

            # 오버레이
            color = (229, 57, 53) if rate >= DEFECT_LIMIT else (76, 175, 80)
            cv2.rectangle(frame, (0, 0), (w_f, 36), (0, 0, 0), -1)
            suffix = 'H' if (model and model.endswith('H')) else 'L'
            model_color = (255, 107, 107) if suffix == 'H' else (91, 155, 213)

            # 상단 오버레이 텍스트
            if model:
                txt = f"Model: {model}  Fiber: {fibers:,}  Count: {cnt}  {rate:.2f}%  [{'FAIL' if rate >= DEFECT_LIMIT else 'PASS'}]"
            else:
                txt = 'MODEL NOT SELECTED'
            cv2.putText(frame, txt, (8, 26),
                        cv2.FONT_HERSHEY_DUPLEX, 0.7, color, 1)

            # 불량 시 테두리 + 판정 텍스트 표시 (오버레이 없음)
            if rate >= DEFECT_LIMIT:
                cv2.rectangle(frame, (0, 0), (w_f - 1, h_f - 1), (229, 57, 53), 14)
                fail_text = 'FAIL'
                (tw, th), _ = cv2.getTextSize(fail_text, cv2.FONT_HERSHEY_DUPLEX, 2.0, 3)
                cv2.putText(frame, fail_text,
                            (w_f - tw - 20, h_f - 20),
                            cv2.FONT_HERSHEY_DUPLEX, 2.0, (229, 57, 53), 3)
            else:
                pass_text = 'PASS'
                (tw, th), _ = cv2.getTextSize(pass_text, cv2.FONT_HERSHEY_DUPLEX, 1.2, 2)
                cv2.putText(frame, pass_text,
                            (w_f - tw - 20, h_f - 20),
                            cv2.FONT_HERSHEY_DUPLEX, 1.2, (76, 175, 80), 2)

            # 십자선
            cx, cy = w_f // 2, h_f // 2
            cv2.line(frame, (cx - 20, cy), (cx + 20, cy), (255, 255, 255), 1)
            cv2.line(frame, (cx, cy - 20), (cx, cy + 20), (255, 255, 255), 1)
            cv2.circle(frame, (cx, cy), 8, (255, 255, 255), 1)

            # px/mm 고정값 유지 (캘리브레이션 완료 시)
            if self._px_per_mm_baseline > 0:
                self._px_per_mm = self._px_per_mm_baseline

            pw = max(self.canvas.winfo_width(),  400)
            ph = max(self.canvas.winfo_height(), 300)
            img = Image.fromarray(frame).resize((pw, ph), Image.LANCZOS)
            if not hasattr(self, '_preview_font_18'):
                try:
                    self._preview_font_18 = ImageFont.truetype('C:\\Windows\\Fonts\\malgunbd.ttf', 18)
                except Exception:
                    try:
                        self._preview_font_18 = ImageFont.truetype('C:\\Windows\\Fonts\\malgun.ttf', 18)
                    except Exception:
                        self._preview_font_18 = ImageFont.load_default()
            font_pil = self._preview_font_18

            draw = ImageDraw.Draw(img)
            if not model:
                draw.text((pw // 2 - 80, ph - 28),
                          '  카메라를 연결하세요', fill=(139, 148, 158), font=font_pil)

            # ── 캘리브레이션 상태 (우상단, 완료 시만 표시) ────────
            _pnow = getattr(self, '_px_per_mm', 0.0)
            if _pnow > 0:
                draw.rectangle([pw - 250, 4, pw - 4, 28], fill=(14, 18, 24))
                draw.text((pw - 246, 6),
                          f'✓ 캘리브 완료  1mm={_pnow:.1f}px',
                          fill=(80, 220, 100), font=font_pil)

            # ── 드래그 캘리브레이션 라인 ─────────────────────────
            _px_mm   = getattr(self, '_px_per_mm', 0.0)
            _in_cal  = getattr(self, '_cal_mode',  False)
            _in_meas = getattr(self, '_measure_mode', False)
            _c_pts   = getattr(self, '_cal_pts',   [])
            _m_pts   = getattr(self, '_measure_pts', [])

            _p1x = int(self._cal_line_p1[0] * pw)
            _p1y = int(self._cal_line_p1[1] * ph)
            _p2x = int(self._cal_line_p2[0] * pw)
            _p2y = int(self._cal_line_p2[1] * ph)
            _line_dx = _p2x - _p1x
            _line_dy = _p2y - _p1y
            _line_px  = _math.sqrt(_line_dx**2 + _line_dy**2)

            _lclr = ((255, 215, 0)   if _in_cal else
                     (0, 220, 255)   if _px_mm > 0 else
                     (160, 170, 180))

            # 라인 본체
            draw.line([(_p1x, _p1y), (_p2x, _p2y)], fill=_lclr, width=2)
            # 끝점 수직 눈금 (라인에 수직으로)
            _ang = _math.atan2(_line_dy, _line_dx) + _math.pi / 2
            for _ex, _ey in ((_p1x, _p1y), (_p2x, _p2y)):
                _tx, _ty = int(_math.cos(_ang)*10), int(_math.sin(_ang)*10)
                draw.line([(_ex-_tx, _ey-_ty), (_ex+_tx, _ey+_ty)],
                           fill=_lclr, width=2)
            # 중간 눈금
            _mx_, _my_ = (_p1x+_p2x)//2, (_p1y+_p2y)//2
            _tx2, _ty2 = int(_math.cos(_ang)*6), int(_math.sin(_ang)*6)
            draw.line([(_mx_-_tx2, _my_-_ty2), (_mx_+_tx2, _my_+_ty2)],
                       fill=_lclr, width=1)
            # 끝점 서클 (드래그 핸들)
            _ep_r = 12 if (_in_cal or self._line_dragging) else 8
            for _ex, _ey in ((_p1x, _p1y), (_p2x, _p2y)):
                draw.ellipse([_ex-_ep_r, _ey-_ep_r, _ex+_ep_r, _ey+_ep_r],
                              outline=_lclr, width=2)

            # 방향키로 미세조정 중인 활성 핸들 표시 (빨간 점)
            _active = getattr(self, '_line_active', None)
            if _active and not _in_cal and not _in_meas:
                if _active == 'p1':
                    _ax, _ay = _p1x, _p1y
                elif _active == 'p2':
                    _ax, _ay = _p2x, _p2y
                else:
                    _ax, _ay = _mx_, _my_
                draw.ellipse([_ax-4, _ay-4, _ax+4, _ay+4], fill=(255, 60, 60))

            # 라인 레이블
            if _in_cal:
                _ll = f'▼ {len(_c_pts)}/2 끝점 클릭'
            elif _px_mm > 0:
                _ll_mm = _line_px / _px_mm
                _ll = f'{_ll_mm:.2f} mm'
            else:
                _ll = ''   # 캘리브레이션 전: 라인 레이블 숨김
            _lbl_x = min(_p1x, _p2x)
            _lbl_y = min(_p1y, _p2y) - 28
            draw.text((_lbl_x, max(_lbl_y, 4)), _ll, fill=_lclr, font=font_pil)

            # 캘리브레이션 클릭 포인트
            if _in_cal and _c_pts:
                for _i, (cnx, cny) in enumerate(_c_pts):
                    _cx_s, _cy_s = int(cnx * pw), int(cny * ph)
                    draw.ellipse([_cx_s-8, _cy_s-8, _cx_s+8, _cy_s+8],
                                  fill=_lclr)
                    draw.text((_cx_s+10, _cy_s-14),
                               f'P{_i+1}', fill=_lclr, font=font_pil)
                if len(_c_pts) == 2:
                    draw.line([int(_c_pts[0][0]*pw), int(_c_pts[0][1]*ph),
                                int(_c_pts[1][0]*pw), int(_c_pts[1][1]*ph)],
                               fill=_lclr, width=2)

            # ── 1mm 기준 스케일 바 (우하단, 작게) ──────────────
            if _px_mm > 0:
                _sb_px2, _sb_mm2 = _nice_scale(_px_mm, target_px=100)
                _s2lbl = (f'{_sb_mm2*1000:.0f}μm' if _sb_mm2 < 0.1
                           else f'{_sb_mm2:.2f}mm')
                _s2x0  = pw - SCALE_MARGIN - _sb_px2
                _s2x1  = pw - SCALE_MARGIN
                _s2y   = ph - SCALE_MARGIN - 8
                draw.line([(_s2x0, _s2y), (_s2x1, _s2y)], fill=(0, 180, 220), width=2)
                for _ex in (_s2x0, _s2x1):
                    draw.line([(_ex, _s2y-6), (_ex, _s2y+6)], fill=(0, 180, 220), width=2)
                draw.text((_s2x0, _s2y-24), _s2lbl, fill=(0, 180, 220), font=font_pil)

            # ── 이탈 측정 오버레이 ─────────────────────────────
            if _in_meas and _m_pts:
                _m_clr = (0, 220, 255)
                _pts_s = [(int(nx * pw), int(ny * ph)) for nx, ny in _m_pts]
                for _idx, (sx, sy) in enumerate(_pts_s):
                    draw.ellipse([sx - 8, sy - 8, sx + 8, sy + 8],
                                  outline=_m_clr, width=2)
                    draw.line([sx - 14, sy, sx + 14, sy], fill=_m_clr, width=1)
                    draw.line([sx, sy - 14, sx, sy + 14], fill=_m_clr, width=1)
                    draw.text((sx + 10, sy - 18), f'P{_idx + 1}',
                               fill=_m_clr, font=font_pil)
                if len(_pts_s) == 2:
                    draw.line([_pts_s[0][0], _pts_s[0][1],
                                _pts_s[1][0], _pts_s[1][1]], fill=_m_clr, width=2)
                    _mxc = (_pts_s[0][0] + _pts_s[1][0]) // 2
                    _myc = (_pts_s[0][1] + _pts_s[1][1]) // 2
                    _rv  = getattr(self, '_meas_result_var', None)
                    if _rv:
                        draw.text((_mxc + 6, _myc - 22),
                                   _rv.get(), fill=(0, 255, 180), font=font_pil)
                elif len(_pts_s) == 1:
                    draw.text((_pts_s[0][0] + 14, _pts_s[0][1] - 10),
                               '두 번째 점 클릭',
                               fill=_m_clr, font=font_pil)

            # ── 실시간 돋보기 오버레이 ────────────────────────────
            if getattr(self, '_magnifier_on', False):
                _mnx = getattr(self, '_mouse_nx', -1.0)
                _mny = getattr(self, '_mouse_ny', -1.0)
                if 0.0 <= _mnx <= 1.0 and 0.0 <= _mny <= 1.0:
                    _mx  = int(_mnx * pw)
                    _my  = int(_mny * ph)
                    _r   = 50          # 크롭 반경 → 100×100 → 2× → 200×200
                    _cx1 = max(0, _mx - _r);  _cy1 = max(0, _my - _r)
                    _cx2 = min(pw, _mx + _r); _cy2 = min(ph, _my + _r)
                    _crop = img.crop((_cx1, _cy1, _cx2, _cy2))
                    _cw_m = _cx2 - _cx1; _ch_m = _cy2 - _cy1
                    _zw   = _cw_m * 2;   _zh   = _ch_m * 2
                    if _zw > 0 and _zh > 0:
                        _zoom = _crop.resize((_zw, _zh), Image.LANCZOS)
                        _zdraw = ImageDraw.Draw(_zoom)
                        # 십자선
                        _zcx = (_mx - _cx1) * 2
                        _zcy = (_my - _cy1) * 2
                        _zdraw.line([(_zcx-28, _zcy), (_zcx+28, _zcy)],
                                    fill=(0, 255, 80), width=1)
                        _zdraw.line([(_zcx, _zcy-28), (_zcx, _zcy+28)],
                                    fill=(0, 255, 80), width=1)
                        _zdraw.ellipse([_zcx-4, _zcy-4, _zcx+4, _zcy+4],
                                       outline=(0, 255, 80), width=1)
                        # 황금색 원형 테두리
                        _zdraw.ellipse([0, 0, _zw-1, _zh-1],
                                       outline=(255, 215, 0), width=3)
                        # 원형 마스크 적용
                        _mask = Image.new('L', (_zw, _zh), 0)
                        ImageDraw.Draw(_mask).ellipse([0, 0, _zw-1, _zh-1], fill=255)
                        # 배치: 마우스 오른쪽 (화면 우측 벗어나면 왼쪽)
                        _px = _mx + 20
                        _py = _my - _zh // 2
                        if _px + _zw > pw:
                            _px = _mx - _zw - 20
                        _py = max(0, min(ph - _zh, _py))
                        _zoom_rgba = _zoom.convert('RGBA')
                        _zoom_rgba.putalpha(_mask)
                        img.paste(_zoom_rgba, (_px, _py), _zoom_rgba)

            photo = ImageTk.PhotoImage(img)
            self.canvas.create_image(0, 0, anchor='nw', image=photo)
            self.image = photo
            # 캔버스 테두리 색: 불량=빨강, 정상=없음 (즉시 반영)
            if rate >= DEFECT_LIMIT:
                self.canvas.configure(highlightthickness=4,
                                      highlightbackground='#e53935')
            else:
                self.canvas.configure(highlightthickness=0)
        else:
            pw = max(self.canvas.winfo_width(),  400)
            ph = max(self.canvas.winfo_height(), 300)
            self.canvas.delete('all')
            self.canvas.create_text(pw // 2, ph // 2,
                                     text='하단 포트 번호를 선택하세요',
                                     fill=TXT_G, font=('맑은 고딕', 13))

        self.after(33, self._update_preview)

    # ── 치수 측정 패널 ────────────────────────────────────────────────────────

    def _build_dimension_panel(self, parent):
        outer = tk.Frame(parent, bg=CARD_BG,
                         highlightbackground='#2c5f9e', highlightthickness=1)
        outer.pack(fill='x', pady=(4, 0))

        # ── 헤더 ──────────────────────────────────────────────
        hdr = tk.Frame(outer, bg='#1a3a5c')
        hdr.pack(fill='x')
        tk.Label(hdr, text='  \U0001f4cf  이탈 거리 측정  (카메라 2점 클릭)',
                 font=('맑은 고딕', 9, 'bold'), bg='#1a3a5c', fg=TXT_W
                 ).pack(side='left', padx=8, pady=4)

        # ── STEP 1: 캘리브레이션 (접이식 — 한 번 설정하면 평소엔 접어둠) ──
        s1 = tk.Frame(outer, bg='#151008',
                      highlightbackground='#7a5c00', highlightthickness=1)
        s1.pack(fill='x', padx=4, pady=(4, 2))

        s1_hdr = tk.Frame(s1, bg='#151008', cursor='hand2')
        s1_hdr.pack(fill='x')
        self._cal_toggle_var = tk.StringVar(value='▼')
        tk.Label(s1_hdr, textvariable=self._cal_toggle_var,
                 font=('맑은 고딕', 9, 'bold'), bg='#151008', fg=ACC_YEL,
                 cursor='hand2').pack(side='left', padx=(6, 2), pady=(4, 2))
        tk.Label(s1_hdr, text='① 검사구역 캘리브레이션',
                 font=('맑은 고딕', 9, 'bold'), bg='#151008', fg=ACC_YEL,
                 cursor='hand2').pack(side='left', pady=(4, 2))

        s1_body = tk.Frame(s1, bg='#151008')
        s1_body.pack(fill='x')

        def _toggle_s1(_ev=None):
            if s1_body.winfo_ismapped():
                s1_body.pack_forget()
                self._cal_toggle_var.set('▶')
            else:
                s1_body.pack(fill='x')
                self._cal_toggle_var.set('▼')

        s1_hdr.bind('<Button-1>', _toggle_s1)
        for _w in s1_hdr.winfo_children():
            _w.bind('<Button-1>', _toggle_s1)

        # 검사구역 직경 입력 행 — 좁은 창에서도 안 잘리게 입력칸과 버튼을 분리
        r_cal = tk.Frame(s1_body, bg='#151008')
        r_cal.pack(fill='x', padx=6, pady=(0, 2))
        tk.Label(r_cal, text='검사구역 (mm):',
                 font=('맑은 고딕', 9), bg='#151008', fg=TXT_G).pack(side='left')
        self._cal_ref_var = tk.StringVar(value='2.00')
        self._cal_entry = tk.Entry(r_cal, textvariable=self._cal_ref_var, width=5,
                 font=('맑은 고딕', 9, 'bold'), bg='#21262d', fg=TXT_W,
                 relief='flat', insertbackground=TXT_W)
        self._cal_entry.pack(side='left', padx=2)
        tk.Label(r_cal, text='mm', font=('맑은 고딕', 9),
                 bg='#151008', fg=TXT_G).pack(side='left')

        self.btn_magnifier = tk.Button(
            r_cal, text='돋보기 OFF',
            font=('맑은 고딕', 9), bg='#21262d', fg='#8b949e',
            relief='flat', cursor='hand2', padx=5, pady=2,
            command=self._toggle_magnifier)
        self.btn_magnifier.pack(side='right', padx=2)

        # 측정 시작 버튼 — 자기 줄을 다 차지해서 좁은 창에서도 안 잘림
        r_cal_btn = tk.Frame(s1_body, bg='#151008')
        r_cal_btn.pack(fill='x', padx=6, pady=(2, 2))
        self.btn_cal = tk.Button(r_cal_btn, text='검사구역 직경 측정 시작',
                                  font=('맑은 고딕', 9, 'bold'),
                                  bg='#1a3a6b', fg='#fff',
                                  relief='flat', cursor='hand2', padx=6, pady=2)
        self.btn_cal.configure(command=self._start_calibration)
        self.btn_cal.pack(fill='x')

        # 측정 결과 표시 행
        r_result = tk.Frame(s1_body, bg='#151008')
        r_result.pack(fill='x', padx=6, pady=(0, 2))
        tk.Label(r_result, text='측정값:',
                 font=('맑은 고딕', 8), bg='#151008', fg=TXT_G).pack(side='left')
        self._measured_px_var = tk.StringVar(value='— px')
        tk.Label(r_result, textvariable=self._measured_px_var,
                 font=('맑은 고딕', 9, 'bold'), bg='#151008', fg=ACC_YEL
                 ).pack(side='left', padx=4)
        self._cal_result_var = tk.StringVar(value='')
        tk.Label(r_result, textvariable=self._cal_result_var,
                 font=('맑은 고딕', 9, 'bold'), bg='#151008', fg=ACC_GRN
                 ).pack(side='right', padx=4)

        self._cal_status_var = tk.StringVar(
            value='① 루페로 검사구역 실측(mm) 입력\n② 측정 시작 클릭\n③ P1 → P2 순서로 클릭')
        tk.Label(s1_body, textvariable=self._cal_status_var,
                 font=('맑은 고딕', 9, 'bold'), bg='#151008', fg='#ffd54a',
                 wraplength=220, justify='left').pack(anchor='w', padx=6, pady=(0, 8))

        # 고정 배율 단일 캘리브레이션 — 저장은 _on_canvas_click 참고
        self._cal_s1_body   = s1_body
        self._cal_toggle_fn = _toggle_s1
        if 'by_zoom' not in self._cal_store and '__housing__' in self._cal_store:
            # 구버전 단일 캘리브레이션 → __default__ 항목으로 이전
            self._cal_store['by_zoom'] = {'__default__': self._cal_store['__housing__']}
        _saved = self._cal_store.get('by_zoom', {}).get('__default__')
        if _saved:
            self._apply_saved_cal(_saved)
            # 이미 캘리브레이션이 저장돼 있으면 평소엔 접어서 화면 밀도를 낮춤
            _toggle_s1()

        # ── STEP 2: 이탈 측정 ─────────────────────────────────
        s2 = tk.Frame(outer, bg='#0a1a0a',
                      highlightbackground='#2e8b57', highlightthickness=1)
        s2.pack(fill='x', padx=4, pady=(2, 4))

        tk.Label(s2, text='  ② 이탈 거리 측정  — 두 점 클릭',
                 font=('맑은 고딕', 9, 'bold'), bg='#0a1a0a', fg=ACC_GRN
                 ).pack(anchor='w', padx=4, pady=(4, 2))

        self.btn_measure = tk.Button(
            s2, text='\U0001f4cf  측정 모드 OFF',
            font=('맑은 고딕', 9, 'bold'),
            bg='#21262d', fg=TXT_G,
            relief='flat', cursor='hand2', pady=4,
            command=self._toggle_measure_mode)
        self.btn_measure.pack(fill='x', padx=6, pady=(0, 2))

        # 이탈 거리 결과 (크게)
        res_box = tk.Frame(s2, bg='#0d2214',
                           highlightbackground='#2e8b57', highlightthickness=1)
        res_box.pack(fill='x', padx=6, pady=2)
        tk.Label(res_box, text='이탈 거리',
                 font=('맑은 고딕', 9), bg='#0d2214', fg=TXT_G).pack(pady=(4, 0))
        self._meas_result_var = tk.StringVar(value='— mm')
        tk.Label(res_box, textvariable=self._meas_result_var,
                 font=('맑은 고딕', 20, 'bold'), bg='#0d2214', fg=ACC_GRN
                 ).pack(pady=(0, 4))

        tk.Button(s2, text='측정 초기화', font=('맑은 고딕', 9),
                  bg='#2a2a2a', fg=TXT_W, relief='flat', cursor='hand2', pady=2,
                  command=self._reset_meas).pack(fill='x', padx=6, pady=(0, 6))

        # ── STEP 3: 제품 컷팅 이미지 캡처 (측정 아님 — 단순 기록용 캡처) ──────
        # 이 패널의 실제 색상은 아래 self._apply_cut_capture_theme() 호출이
        # 라이트/다크 테마에 맞춰 직접 칠한다 (CUT_CAPTURE_COLORS 참고).
        # 여기서 지정하는 색은 위젯 생성 시 잠깐 보이는 기본값일 뿐이므로
        # 의미 없음 — 실제 색은 아래 _apply_cut_capture_theme()이 즉시 덮어씀.
        s3 = tk.Frame(outer, highlightthickness=1)
        s3.pack(fill='x', padx=4, pady=(2, 4))
        self._s3_panel = s3

        self._cut_title_lbl = tk.Label(
            s3, text='  ③ 제품 컷팅 이미지 캡처', font=('맑은 고딕', 9, 'bold'))
        self._cut_title_lbl.pack(anchor='w', padx=4, pady=(4, 2))
        self._cut_subtitle_lbl = tk.Label(
            s3, text='  (위 검사 정보란과 별개 — 이 캡처 전용)', font=('맑은 고딕', 7))
        self._cut_subtitle_lbl.pack(anchor='w', padx=4, pady=(0, 2))

        cs = ttk.Style()
        cs.theme_use('clam')

        self._cut_field_rows   = []
        self._cut_field_labels = []
        for _lbl, _var_attr, _recent_attr in (
                ('제품 LOT', '_cut_lot_var', '_cut_lot_recent'),
                ('일련번호', '_cut_serial_var', '_cut_serial_recent'),
                ('칼날번호', '_cut_blade_var', '_cut_blade_recent')):
            _row = tk.Frame(s3)
            _row.pack(fill='x', padx=6, pady=(0, 2))
            self._cut_field_rows.append(_row)
            _flbl = tk.Label(_row, text=f'{_lbl}:', font=('맑은 고딕', 9, 'bold'),
                              width=7, anchor='w')
            _flbl.pack(side='left')
            self._cut_field_labels.append(_flbl)
            _cb = ttk.Combobox(
                _row, textvariable=getattr(self, _var_attr),
                values=getattr(self, _recent_attr),
                font=('맑은 고딕', 10, 'bold'), style='CutCapture.TCombobox')
            _cb.pack(side='left', fill='x', expand=True, padx=2)
            setattr(self, _var_attr + '_cb', _cb)

        self._cut_side_row = tk.Frame(s3)
        self._cut_side_row.pack(fill='x', padx=6, pady=(0, 4))
        self._btn_cut_r = tk.Button(
            self._cut_side_row, text='➡ 오른쪽 (R)', font=('맑은 고딕', 9, 'bold'),
            bg='#21262d', fg=ACC_RED, relief='flat', cursor='hand2', pady=4,
            command=lambda: self._select_cutting_side('R'))
        self._btn_cut_r.pack(side='left', fill='x', expand=True, padx=(0, 2))
        self._btn_cut_l = tk.Button(
            self._cut_side_row, text='⬅ 왼쪽 (L)', font=('맑은 고딕', 9, 'bold'),
            bg='#21262d', fg=ACC_BLU, relief='flat', cursor='hand2', pady=4,
            command=lambda: self._select_cutting_side('L'))
        self._btn_cut_l.pack(side='left', fill='x', expand=True, padx=(2, 0))

        self._cut_count_hdr_lbl = tk.Label(
            s3, text='컷팅 횟수', font=('맑은 고딕', 9, 'bold'))
        self._cut_count_hdr_lbl.pack(anchor='w', padx=6)
        self._cut_grid = tk.Frame(s3)
        self._cut_grid.pack(fill='x', padx=6, pady=(2, 2))
        self._cut_count_btns = {}
        for i, c in enumerate((1, 500, 1000, 1500, 2000, 2500, 5000)):
            b = tk.Button(
                self._cut_grid, text=f'{c}회', font=('맑은 고딕', 9, 'bold'),
                relief='flat', cursor='hand2', pady=4,
                command=lambda c=c: self._select_cutting_count(c))
            b.grid(row=i // 4, column=i % 4, sticky='nsew', padx=1, pady=1)
            self._cut_count_btns[c] = b
        for col in range(4):
            self._cut_grid.columnconfigure(col, weight=1)

        self._cut_custom_row = tk.Frame(s3)
        self._cut_custom_row.pack(fill='x', padx=6, pady=(2, 4))
        self._cut_custom_lbl = tk.Label(
            self._cut_custom_row, text='직접입력:', font=('맑은 고딕', 8, 'bold'))
        self._cut_custom_lbl.pack(side='left')
        self._cut_count_entry_var = tk.StringVar(value='')
        self._cut_count_entry = tk.Entry(
            self._cut_custom_row, textvariable=self._cut_count_entry_var, width=8,
            font=('맑은 고딕', 10, 'bold'), relief='flat', insertwidth=2)
        self._cut_count_entry.pack(side='left', padx=4)
        self._cut_unit_lbl = tk.Label(
            self._cut_custom_row, text='회', font=('맑은 고딕', 8, 'bold'))
        self._cut_unit_lbl.pack(side='left')
        self._cut_apply_btn = tk.Button(
            self._cut_custom_row, text='적용', font=('맑은 고딕', 8),
            relief='flat', cursor='hand2', padx=8,
            command=self._apply_custom_cutting_count)
        self._cut_apply_btn.pack(side='left', padx=6)
        self._cut_count_entry.bind('<Return>', lambda e: self._apply_custom_cutting_count())

        self._cutting_status_var = tk.StringVar(value='측면과 횟수를 선택하세요')
        self._cut_status_lbl = tk.Label(
            s3, textvariable=self._cutting_status_var, font=('맑은 고딕', 8, 'bold'),
            wraplength=220, justify='left')
        self._cut_status_lbl.pack(anchor='w', padx=6, pady=(0, 4))

        self._cut_capture_btn = tk.Button(
            s3, text='📸  제품 컷팅 이미지 캡처', font=('맑은 고딕', 9, 'bold'),
            relief='flat', cursor='hand2', pady=5,
            command=self._capture_cutting_image)
        self._cut_capture_btn.pack(fill='x', padx=6, pady=(0, 6))

        self._apply_cut_capture_theme()

    def _select_cutting_side(self, side):
        self._cutting_side = side
        for s, btn, clr in (('R', self._btn_cut_r, ACC_RED),
                             ('L', self._btn_cut_l, ACC_BLU)):
            if s == side:
                btn.configure(bg=clr, fg='#fff')
            else:
                btn.configure(bg='#21262d', fg=clr)
        self._update_cutting_status()

    def _select_cutting_count(self, count):
        self._cutting_count = count
        self._cut_count_entry_var.set('')
        self._apply_cut_capture_theme()
        self._update_cutting_status()

    def _apply_custom_cutting_count(self):
        raw = self._cut_count_entry_var.get().strip()
        try:
            val = int(raw)
            if val <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning('입력 오류', '컷팅 횟수는 양의 정수로 입력하세요.')
            return
        self._cutting_count = val
        self._apply_cut_capture_theme()
        self._update_cutting_status()

    def _update_cutting_status(self):
        side_txt = {'R': '오른쪽(R)', 'L': '왼쪽(L)', None: '미선택'}[self._cutting_side]
        cnt_txt = f'{self._cutting_count}회' if self._cutting_count else '미선택'
        self._cutting_status_var.set(f'현재 선택: {side_txt} · {cnt_txt}')

    def _push_cut_recent(self, list_attr, value):
        """컷팅 캡처 전용 LOT/일련번호/칼날번호 드롭다운에 최근 입력값 기록."""
        lst = getattr(self, list_attr)
        if value in lst:
            lst.remove(value)
        lst.insert(0, value)
        del lst[10:]
        cb = getattr(self, list_attr.replace('_recent', '_var') + '_cb', None)
        if cb is not None:
            cb['values'] = lst
        self._save_config()

    def _capture_cutting_image(self):
        """검사/측정용이 아닌 단순 기록용 캡처 — 컷팅 블레이드 좌/우 절단면을
        컷팅 횟수 구간별로 남겨서 마모 등을 추적하기 위한 용도.
        위쪽 검사정보(LOT/일련번호)와 헷갈리지 않게 이 캡처 전용 입력란을 씀."""
        if self._cutting_side is None:
            messagebox.showwarning('선택 필요', '먼저 오른쪽/왼쪽 컷팅면을 선택하세요!')
            return
        if not self._cutting_count:
            messagebox.showwarning('선택 필요', '먼저 컷팅 횟수를 선택하세요!')
            return

        with self._frame_lock:
            frame = self._frame.copy() if self._frame is not None else None
        if frame is None:
            messagebox.showwarning('캡처 실패', '카메라 화면이 없습니다.')
            return

        lot    = self._cut_lot_var.get().strip()
        serial = self._cut_serial_var.get().strip()
        blade  = self._cut_blade_var.get().strip()
        if not lot:
            messagebox.showwarning('입력 필요', '제품 LOT 번호를 먼저 입력하세요!')
            return
        if not serial:
            messagebox.showwarning('입력 필요', '일련번호를 먼저 입력하세요!')
            return

        lot_folder  = _safe_filename(lot, 'NO_LOT')
        safe_serial = _safe_filename(serial, 'NO_SN')

        # 동일 일련번호가 '다른 컷팅 횟수 회차'로 이미 촬영된 적 있는지 검사 —
        # 작업자가 제품 바뀐 걸 깜빡하고 일련번호를 안 바꿨을 가능성 경고.
        # 같은 회차의 좌/우(L/R) 캡처는 한 제품을 양면 촬영하는 정상 흐름이므로 제외.
        _fname_re = re.compile(
            r'^' + re.escape(safe_serial) + r' (\d+) [LR](?:_\d+)?\.jpg$')
        existing_matches = []
        if os.path.isdir(CUTTING_CAPTURE_DIR):
            for _root, _dirs, _files in os.walk(CUTTING_CAPTURE_DIR):
                for _f in _files:
                    _m = _fname_re.match(_f)
                    if _m and int(_m.group(1)) != self._cutting_count:
                        existing_matches.append(
                            os.path.relpath(os.path.join(_root, _f), CUTTING_CAPTURE_DIR))
        if existing_matches:
            _preview = '\n'.join(existing_matches[:10])
            if len(existing_matches) > 10:
                _preview += f'\n... 외 {len(existing_matches) - 10}건'
            if not messagebox.askyesno(
                    '중복 일련번호',
                    f'"{serial}" 은(는) 이미 촬영된 일련번호입니다.\n\n'
                    f'기존 촬영 파일:\n{_preview}\n\n'
                    '그래도 계속 촬영하시겠습니까?'):
                return

        now = datetime.datetime.now()
        ts  = now.strftime('%Y%m%d_%H%M%S')
        save_dir    = os.path.join(CUTTING_CAPTURE_DIR, lot_folder)

        side  = self._cutting_side
        count = self._cutting_count
        base_name = f'{safe_serial} {count} {side}.jpg'
        fname = os.path.join(save_dir, base_name)
        if os.path.exists(fname):
            base_name = f'{safe_serial} {count} {side}_{ts}.jpg'
            fname     = os.path.join(save_dir, base_name)

        img  = Image.fromarray(frame)
        draw = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype('C:\\Windows\\Fonts\\malgunbd.ttf', 22)
        except Exception:
            font = ImageFont.load_default()

        side_label = '오른쪽 컷팅면 (R)' if side == 'R' else '왼쪽 컷팅면 (L)'
        blade_txt = f'   칼날: {blade}' if blade else ''
        overlay = (f'{side_label}   LOT: {lot}   SN: {safe_serial}   '
                   f'컷팅 {count}회{blade_txt}   '
                   f'[{now.strftime("%Y-%m-%d %H:%M:%S")}]')
        draw.rectangle([0, img.height - 40, img.width, img.height], fill=(0, 0, 0))
        draw.text((10, img.height - 33), overlay, fill=(255, 213, 74), font=font)

        try:
            os.makedirs(save_dir, exist_ok=True)
            img.save(fname, quality=95)
        except Exception as ex:
            messagebox.showerror(
                '캡처 저장 실패',
                'LOT/일련번호에 사용할 수 없는 문자가 포함되었거나 '
                '디스크 오류로 컷팅 이미지를 저장하지 못했습니다.\n\n'
                f'상세 정보: {ex}')
            return

        self._push_cut_recent('_cut_lot_recent', lot)
        self._push_cut_recent('_cut_serial_recent', serial)
        if blade:
            self._push_cut_recent('_cut_blade_recent', blade)

        self.lbl_flash.configure(
            text=f'컷팅 이미지 저장: {lot_folder}/{os.path.basename(fname)}',
            fg='#4a9fd4')
        self.after(3000, lambda: self.lbl_flash.configure(text=''))

    def _reset_meas(self):
        self._measure_pts  = []
        self._deviation_val = 0.0
        if self._meas_result_var:
            self._meas_result_var.set('— mm')

    def _start_calibration(self):
        try:
            ref = float(self._cal_ref_var.get())
            if ref <= 0:
                raise ValueError
        except (ValueError, TypeError):
            messagebox.showwarning('입력 오류', '기준 거리를 양수로 입력하세요 (mm)')
            return
        # 측정 모드 해제
        self._measure_mode = False
        self._measure_pts  = []
        if hasattr(self, 'btn_measure'):
            self.btn_measure.configure(bg='#21262d', fg=TXT_G,
                                       text='\U0001f4cf  측정 모드 OFF')
        # 캘리브레이션 시작
        self._cal_mode = True
        self._cal_pts  = []
        self.btn_cal.configure(bg=ACC_YEL, fg='#111')
        self._cal_status_var.set(
            f'검사구역 한쪽 끝 P1을 클릭하세요  ({ref:.2f} mm 기준)')

    def _toggle_measure_mode(self):
        if self._measure_mode:
            self._measure_mode = False
            self.btn_measure.configure(bg='#21262d', fg=TXT_G,
                                       text='\U0001f4cf  측정 모드 OFF')
        else:
            if self._px_per_mm <= 0:
                messagebox.showwarning(
                    '캘리브레이션 필요',
                    '먼저 캘리브레이션을 완료해야 측정할 수 있습니다.')
                return
            self._cal_mode = False
            self._cal_pts  = []
            self._measure_mode = True
            self._measure_pts  = []
            self.btn_measure.configure(bg='#1a5c2e', fg='#fff',
                                       text='\U0001f4cf  측정 모드 ON  ― 두 점 클릭')
            if self._meas_result_var:
                self._meas_result_var.set('첫 번째 점을 클릭하세요...')

    # ── 캔버스 드래그 이벤트 ──────────────────────────────────────────────────

    def _line_hit(self, ex, ey, cw, ch):
        """클릭 위치가 캘리브 라인 근처인지 확인. 'p1'|'p2'|'line'|None 반환."""
        p1x = int(self._cal_line_p1[0] * cw)
        p1y = int(self._cal_line_p1[1] * ch)
        p2x = int(self._cal_line_p2[0] * cw)
        p2y = int(self._cal_line_p2[1] * ch)
        R = 16  # 끝점 히트 반경
        if _math.sqrt((ex-p1x)**2 + (ey-p1y)**2) <= R:
            return 'p1'
        if _math.sqrt((ex-p2x)**2 + (ey-p2y)**2) <= R:
            return 'p2'
        # 라인 본체 히트 (중간 70%)
        lx, ly = p2x-p1x, p2y-p1y
        ll = _math.sqrt(lx*lx + ly*ly)
        if ll > 0:
            t = ((ex-p1x)*lx + (ey-p1y)*ly) / (ll*ll)
            if 0.15 <= t <= 0.85:
                d = abs((ey-p1y)*lx - (ex-p1x)*ly) / ll
                if d <= 10:
                    return 'line'
        return None

    def _nudge_line(self, dx_px, dy_px):
        if self._cal_mode or self._measure_mode:
            return
        cw = self.canvas.winfo_width()
        ch = self.canvas.winfo_height()
        if cw <= 0 or ch <= 0:
            return
        ddx, ddy = dx_px / cw, dy_px / ch
        target = self._line_active or 'p2'
        if target in ('p1', 'line'):
            self._cal_line_p1 = (max(0.0, min(1.0, self._cal_line_p1[0] + ddx)),
                                  max(0.0, min(1.0, self._cal_line_p1[1] + ddy)))
        if target in ('p2', 'line'):
            self._cal_line_p2 = (max(0.0, min(1.0, self._cal_line_p2[0] + ddx)),
                                  max(0.0, min(1.0, self._cal_line_p2[1] + ddy)))

    def _on_canvas_press(self, event):
        cw = self.canvas.winfo_width()
        ch = self.canvas.winfo_height()
        if cw <= 0 or ch <= 0:
            return
        self.canvas.focus_set()
        hit = self._line_hit(event.x, event.y, cw, ch)
        if hit and not self._cal_mode and not self._measure_mode:
            self._line_dragging = hit
            self._line_active   = hit
            mx, my = event.x / cw, event.y / ch
            if hit == 'line':
                mid_x = (self._cal_line_p1[0] + self._cal_line_p2[0]) / 2
                mid_y = (self._cal_line_p1[1] + self._cal_line_p2[1]) / 2
                self._line_drag_ref = (mx, my, mid_x, mid_y)
            elif hit == 'p1':
                self._line_drag_ref = (mx, my,
                                        self._cal_line_p1[0], self._cal_line_p1[1])
            else:
                self._line_drag_ref = (mx, my,
                                        self._cal_line_p2[0], self._cal_line_p2[1])
        else:
            self._line_dragging = None
            self._on_canvas_click(event)

    def _on_canvas_drag(self, event):
        cw = self.canvas.winfo_width()
        ch = self.canvas.winfo_height()
        if cw <= 0 or ch <= 0 or not self._line_dragging:
            return
        mx, my = event.x / cw, event.y / ch
        ox, oy, bx, by = self._line_drag_ref
        dx, dy = mx - ox, my - oy
        if self._line_dragging == 'line':
            vx = (self._cal_line_p2[0] - self._cal_line_p1[0]) / 2
            vy = (self._cal_line_p2[1] - self._cal_line_p1[1]) / 2
            cx_ = max(0.0, min(1.0, bx + dx))
            cy_ = max(0.0, min(1.0, by + dy))
            self._cal_line_p1 = (cx_ - vx, cy_ - vy)
            self._cal_line_p2 = (cx_ + vx, cy_ + vy)
        elif self._line_dragging == 'p1':
            self._cal_line_p1 = (max(0.0, min(1.0, bx + dx)),
                                   max(0.0, min(1.0, by + dy)))
        else:
            self._cal_line_p2 = (max(0.0, min(1.0, bx + dx)),
                                   max(0.0, min(1.0, by + dy)))

    def _on_canvas_release(self, event):
        self._line_dragging = None

    def _on_canvas_hover(self, event):
        cw = self.canvas.winfo_width()
        ch = self.canvas.winfo_height()
        if cw <= 0 or ch <= 0:
            return
        # 돋보기용 마우스 위치 갱신
        self._mouse_nx = max(0.0, min(1.0, event.x / cw))
        self._mouse_ny = max(0.0, min(1.0, event.y / ch))
        if (not self._cal_mode and not self._measure_mode and
                self._line_hit(event.x, event.y, cw, ch)):
            self.canvas.configure(cursor='fleur')
        else:
            self.canvas.configure(cursor='crosshair')

    def _on_canvas_click(self, event):
        cw = self.canvas.winfo_width()
        ch = self.canvas.winfo_height()
        if cw <= 0 or ch <= 0:
            return
        nx, ny = event.x / cw, event.y / ch

        if self._cal_mode:
            # 드래그 라인 끝점 스냅
            _lp1x = int(self._cal_line_p1[0] * cw)
            _lp1y = int(self._cal_line_p1[1] * ch)
            _lp2x = int(self._cal_line_p2[0] * cw)
            _lp2y = int(self._cal_line_p2[1] * ch)
            d0 = _math.sqrt((event.x - _lp1x)**2 + (event.y - _lp1y)**2)
            d1 = _math.sqrt((event.x - _lp2x)**2 + (event.y - _lp2y)**2)
            if d0 <= SCALE_SNAP_R:
                nx, ny = self._cal_line_p1
            elif d1 <= SCALE_SNAP_R:
                nx, ny = self._cal_line_p2

            self._cal_pts.append((nx, ny))
            n = len(self._cal_pts)
            # 돋보기는 마우스 따라다니는 실시간 오버레이 — 클릭 시 별도 동작 없음
            if n == 1:
                try:
                    ref = float(self._cal_ref_var.get())
                except (ValueError, TypeError):
                    ref = 40.0
                self._cal_status_var.set(
                    f'반대쪽 끝 P2를 클릭하세요  ({ref:.2f} mm 기준)')
            elif n >= 2:
                p1, p2 = self._cal_pts[0], self._cal_pts[1]
                dx = (p2[0] - p1[0]) * cw
                dy = (p2[1] - p1[1]) * ch
                px_dist = _math.sqrt(dx*dx + dy*dy)
                try:
                    ref_mm = float(self._cal_ref_var.get())
                except (ValueError, TypeError):
                    ref_mm = 1.0
                # ── P1-P2 거리로 px/mm 확정 ──────────────────────────
                self._px_per_mm          = px_dist / ref_mm if ref_mm > 0 else 0
                self._cal_mode           = False
                self._cal_pts            = []
                self._px_per_mm_baseline = self._px_per_mm

                self._cal_store.setdefault('by_zoom', {})['__default__'] = {
                    'px_per_mm': round(self._px_per_mm, 4),
                    'ref_mm':    round(ref_mm,           2),
                    'port_px':   round(px_dist,          2),
                }
                self._save_cal_file()
                self.btn_cal.configure(bg='#1a5c2e', fg='#fff')
                if hasattr(self, '_measured_px_var'):
                    self._measured_px_var.set(f'{px_dist:.1f} px')
                if hasattr(self, '_cal_result_var'):
                    self._cal_result_var.set(f'거리: {ref_mm:.1f}mm  ✓')
                self._cal_status_var.set(
                    f'✓ 완료  {ref_mm:.2f}mm = {px_dist:.1f}px'
                    f'  →  1mm = {self._px_per_mm:.2f}px  ✓')
        elif self._measure_mode:
            # 측정 완료 상태(2점 있음)에서 새 클릭 → 초기화 후 재시작
            if len(self._measure_pts) >= 2:
                self._measure_pts = []
            self._measure_pts.append((nx, ny))
            n = len(self._measure_pts)
            if n == 1:
                if self._meas_result_var:
                    self._meas_result_var.set('두 번째 점을 클릭하세요...')
            elif n == 2:
                p1, p2 = self._measure_pts
                dx = (p2[0] - p1[0]) * cw
                dy = (p2[1] - p1[1]) * ch
                px_dist = _math.sqrt(dx*dx + dy*dy)
                mm = round(px_dist / self._px_per_mm, 2) if self._px_per_mm > 0 else 0
                self._deviation_val = mm
                if self._meas_result_var:
                    self._meas_result_var.set(f'{mm:.2f} mm')

    # ── 줌 캘리브레이션 파일 ─────────────────────────────────────────────────

    def _load_cal_file(self):
        try:
            if os.path.exists(_CAL_FILE):
                with open(_CAL_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def _apply_saved_cal(self, saved):
        self._px_per_mm          = float(saved.get('px_per_mm', 0))
        self._px_per_mm_baseline = self._px_per_mm
        _ref_mm  = float(saved.get('ref_mm',  2.0))
        _port_px = float(saved.get('port_px', 0))
        self._cal_ref_var.set(f'{_ref_mm:.2f}')
        self._measured_px_var.set(f'{_port_px:.1f} px' if _port_px > 0 else '—')
        self._cal_result_var.set(f'거리: {_ref_mm:.1f}mm  ✓')
        self._cal_status_var.set(
            f'✓ 저장된 캘리브 로드  {_ref_mm:.1f}mm → 1mm={self._px_per_mm:.2f}px')
        self.btn_cal.configure(bg='#1a5c2e', fg='#fff')

    def _save_cal_file(self):
        # 임시 파일에 먼저 쓰고 os.replace로 교체 — 저장 도중 강제종료/정전이
        # 나도 원본 캘리브레이션 파일이 손상(부분 기록)되지 않도록 함
        tmp_path = _CAL_FILE + '.tmp'
        try:
            with open(tmp_path, 'w', encoding='utf-8') as f:
                json.dump(self._cal_store, f, ensure_ascii=False, indent=2)
                f.flush()
                os.fsync(f.fileno())
            os.replace(tmp_path, _CAL_FILE)
        except Exception as ex:
            messagebox.showerror(
                '캘리브레이션 저장 실패',
                '캘리브레이션 값을 저장하지 못했습니다. 다음 실행 시 다시 설정해야 할 수 있습니다.\n\n'
                f'파일 위치: {_CAL_FILE}\n상세 정보: {ex}')
            try:
                os.remove(tmp_path)
            except OSError:
                pass

    # ── 종료 ──────────────────────────────────────────────────────────────────

    def _on_close(self):
        cnt = self.count.get()
        if cnt > 0:
            if not messagebox.askyesno('종료 확인',
                                        f'카운트 {cnt}건이 저장되지 않았습니다.\n종료하시겠습니까?'):
                return
        self.running = False
        self._led_on = False   # keep-alive 스레드가 다음 루프에서 스스로 멈추도록
        # 카메라 스레드(_cam_loop)가 cam.read()를 쉴 새 없이 반복 호출하는 구조라,
        # 종료 순간에 OpenCV/DirectShow 네이티브 호출 도중일 확률이 높다. 그 상태에서
        # 바로 강제종료하면 Windows가 그 커널 I/O가 끝날 때까지 프로세스 정리를
        # 미루는 것으로 추정됨(종료 시 프로세스가 남는 문제의 원인 가설).
        # running=False 신호를 준 뒤 카메라 스레드가 스스로 루프를 빠져나와
        # cam.release()할 시간을 짧게 준 다음 종료한다.
        time.sleep(0.4)
        # DNX64 SDK가 내부적으로 자체 폴링/모니터링 스레드를 갖고 있을 가능성이
        # 있어 보여서(StartMonitoring/StopMonitoring API 존재), 명시적으로
        # 정지시켜본다.
        try:
            if _DNX64._dll is not None:
                _DNX64._dll.StopMonitoring()
        except Exception:
            pass
        # 우리 프로세스 내부(스레드 watchdog, os._exit 등) 방식으로는 잡히지 않는
        # 원인 불명의 좀비화가 실제로 관찰됨(작업관리자로만 종료되던 현상).
        # 우리 프로세스 상태와 완전히 무관하게, 외부의 독립된 프로세스가 잠시
        # 기다렸다가 우리 PID를 강제로 taskkill 하도록 해서 확실하게 죽인다.
        try:
            subprocess.Popen(
                ['cmd', '/c', f'timeout /t 2 /nobreak >nul & taskkill /F /PID {os.getpid()}'],
                creationflags=subprocess.CREATE_NO_WINDOW)
        except Exception:
            pass
        # 메인 스레드가 destroy() 도중 멈추면(카메라/DirectShow 핸들 문제 등) 아래
        # _force_exit 자체가 실행 안 될 수 있으므로, 별도 스레드에서도 데드라인을
        # 걸어 무조건 프로세스를 죽이는 안전장치를 걸어둔다.
        def _watchdog():
            time.sleep(1.5)
            os._exit(0)
        threading.Thread(target=_watchdog, daemon=True).start()
        self._force_exit()

    def _force_exit(self):
        """DinoCapture와 카메라를 동시에 잡고 있으면 DirectShow/COM 핸들이 풀리지
        않아 destroy() 후에도 프로세스가 남는 경우가 있다(작업관리자로만 종료 가능).
        저장이 끝난 이 시점 이후로는 정리할 상태가 없으므로 즉시 강제 종료한다."""
        try:
            self.destroy()
        except Exception:
            pass
        os._exit(0)


if __name__ == '__main__':
    app = HoleCounter()
    app.update_idletasks()
    sw = app.winfo_screenwidth()
    sh = app.winfo_screenheight()
    win_w = min(1400, sw - 40)
    win_h = min(800,  sh - 60)
    x = (sw - win_w) // 2
    y = (sh - win_h) // 2
    app.geometry(f'{win_w}x{win_h}+{x}+{y}')
    app.mainloop()
